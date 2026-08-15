"""
Menina dos Raios Ltda â€” Backend v15
Multi-usuÃ¡rio Â· SessÃµes Â· HistÃ³rico por conta Â· Placa Â· Hora Â· PreÃ§o por data
"""
import os, sys, re, uuid, sqlite3, webbrowser, threading, io, json, hashlib, socket, hmac, logging, shutil, base64
from contextvars import ContextVar
from contextlib import contextmanager
from collections import defaultdict
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Optional, List, Dict, Any
import uvicorn, bcrypt
from functools import lru_cache
from fastapi import FastAPI, HTTPException, UploadFile, File, Header, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, StreamingResponse
try:
    from .app_notes_domain import _clean_app_note
    from .app_notes_service import app_note_catalog_from_rows, app_note_dict_from_row
    from .backup_admin import _copy_sqlite_consistent, _is_sqlite_file, _valid_backup_name, backup_db_sources_from_paths, backup_expected_databases_from_paths, backup_files_from_dir, backup_manifest_databases_from_zip, backup_path_for_filename, restore_zip_backup_with, safety_backup_before_restore_with
    from .company_config import company_db_path_for, company_key_from
    from .monteiro_permissions import payment_role_allowed
    from .monteiro_periods import _pal_period_where, _pay_period_map
    from .orcamentos import QuoteItemsLimitError, _quote_companies, _quote_company, quote_totals_from_items
    from .permissions_tabs import TAB_PERMISSION_ALIASES, _expand_tab_keys, permissions_configured_from_map, session_has_any_tab_from_map, tab_permissions_map_from_db
    from .security_auth import LOGIN_RATE_BLOCK_SECS, LOGIN_RATE_MAX_FAILS, LOGIN_RATE_WINDOW, _LOGIN_ATTEMPTS, _check_login_rate, _record_login, _time_mod
    from .security_request import _client_ip, _is_trusted_proxy_host
    from .schemas import AdminMessageIn, ClientIn, LoginIn, PriceUpdate, SaleIn, UserIn
    from .utils import _add_months, _calendar_event_dict, _normalize_client, _normalize_name, _safe_txt, _wa_failure_hint, _wa_log_response
except ImportError:
    from app_notes_domain import _clean_app_note
    from app_notes_service import app_note_catalog_from_rows, app_note_dict_from_row
    from backup_admin import _copy_sqlite_consistent, _is_sqlite_file, _valid_backup_name, backup_db_sources_from_paths, backup_expected_databases_from_paths, backup_files_from_dir, backup_manifest_databases_from_zip, backup_path_for_filename, restore_zip_backup_with, safety_backup_before_restore_with
    from company_config import company_db_path_for, company_key_from
    from monteiro_permissions import payment_role_allowed
    from monteiro_periods import _pal_period_where, _pay_period_map
    from orcamentos import QuoteItemsLimitError, _quote_companies, _quote_company, quote_totals_from_items
    from permissions_tabs import TAB_PERMISSION_ALIASES, _expand_tab_keys, permissions_configured_from_map, session_has_any_tab_from_map, tab_permissions_map_from_db
    from security_auth import LOGIN_RATE_BLOCK_SECS, LOGIN_RATE_MAX_FAILS, LOGIN_RATE_WINDOW, _LOGIN_ATTEMPTS, _check_login_rate, _record_login, _time_mod
    from security_request import _client_ip, _is_trusted_proxy_host
    from schemas import AdminMessageIn, ClientIn, LoginIn, PriceUpdate, SaleIn, UserIn
    from utils import _add_months, _calendar_event_dict, _normalize_client, _normalize_name, _safe_txt, _wa_failure_hint, _wa_log_response

LOG_LEVEL = os.environ.get("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
logger = logging.getLogger("menina")

def _debug_print(*args, **kwargs):
    """Mantem prints antigos sem vazar detalhes em producao."""
    if logger.isEnabledFor(logging.DEBUG):
        logger.debug(" ".join(str(a) for a in args))

print = _debug_print

# â”€â”€ Password hashing â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# MigraÃ§Ã£o transparente de SHA-256 (legacy) â†’ bcrypt.
# Hashes bcrypt comeÃ§am com "$2a$"/"$2b$"/"$2y$". SHA-256 = 64 chars hex.
def hash_password(plain: str) -> str:
    """Gera hash bcrypt com salt Ãºnico (cost 12 = ~200ms, padrÃ£o seguro)."""
    return bcrypt.hashpw(plain.encode("utf-8"), bcrypt.gensalt(rounds=12)).decode("ascii")

def verify_password(plain: str, stored: str) -> bool:
    """Aceita bcrypt OU SHA-256 legado (para usuÃ¡rios antigos)."""
    if not stored: return False
    if stored.startswith("$2"):
        try: return bcrypt.checkpw(plain.encode("utf-8"), stored.encode("ascii"))
        except Exception: return False
    # legacy SHA-256
    return hashlib.sha256(plain.encode("utf-8")).hexdigest() == stored

def needs_rehash(stored: str) -> bool:
    """True se hash Ã© SHA-256 antigo e precisa ser migrado para bcrypt."""
    return bool(stored) and not stored.startswith("$2")

# â”€â”€ PolÃ­tica de senha â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Bloqueia senhas Ã³bvias mesmo que atendam o tamanho mÃ­nimo. Lista enxuta com
# as senhas mais comuns globalmente + variantes ligadas Ã  empresa.
COMMON_PASSWORDS = {
    # NumÃ©ricas
    "12345678","123456789","1234567890","12345","123456","1234567",
    "11111111","00000000","12121212","87654321","123123123","11223344",
    # Top universais
    "password","password1","password12","password123","passw0rd",
    "qwerty","qwerty12","qwerty123","qwertyuiop","qwerty1234",
    "asdfgh","asdfghjk","asdfghjkl","zxcvbnm","1qaz2wsx","1q2w3e4r",
    "abc123","abc12345","abcd1234","aaaaaaaa","letmein",
    "iloveyou","welcome","welcome1","master","dragon","monkey","trustno1",
    "sunshine","princess","football","baseball","superman",
    # PortuguÃªs / brasileiras
    "senha","senha123","senha1234","minhasenha","senha2024","senha2025","senha2026",
    "brasil","brasil123","futebol","corinthians","flamengo",
    # Admin / sistema
    "admin","admin123","admin1234","administrator","administrator1","root","root123",
    "system","sistema","master123",
    # Variantes ligadas Ã  empresa
    "menina","raios","meninadosraios","menina123","raios123","monteiro","monteiro123",
    "bmmonteiro","mingau","boavista","boavista123","roraima","roraima123",
}

def validate_password(pw: str, username: str = "") -> None:
    """Valida senha contra polÃ­tica mÃ­nima. Levanta HTTPException(400) se falhar."""
    if not pw or len(pw) < 8:
        raise HTTPException(400, "Senha muito curta (mÃ­nimo 8 caracteres).")
    if len(pw) > 200:
        raise HTTPException(400, "Senha muito longa (mÃ¡ximo 200 caracteres).")
    lower = pw.lower()
    if lower in COMMON_PASSWORDS:
        raise HTTPException(400, "Esta senha Ã© muito comum. Escolha uma diferente.")
    if username and lower == username.lower():
        raise HTTPException(400, "A senha nÃ£o pode ser igual ao nome de usuÃ¡rio.")

BASE_DIR   = Path(__file__).parent
BACKUP_DIR = Path(os.environ.get("BACKUP_DIR", str(BASE_DIR.parent / "backups"))).resolve()
BACKUP_DIR.mkdir(parents=True, exist_ok=True)
LEGACY_BACKUP_DIR = BASE_DIR / "backups"
if LEGACY_BACKUP_DIR.exists() and LEGACY_BACKUP_DIR.resolve() != BACKUP_DIR:
    for old_backup in LEGACY_BACKUP_DIR.glob("bm_backup_*"):
        target = BACKUP_DIR / old_backup.name
        if old_backup.is_file() and not target.exists():
            try:
                shutil.copy2(old_backup, target)
            except Exception as e:
                print(f"migracao de backup falhou para {old_backup.name}: {e}")
MAX_BACKUPS = 30  # manter Ãºltimos 30 dias
STATIC_DIR = BASE_DIR / "static"
DB_PATH    = BASE_DIR / "bm_monteiro.db"
COMPANY_DBS = {
    "raios": DB_PATH,
    "estrada": BASE_DIR / "menina_estrada.db",
}
CURRENT_COMPANY: ContextVar[str] = ContextVar("CURRENT_COMPANY", default="raios")
APP_NOTES_DB_PATH = BASE_DIR / "app_notes.db"
APP_NOTES_TOKEN = os.environ.get("APP_NOTES_TOKEN", "6ab5af8ad03f2f9c4a2d6589838c7e0bee6a56886910b5df98b01557e0138fce")
APP_CALENDAR_TOKEN = os.environ.get("APP_CALENDAR_TOKEN", APP_NOTES_TOKEN)
MONTEIRO_NOTES_DIR = BASE_DIR / "monteiro_notas"
MONTEIRO_NOTES_DIR.mkdir(exist_ok=True)
SESSION_HOURS = 10
SESSION_IDLE_MINUTES = 20
MAX_EXCEL_UPLOAD = 10 * 1024 * 1024
MAX_NF_UPLOAD = 12 * 1024 * 1024
MAX_VERIFY_PDF_UPLOAD = 15 * 1024 * 1024
MAX_BACKUP_UPLOAD = 250 * 1024 * 1024
PDF_MAX_CONCURRENT = max(1, int(os.environ.get("PDF_MAX_CONCURRENT", "1")))
_PDF_SEMAPHORE = threading.BoundedSemaphore(PDF_MAX_CONCURRENT)

async def read_upload_limited(file: UploadFile, max_bytes: int, label: str) -> bytes:
    content = await file.read(max_bytes + 1)
    if not content:
        raise HTTPException(400, "Arquivo vazio.")
    if len(content) > max_bytes:
        mb = max_bytes // (1024 * 1024)
        raise HTTPException(400, f"{label} muito grande. Limite: {mb} MB.")
    return content

@contextmanager
def pdf_generation_slot():
    if not _PDF_SEMAPHORE.acquire(blocking=False):
        raise HTTPException(429, "Outro PDF ainda esta sendo gerado. Tente novamente em instantes.")
    try:
        yield
    finally:
        _PDF_SEMAPHORE.release()

# â”€â”€ PreÃ§os padrÃ£o â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
DEFAULT_PRICES = {
    "MAC_UNIT":    {"label":"Macaxeira (Unidade)",       "price":7.50,  "is_variable":False},
    "MAC_PCT":     {"label":"Macaxeira com Casca (KG)",  "price":7.50,  "is_variable":False},
    "MAC_VACUO":   {"label":"Macaxeira a VÃ¡cuo",         "price":7.50,  "is_variable":False},
    "MAC_CHIPS":   {"label":"Macaxeira Chips",           "price":10.00, "is_variable":False},
    "MASSA_MAC":   {"label":"Massa de Macaxeira",        "price":6.36,  "is_variable":False},
    "ALHO_250G":   {"label":"Alho 250g",                 "price":12.00, "is_variable":False},
    "ALHO_KG":     {"label":"Alho KG",                   "price":45.00, "is_variable":False},
    "PASTA_ALHO":  {"label":"Pasta de Alho",             "price":20.00, "is_variable":False},
    "PRE_COZIDA":  {"label":"PrÃ©-Cozida",                "price":10.00, "is_variable":False},
    "ABOBORA":     {"label":"AbÃ³bora (variÃ¡vel)",        "price":2.20,
                    "price_min":1.00,"price_max":4.00,   "is_variable":True},
    "ABOBORA_JAC": {"label":"AbÃ³bora JacarÃ© (variÃ¡vel)", "price":3.00,
                    "price_min":3.00,"price_max":4.00,   "is_variable":True},
}

PRODUCT_CANONICAL = {
    "MACAXEIRA A VÃCUO":"Macaxeira a VÃ¡cuo","MACAXEIRA CHIPS":"Macaxeira Chips",
    "MACAXEIRA COM CASCA":"Macaxeira com Casca (KG)","MACAXEIRA COM CASCA KG":"Macaxeira com Casca (KG)",
    "MACAXEIRA COM CASCA (KG)":"Macaxeira com Casca (KG)","MACAXEIRA C CASCA":"Macaxeira com Casca (KG)",
    "MACAXEIRA C/ CASCA":"Macaxeira com Casca (KG)","MACAXEIRA CASCA":"Macaxeira com Casca (KG)",
    "MASSA DE MACAXEIRA":"Massa de Macaxeira","ALHO 250G":"Alho 250g",
    "ALHO 250g":"Alho 250g","ALHO DESCASCADO":"Alho KG","ALHO KG":"Alho KG",
    "ALHO":"Alho 250g","PASTA DE ALHO":"Pasta de Alho",
    "PRE COZIDA":"PrÃ©-Cozida","PRÃ‰ COZIDA":"PrÃ©-Cozida",
    "MACAXEIRA A VÃCUO":"Macaxeira a VÃ¡cuo","MACAXEIRA VACUO":"Macaxeira a VÃ¡cuo","ABÃ“BORA JACARÃ‰":"AbÃ³bora JacarÃ©","ABOBORA JACARE":"AbÃ³bora JacarÃ©",
    "ABÃ“BORA":"AbÃ³bora JacarÃ©","ABOBORA":"AbÃ³bora JacarÃ©",
}
def norm_p(p, sale_type="NF"):
    """Normaliza produto usando contexto de tipo de venda para macaxeira genÃ©rica."""
    import re
    import unicodedata
    raw = str(p or "").strip()
    pu = raw.upper().strip()
    folded = unicodedata.normalize("NFD", pu).encode("ascii", "ignore").decode("ascii")
    folded = re.sub(r"[^A-Z0-9]+", " ", folded).strip()
    # Macaxeira sem especificaÃ§Ã£o: PR = com casca, NF/outros = a vÃ¡cuo
    if folded == "MACAXEIRA":
        return "Macaxeira com Casca (KG)" if sale_type == "PR" else "Macaxeira a VÃ¡cuo"
    if "MACAXEIRA" in folded:
        if "CASCA" in folded or "KG" in folded or "KILO" in folded or "QUILO" in folded or "PCT" in folded or "PACOTE" in folded:
            return "Macaxeira com Casca (KG)"
        if "VACUO" in folded and "SEM VACUO" not in folded:
            return "Macaxeira a VÃ¡cuo"
    if pu in PRODUCT_CANONICAL:
        return PRODUCT_CANONICAL[pu]
    for k, v in PRODUCT_CANONICAL.items():
        kf = unicodedata.normalize("NFD", k.upper()).encode("ascii", "ignore").decode("ascii")
        kf = re.sub(r"[^A-Z0-9]+", " ", kf).strip()
        if kf and kf in folded: return v
    return raw.title()

def parse_avaria_text(text: str, prices: Dict[str,float]) -> tuple:
    raw=str(text).strip()
    if raw in ['0','','nan','None']: return 0.0, []
    clean=re.sub(r'\([^)]*\)','',raw,flags=re.IGNORECASE).strip()
    if not clean or clean=='0': return 0.0, []
    parts=re.split(r'[,;/]|(?<=[A-Za-zÀ-ÿ\d])\s*-\s*(?=\d)',clean)
    total=0.0; items=[]
    for part in parts:
        p=part.strip().upper()
        if not p: continue
        m=re.match(r'^(\d+(?:[.,]\d+)?)\s*(.+)$',p)
        if not m: continue
        qty=float(m.group(1).replace(',','.')); desc=m.group(2).strip()
        if re.search(r'V[AÁÃ]CUO',desc): key='MAC_VACUO'; product='Macaxeira a VÃ¡cuo'
        elif re.search(r'CHIPS',desc) and re.search(r'MAC',desc): key='MAC_CHIPS'; product='Macaxeira Chips'
        elif re.search(r'MASSA.*MAC|MAC.*MASSA',desc): key='MASSA_MAC'; product='Massa de Macaxeira'
        elif re.search(r'PCT|PACOTE|PCK|KG.*MAC|MAC.*KG',desc) and re.search(r'MAC',desc): key='MAC_PCT'; product='Macaxeira com Casca (KG)'
        elif re.search(r'MAC|^M$|^MS$|MACAXEIRAS?',desc): key='MAC_UNIT'; product='Macaxeira (Unidade)'
        elif re.search(r'PASTA.*ALHO|ALHO.*PASTA',desc): key='PASTA_ALHO'; product='Pasta de Alho'
        elif re.search(r'DESCASC|KG.*ALHO|ALHO.*KG',desc): key='ALHO_KG'; product='Alho KG'
        elif re.search(r'ALHOS?',desc): key='ALHO_250G'; product='Alho 250g'
        elif re.search(r'PR[EÃ‰]\s*C|COZID',desc): key='PRE_COZIDA'; product='PrÃ©-Cozida'
        else: continue
        price=prices.get(key,0)
        value=round(qty*price,2)
        total=round(total+value,2)
        items.append({'product':product,'key':key,'qty':qty,'price':price,'value':value})
    return total, items

# â”€â”€ DB â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _company_key(value: str = "") -> str:
    return company_key_from(value, COMPANY_DBS)

def _company_db_path(company: str = "") -> Path:
    return company_db_path_for(company, COMPANY_DBS, DB_PATH)

def get_db(company: str = None):
    path=_company_db_path(company if company is not None else CURRENT_COMPANY.get())
    conn=sqlite3.connect(str(path),check_same_thread=False)
    conn.row_factory=sqlite3.Row
    # ConcorrÃªncia: WAL permite leitura simultÃ¢nea durante escrita (backup nÃ£o trava)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA busy_timeout=5000")
    conn.execute("PRAGMA cache_size=-8000")   # ~8MB de cache em RAM
    conn.execute("PRAGMA temp_store=MEMORY")  # ordenaÃ§Ãµes/temporÃ¡rios em RAM
    return conn

def get_control_db():
    """Banco compartilhado do grupo: usuarios, sessoes, permissoes e avisos."""
    return get_db("raios")

def init_db(company: str = None):
    conn=get_db(company)
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id           TEXT PRIMARY KEY,
        username     TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        full_name    TEXT,
        role         TEXT NOT NULL DEFAULT 'editor',
        active       INTEGER NOT NULL DEFAULT 1,
        created_at   TEXT DEFAULT (datetime('now')),
        created_by   TEXT
    );
    CREATE TABLE IF NOT EXISTS sessions (
        token      TEXT PRIMARY KEY,
        user_id    TEXT NOT NULL,
        username   TEXT NOT NULL,
        full_name  TEXT,
        role       TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now')),
        last_seen  TEXT DEFAULT (datetime('now')),
        expires_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS admin_messages (
        id          TEXT PRIMARY KEY,
        user_id     TEXT NOT NULL,
        title       TEXT NOT NULL,
        body        TEXT NOT NULL,
        created_at  TEXT DEFAULT (datetime('now')),
        expires_at  TEXT,
        seen_at     TEXT,
        removed_by_admin INTEGER NOT NULL DEFAULT 0,
        created_by  TEXT,
        FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_admin_messages_user ON admin_messages(user_id, seen_at, removed_by_admin, expires_at);
    CREATE TABLE IF NOT EXISTS sales (
        id              TEXT PRIMARY KEY,
        sale_type       TEXT NOT NULL CHECK(sale_type IN ('NF','PR','AVULSO','AVARIA')),
        sale_date       TEXT NOT NULL,
        sale_time       TEXT,
        client          TEXT,
        product         TEXT,
        nf_number       TEXT,
        quantity        REAL NOT NULL DEFAULT 0,
        unit_price      REAL NOT NULL DEFAULT 0,
        total           REAL NOT NULL DEFAULT 0,
        notes           TEXT,
        delivery_person TEXT,
        plate           TEXT,
        source          TEXT DEFAULT 'manual',
        created_by      TEXT,
        created_at      TEXT NOT NULL DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_date ON sales(sale_date);
    CREATE INDEX IF NOT EXISTS idx_type ON sales(sale_type);
    CREATE INDEX IF NOT EXISTS idx_sales_client ON sales(client);
    CREATE INDEX IF NOT EXISTS idx_sales_driver ON sales(delivery_person);
    CREATE INDEX IF NOT EXISTS idx_sales_type_date ON sales(sale_type, sale_date);
    CREATE INDEX IF NOT EXISTS idx_sales_client_date ON sales(client, sale_date);
    -- Ãndice composto p/ o ORDER BY de /api/sales (evita TEMP B-TREE / sort)
    CREATE INDEX IF NOT EXISTS idx_sales_date_time ON sales(sale_date DESC, sale_time DESC);
    CREATE INDEX IF NOT EXISTS idx_sales_type_date_time ON sales(sale_type, sale_date DESC, sale_time DESC);
    CREATE TABLE IF NOT EXISTS import_log (
        id          TEXT PRIMARY KEY,
        filename    TEXT,
        imported_at TEXT DEFAULT (datetime('now')),
        rows_added  INTEGER DEFAULT 0,
        status      TEXT,
        imported_by TEXT
    );
    CREATE TABLE IF NOT EXISTS product_prices (
        key         TEXT PRIMARY KEY,
        label       TEXT NOT NULL,
        price       REAL NOT NULL DEFAULT 0,
        price_min   REAL,
        price_max   REAL,
        is_variable INTEGER NOT NULL DEFAULT 0,
        updated_at  TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS price_history (
        id             TEXT PRIMARY KEY,
        key            TEXT NOT NULL,
        price          REAL NOT NULL,
        effective_date TEXT NOT NULL,
        note           TEXT,
        changed_by     TEXT,
        created_at     TEXT DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_ph ON price_history(key,effective_date);
    CREATE TABLE IF NOT EXISTS audit_log (
        id            TEXT PRIMARY KEY,
        timestamp     TEXT DEFAULT (datetime('now')),
        user_id       TEXT,
        username      TEXT,
        product_key   TEXT,
        product_label TEXT,
        field_changed TEXT,
        old_value     TEXT,
        new_value     TEXT,
        effective_date TEXT,
        note          TEXT,
        action        TEXT
    );
    CREATE TABLE IF NOT EXISTS whatsapp_contacts (
        id         TEXT PRIMARY KEY,
        name       TEXT NOT NULL,
        phone      TEXT NOT NULL,
        active     INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now')),
        created_by TEXT
    );
    CREATE TABLE IF NOT EXISTS whatsapp_config (
        key   TEXT PRIMARY KEY,
        value TEXT
    );
    CREATE TABLE IF NOT EXISTS whatsapp_log (
        id         TEXT PRIMARY KEY,
        phone      TEXT,
        contact    TEXT,
        event_type TEXT,
        message    TEXT,
        status     TEXT,
        response   TEXT,
        sent_at    TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS whatsapp_auto_rules (
        id            TEXT PRIMARY KEY,
        rule_type     TEXT NOT NULL DEFAULT 'trigger',
        name          TEXT NOT NULL,
        enabled       INTEGER NOT NULL DEFAULT 1,
        event_key     TEXT,
        keyword       TEXT,
        message       TEXT NOT NULL DEFAULT '',
        recipients    TEXT,
        priority      INTEGER DEFAULT 0,
        schedule_from TEXT,
        schedule_to   TEXT,
        cooldown_min  INTEGER DEFAULT 0,
        repeatable    INTEGER DEFAULT 1,
        max_per_day   INTEGER DEFAULT 0,
        last_run_at   TEXT,
        last_status   TEXT,
        created_at    TEXT DEFAULT (datetime('now')),
        updated_at    TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS whatsapp_templates (
        id         TEXT PRIMARY KEY,
        name       TEXT NOT NULL,
        category   TEXT,
        content    TEXT NOT NULL DEFAULT '',
        created_at TEXT DEFAULT (datetime('now')),
        updated_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT);
    CREATE TABLE IF NOT EXISTS clients (
        id         TEXT PRIMARY KEY,
        name       TEXT NOT NULL,
        cnpj       TEXT,
        cpf        TEXT,
        phone      TEXT,
        email      TEXT,
        address    TEXT,
        city       TEXT,
        notes      TEXT,
        active     INTEGER NOT NULL DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now')),
        created_by TEXT
    );
    CREATE TABLE IF NOT EXISTS vehicles (
        id         TEXT PRIMARY KEY,
        name       TEXT NOT NULL,
        plate      TEXT NOT NULL,
        driver     TEXT,
        active     INTEGER NOT NULL DEFAULT 1,
        notes      TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS drivers (
        id         TEXT PRIMARY KEY,
        name       TEXT NOT NULL UNIQUE,
        active     INTEGER NOT NULL DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now')),
        updated_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS app_config (
        key   TEXT PRIMARY KEY,
        value TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS quote_products (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        code TEXT,
        unit TEXT NOT NULL DEFAULT 'UND',
        default_price REAL NOT NULL DEFAULT 0,
        description TEXT,
        active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now')),
        updated_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS quotes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        quote_number INTEGER UNIQUE,
        company_key TEXT NOT NULL DEFAULT 'estrada',
        client_name TEXT NOT NULL,
        attention TEXT,
        client_cnpj TEXT,
        client_ie TEXT,
        client_phone TEXT,
        client_email TEXT,
        client_address TEXT,
        client_district TEXT,
        client_city TEXT,
        client_state TEXT,
        client_zip TEXT,
        issue_date TEXT NOT NULL,
        issue_time TEXT,
        validity_days INTEGER NOT NULL DEFAULT 3,
        delivery_deadline TEXT,
        payment_terms TEXT,
        observations TEXT,
        discount REAL NOT NULL DEFAULT 0,
        subtotal REAL NOT NULL DEFAULT 0,
        total REAL NOT NULL DEFAULT 0,
        status TEXT NOT NULL DEFAULT 'rascunho',
        created_by TEXT,
        created_at TEXT DEFAULT (datetime('now')),
        updated_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS quote_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        quote_id INTEGER NOT NULL,
        product_id INTEGER,
        item_order INTEGER NOT NULL DEFAULT 1,
        code TEXT,
        description TEXT NOT NULL,
        quantity REAL NOT NULL DEFAULT 1,
        unit TEXT NOT NULL DEFAULT 'UND',
        unit_price REAL NOT NULL DEFAULT 0,
        discount REAL NOT NULL DEFAULT 0,
        subtotal REAL NOT NULL DEFAULT 0,
        FOREIGN KEY(quote_id) REFERENCES quotes(id) ON DELETE CASCADE,
        FOREIGN KEY(product_id) REFERENCES quote_products(id)
    );
    CREATE INDEX IF NOT EXISTS idx_quotes_client ON quotes(client_name);
    CREATE INDEX IF NOT EXISTS idx_quotes_created ON quotes(created_at DESC);
    CREATE INDEX IF NOT EXISTS idx_quote_items_quote ON quote_items(quote_id);
    """)
    conn.executescript("""
        INSERT OR IGNORE INTO whatsapp_config (key, value) VALUES
            ('provider', 'ultramsg'), ('api_url', ''), ('api_token', ''),
            ('instance_id', ''), ('notify_boleto', '1'), ('notify_avaria', '1'),
            ('notify_inativo', '1'), ('avaria_min', '25'), ('inativo_dias', '30'),
            ('auto_period', '7'),
            ('bot_active', '1'), ('auto_reply_enabled', '1'),
            ('auto_reply_from', '08:00'), ('auto_reply_to', '19:00'),
            ('min_interval_secs', '30'), ('block_groups', '1'),
            ('block_duplicate', '1'), ('max_per_day_per_client', '10'),
            ('test_mode', '0'), ('motivation_enabled', '1'),
            ('motivation_time', '07:00'), ('motivation_last_success', ''),
            ('motivation_last_attempt', '');
    """)
    # Migrations for existing DBs
    for col_def in [
        ("sales","sale_time","TEXT"),("sales","plate","TEXT"),
        ("sales","created_by","TEXT"),
        ("price_history","changed_by","TEXT"),
        ("audit_log","user_id","TEXT"),("audit_log","username","TEXT"),
        ("audit_log","action","TEXT"),
        ("import_log","imported_by","TEXT"),
        ("product_prices","active","INTEGER NOT NULL DEFAULT 1"),
        ("sales","delivered","TEXT"),
        ("sales","delivered_at","TEXT"),
        ("sessions","last_seen","TEXT DEFAULT (datetime('now'))"),
        ("quotes","company_key","TEXT NOT NULL DEFAULT 'estrada'"),
    ]:
        try: conn.execute(f"ALTER TABLE {col_def[0]} ADD COLUMN {col_def[1]} {col_def[2]}")
        except sqlite3.OperationalError: pass  # coluna jÃ¡ existe â€” esperado em DB jÃ¡ migrado
    # Usuarios/senhas/sessoes pertencem ao grupo e ficam somente no banco principal.
    if _company_key(company or "raios")=="raios":
        admin_hash=hash_password("admin123")
        conn.execute("INSERT OR IGNORE INTO users(id,username,password_hash,full_name,role) VALUES(?,?,?,?,?)",
                     (str(uuid.uuid4()),"admin",admin_hash,"Administrador","admin"))
        conn.execute("INSERT OR IGNORE INTO settings(key,value) VALUES('admin_password',?)",(admin_hash,))
    # Seed product prices, respeitando produtos removidos pelo administrador.
    deleted_price_keys=set()
    try:
        row=conn.execute("SELECT value FROM settings WHERE key='deleted_product_keys'").fetchone()
        if row and row["value"]:
            deleted_price_keys=set(json.loads(row["value"]))
    except Exception:
        deleted_price_keys=set()
    for key,info in DEFAULT_PRICES.items():
        if key in deleted_price_keys:
            continue
        conn.execute("""INSERT OR IGNORE INTO product_prices(key,label,price,price_min,price_max,is_variable)
                        VALUES(?,?,?,?,?,?)""",
                     (key,info["label"],info["price"],info.get("price_min"),info.get("price_max"),int(info.get("is_variable",False))))
    existing_ph=conn.execute("SELECT COUNT(*) FROM price_history").fetchone()[0]
    if existing_ph==0:
        for key,info in DEFAULT_PRICES.items():
            conn.execute("INSERT OR IGNORE INTO price_history(id,key,price,effective_date,note,changed_by) VALUES(?,?,?,?,?,?)",
                         (str(uuid.uuid4()),key,info["price"],"2026-01-01","PreÃ§o inicial","sistema"))
    # Init boletos table + Ã­ndices
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS boletos (id TEXT PRIMARY KEY, client TEXT NOT NULL, sale_date TEXT NOT NULL, nf_number TEXT, total_val REAL NOT NULL DEFAULT 0, due_date TEXT, status TEXT NOT NULL DEFAULT 'pendente', paid_date TEXT, notes TEXT, created_at TEXT DEFAULT (datetime('now')), UNIQUE(client, sale_date, nf_number))""")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_boletos_status ON boletos(status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_boletos_due ON boletos(due_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_boletos_paid ON boletos(paid_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_boletos_client_date ON boletos(client, sale_date)")
    except sqlite3.OperationalError as e:
        print(f"âš ï¸  init_boletos (create): {e}")
    # Tabela Paladar (mÃ³dulo independente) â€” INTEGER PRIMARY KEY AUTOINCREMENT
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS paladar_sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_group TEXT,
            saledate TEXT NOT NULL,
            product TEXT NOT NULL,
            quantity REAL NOT NULL DEFAULT 1,
            unitprice REAL NOT NULL DEFAULT 0,
            total REAL NOT NULL DEFAULT 0,
            notes TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        )""")
    except sqlite3.OperationalError as e:
        print(f"âš ï¸  init_paladar: {e}")
    # Migration: se a tabela antiga (TEXT id) existir, migrar dados
    try:
        old = conn.execute("PRAGMA table_info(paladar_sales)").fetchall()
        is_old = any(r[1]=='id' and r[2] for r in old if r[2] and 'TEXT' in r[2].upper())
        if is_old:
            conn.executescript("""
                CREATE TABLE paladar_sales_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sale_group TEXT,
                    saledate TEXT NOT NULL,
                    product TEXT NOT NULL,
                    quantity REAL NOT NULL DEFAULT 1,
                    unitprice REAL NOT NULL DEFAULT 0,
                    total REAL NOT NULL DEFAULT 0,
                    notes TEXT,
                    created_at TEXT DEFAULT (datetime('now'))
                );
                INSERT INTO paladar_sales_new (saledate,product,quantity,unitprice,total,notes,created_at)
                    SELECT saledate,product,quantity,unitprice,total,notes,created_at FROM paladar_sales;
                DROP TABLE paladar_sales;
                ALTER TABLE paladar_sales_new RENAME TO paladar_sales;
            """)
            print("âœ…  paladar_sales migrada para INTEGER PRIMARY KEY AUTOINCREMENT")
    except Exception as e:
        print(f"âš ï¸  paladar migration: {e}")
    # Tabela de produtos do Paladar (cadastro separado do sistema principal)
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS paladar_products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            suggested_price REAL DEFAULT 0,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
    except sqlite3.OperationalError as e:
        print(f"âš ï¸  init_paladar_products: {e}")
    # Migration: adicionar colunas de documento e logÃ­stica ao paladar_sales
    for col in ['nf_number','driver','vehicle','plate']:
        try:
            conn.execute(f"ALTER TABLE paladar_sales ADD COLUMN {col} TEXT")
        except sqlite3.OperationalError:
            pass
    # Migration: coluna client para paladar_sales
    try:
        conn.execute("ALTER TABLE paladar_sales ADD COLUMN client TEXT")
    except sqlite3.OperationalError:
        pass
    # Migration: anexos de nota fiscal por grupo de venda Monteiro
    for col in ['invoice_file_path','invoice_original_name','invoice_mime','invoice_uploaded_at']:
        try:
            conn.execute(f"ALTER TABLE paladar_sales ADD COLUMN {col} TEXT")
        except sqlite3.OperationalError:
            pass
    # Indices usados pelos relatorios/filtros mais pesados.
    for idx_sql in [
        "CREATE INDEX IF NOT EXISTS idx_paladar_sales_client_date ON paladar_sales(client, saledate)",
        "CREATE INDEX IF NOT EXISTS idx_paladar_sales_date ON paladar_sales(saledate)",
        "CREATE INDEX IF NOT EXISTS idx_paladar_sales_group ON paladar_sales(sale_group)",
        "CREATE INDEX IF NOT EXISTS idx_paladar_sales_product_date ON paladar_sales(product, saledate)",
    ]:
        try:
            conn.execute(idx_sql)
        except sqlite3.OperationalError as e:
            print(f"indice paladar_sales falhou: {e}")
    # Tabela de pagamentos do Monteiro
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS monteiro_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client TEXT NOT NULL,
            payment_date TEXT NOT NULL,
            amount REAL NOT NULL DEFAULT 0,
            month TEXT,
            year TEXT,
            payment_type TEXT DEFAULT 'repasse',
            notes TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        )""")
    except sqlite3.OperationalError as e:
        print(f"âš ï¸  monteiro_payments: {e}")
    # Migration: add nf_number column + recreate UNIQUE para bancos existentes
    try:
        conn.execute("ALTER TABLE boletos ADD COLUMN nf_number TEXT")
        # Column didn't exist â†’ table has old UNIQUE(client, sale_date). Recreate.
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS boletos_new (
                id TEXT PRIMARY KEY,
                client TEXT NOT NULL,
                sale_date TEXT NOT NULL,
                nf_number TEXT,
                total_val REAL NOT NULL DEFAULT 0,
                due_date TEXT,
                status TEXT NOT NULL DEFAULT 'pendente',
                paid_date TEXT,
                notes TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                UNIQUE(client, sale_date, nf_number)
            );
            INSERT OR IGNORE INTO boletos_new
                SELECT id, client, sale_date,
                    COALESCE(
                        (SELECT MAX(nf_number) FROM sales s
                         WHERE TRIM(s.client) COLLATE NOCASE = TRIM(boletos.client)
                         AND s.sale_date = boletos.sale_date
                         AND s.sale_type != 'AVARIA'
                         LIMIT 1),
                    ''),
                    total_val, due_date, status, paid_date, notes, created_at
                FROM boletos;
            DROP TABLE boletos;
            ALTER TABLE boletos_new RENAME TO boletos;
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_boletos_status ON boletos(status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_boletos_due ON boletos(due_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_boletos_paid ON boletos(paid_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_boletos_client_date ON boletos(client, sale_date)")
    except sqlite3.OperationalError:
        # nf_number jÃ¡ existe â€” migraÃ§Ã£o jÃ¡ rodou, esperado
        pass
    conn.commit(); conn.close()

def load_prices() -> Dict[str,float]:
    conn=get_db()
    rows=conn.execute("SELECT key,price FROM product_prices").fetchall()
    conn.close()
    p={r["key"]:r["price"] for r in rows}
    for k,v in DEFAULT_PRICES.items(): p.setdefault(k,v["price"])
    return p

@lru_cache(maxsize=4096)
def _get_price_on_date_cached(company:str,key:str,date_str:str,bucket:int)->float:
    conn=get_db(company)
    row=conn.execute("SELECT price FROM price_history WHERE key=? AND effective_date<=? ORDER BY effective_date DESC LIMIT 1",
                     (key,date_str)).fetchone()
    conn.close()
    return float(row["price"]) if row else DEFAULT_PRICES.get(key,{}).get("price",0)

def clear_price_cache():
    _get_price_on_date_cached.cache_clear()

def get_price_on_date(key:str,date_str:str)->float:
    bucket=int(_time_mod.time()//300)
    return _get_price_on_date_cached(_company_key(CURRENT_COMPANY.get()),str(key or ""),str(date_str or ""),bucket)

def _cache_bucket(seconds:int=60)->int:
    return int(_time_mod.time()//max(1,seconds))

def clear_sales_cache():
    _summary_cached.cache_clear()
    _summary_array_cached.cache_clear()
    _consolidado_cached.cache_clear()

# â”€â”€ Auth â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def get_session(x_token:str="")->dict:
    if not x_token: return {}
    conn=get_control_db()
    row=conn.execute("SELECT * FROM sessions WHERE token=? AND expires_at>datetime('now')",(x_token,)).fetchone()
    if row:
        has_last_seen="last_seen" in row.keys()
        last=row["last_seen"] if has_last_seen else None
        if last:
            try:
                last_dt=datetime.fromisoformat(str(last).replace(" ","T"))
                if datetime.now()-last_dt > timedelta(minutes=SESSION_IDLE_MINUTES):
                    conn.execute("DELETE FROM sessions WHERE token=?",(x_token,))
                    conn.commit(); conn.close()
                    return {}
            except Exception:
                pass
        if has_last_seen:
            conn.execute("UPDATE sessions SET last_seen=datetime('now') WHERE token=?",(x_token,))
            conn.commit()
    conn.close()
    return dict(row) if row else {}

def require_auth(x_token:str="")->dict:
    sess=get_session(x_token)
    if not sess: raise HTTPException(401,"SessÃ£o invÃ¡lida. FaÃ§a login novamente.")
    return sess

def require_admin(x_token:str="")->dict:
    sess=require_auth(x_token)
    if sess.get("role")!="admin": raise HTTPException(403,"Apenas administradores podem fazer isso.")
    return sess

def require_editor(x_token:str="")->dict:
    sess=require_auth(x_token)
    if sess.get("role") not in ("admin","editor"): raise HTTPException(403,"PermissÃ£o insuficiente.")
    return sess

def require_admin_or_editor(x_token:str="")->dict:
    sess=require_auth(x_token)
    if sess.get("role") not in ("admin","editor"): raise HTTPException(403,"PermissÃ£o insuficiente.")
    return sess

def role_has_tab_permission(role:str, permission_key:str)->bool:
    if role=="admin":
        return True
    conn=get_control_db()
    try:
        row=conn.execute("SELECT value FROM settings WHERE key='tab_permissions'").fetchone()
        defaults={
            "viewer":["consolidado","nf","pr","avulso","avaria","projecao","grafico","clientes","produtividade","boletos","pendentes","produtos"],
            "editor":["consolidado","nf","pr","avulso","avaria","projecao","grafico","clientes","produtividade","boletos","pendentes","produtos","config","cfg_precos"],
            "admin":["consolidado","nf","pr","avulso","avaria","projecao","grafico","clientes","produtividade","boletos","pendentes","produtos","config","cfg_precos","cfg_registro","cfg_importar","cfg_whatsapp","cfg_sebrae"],
        }
        perms=json.loads(row["value"]) if row and row["value"] else defaults
        allowed=perms.get(role,[]) if isinstance(perms,dict) else []
        if role=="editor" and permission_key=="cfg_precos" and ("config" in allowed or any(str(k).startswith("cfg_") for k in allowed)):
            return True
        return any(k in allowed for k in _expand_tab_keys([permission_key]))
    except Exception:
        return False
    finally:
        conn.close()

def _tab_permissions_map()->dict:
    return tab_permissions_map_from_db(get_control_db)

def _permissions_configured()->bool:
    return permissions_configured_from_map(_tab_permissions_map())

def _session_has_any_tab(sess:dict, keys:list)->bool:
    return session_has_any_tab_from_map(sess, keys, _tab_permissions_map())

def require_any_tab_access(x_token:str, keys:list)->dict:
    sess=require_auth(x_token)
    if not _permissions_configured() or _session_has_any_tab(sess, keys):
        return sess
    raise HTTPException(403,"Sem permissao para acessar esta area.")

def require_editor_tab_access(x_token:str, keys:list)->dict:
    sess=require_editor(x_token)
    if not _permissions_configured() or _session_has_any_tab(sess, keys):
        return sess
    raise HTTPException(403,"Sem permissao para editar esta area.")

def require_prices_access(x_token:str="")->dict:
    sess=require_auth(x_token)
    if sess.get("role")=="admin" or role_has_tab_permission(sess.get("role",""),"cfg_precos"):
        return sess
    raise HTTPException(403,"Sem permissao para editar precos.")

def find_legacy_company_user(username:str,password:str)->Optional[dict]:
    """Compatibilidade: encontra login salvo em banco de empresa antigo.
    Usuarios agora pertencem ao grupo, mas essa ponte evita bloquear contas
    que tiveram senha alterada enquanto uma empresa especifica estava ativa."""
    for company in COMPANY_DBS:
        if _company_key(company)=="raios":
            continue
        conn=None
        try:
            conn=get_db(company)
            row=conn.execute("SELECT * FROM users WHERE username=? AND active=1",(username,)).fetchone()
            if row and verify_password(password,row["password_hash"]):
                return dict(row)
        except Exception:
            pass
        finally:
            try:
                if conn: conn.close()
            except Exception:
                pass
    return None

def sync_legacy_user_to_control(conn, legacy:dict)->dict:
    username=legacy.get("username","")
    existing=conn.execute("SELECT id FROM users WHERE username=?",(username,)).fetchone()
    full_name=legacy.get("full_name") or username
    role=legacy.get("role") or "editor"
    pw_hash=legacy.get("password_hash") or ""
    if existing:
        conn.execute("""UPDATE users
                        SET password_hash=?, full_name=?, role=?, active=1
                        WHERE username=?""",(pw_hash,full_name,role,username))
    else:
        conn.execute("""INSERT INTO users(id,username,password_hash,full_name,role,active,created_by)
                        VALUES(?,?,?,?,?,1,?)""",
                     (legacy.get("id") or str(uuid.uuid4()),username,pw_hash,full_name,role,"legacy-sync"))
    conn.commit()
    return dict(conn.execute("SELECT * FROM users WHERE username=? AND active=1",(username,)).fetchone())

def log_action(conn,sess:dict,action:str,product_key:str="",product_label:str="",
               field:str="",old_val:str="",new_val:str="",eff_date:str="",note:str=""):
    """Registra auditoria. NUNCA quebra a operaÃ§Ã£o principal â€” auto-migra colunas
    faltantes em bancos antigos e ignora falhas de log silenciosamente."""
    try:
        # Auto-migraÃ§Ã£o: garante que todas as colunas existem (bancos antigos)
        existing={r[1] for r in conn.execute("PRAGMA table_info(audit_log)").fetchall()}
        needed={"id":"TEXT","timestamp":"TEXT","user_id":"TEXT","username":"TEXT",
                "product_key":"TEXT","product_label":"TEXT","field_changed":"TEXT",
                "old_value":"TEXT","new_value":"TEXT","effective_date":"TEXT",
                "note":"TEXT","action":"TEXT"}
        if not existing:
            conn.execute("""CREATE TABLE IF NOT EXISTS audit_log(
                id TEXT PRIMARY KEY,timestamp TEXT,user_id TEXT,username TEXT,
                product_key TEXT,product_label TEXT,field_changed TEXT,old_value TEXT,
                new_value TEXT,effective_date TEXT,note TEXT,action TEXT)""")
        else:
            for col,typ in needed.items():
                if col not in existing:
                    try: conn.execute(f"ALTER TABLE audit_log ADD COLUMN {col} {typ}")
                    except Exception: pass
        conn.execute("""INSERT INTO audit_log(id,timestamp,user_id,username,product_key,product_label,
                        field_changed,old_value,new_value,effective_date,note,action)
                        VALUES(?,datetime('now'),?,?,?,?,?,?,?,?,?,?)""",
                     (str(uuid.uuid4()),sess.get("user_id",""),sess.get("username",""),
                      product_key,product_label,field,old_val,new_val,eff_date,note,action))
    except Exception:
        pass  # log Ã© secundÃ¡rio; nunca deve quebrar a operaÃ§Ã£o

# â”€â”€ Backup â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def create_backup(label:str="auto")->str:
    """Cria um pacote .zip com todos os bancos registrados do sistema."""
    import json as _j
    import tempfile
    import zipfile
    today=datetime.now().strftime("%Y%m%d")
    if label=="auto":
        existing=list(BACKUP_DIR.glob(f"bm_backup_{today}_*_auto.*"))
        if existing: return str(existing[0].name)  # Already backed up today
    sources=_backup_db_sources()
    if not sources: return ""
    ts=datetime.now().strftime("%Y%m%d_%H%M%S")
    fname=f"bm_backup_{ts}_{label}.zip"
    dest=BACKUP_DIR/fname
    manifest={"created_at":datetime.now().isoformat(),"label":label,"version":"multi-db-v1","databases":[]}
    with tempfile.TemporaryDirectory() as td:
        tmpdir=Path(td)
        with zipfile.ZipFile(dest,"w",compression=zipfile.ZIP_DEFLATED,compresslevel=6) as zf:
            for item in sources:
                src=item["path"]
                tmp=tmpdir/src.name
                _copy_sqlite_consistent(src,tmp)
                arc=f"databases/{src.name}"
                zf.write(tmp,arcname=arc)
                manifest["databases"].append({
                    "key":item["key"],"label":item["label"],"filename":src.name,
                    "size":src.stat().st_size,"archive_path":arc
                })
            zf.writestr("manifest.json",_j.dumps(manifest,ensure_ascii=False,indent=2))
    # Remove backups antigos (manter MAX_BACKUPS)
    all_bk=sorted(_backup_files(),key=lambda p:p.stat().st_mtime)
    for old in all_bk[:-MAX_BACKUPS]: old.unlink(missing_ok=True)
    return fname

def _backup_db_sources():
    return backup_db_sources_from_paths(BASE_DIR, DB_PATH, COMPANY_DBS, APP_NOTES_DB_PATH)

def _backup_expected_databases():
    return backup_expected_databases_from_paths(BASE_DIR, DB_PATH, COMPANY_DBS, APP_NOTES_DB_PATH)

def _backup_files():
    return backup_files_from_dir(BACKUP_DIR)

def _backup_manifest_databases(path:Path):
    return backup_manifest_databases_from_zip(path)

def backup_scheduler():
    """Thread de backup automÃ¡tico diÃ¡rio Ã s 03:00."""
    import time
    while True:
        now=datetime.now()
        # PrÃ³ximas 03:00
        next_run=now.replace(hour=3,minute=0,second=0,microsecond=0)
        if now>=next_run: next_run=next_run+timedelta(days=1)
        wait=(next_run-now).total_seconds()
        time.sleep(wait)
        try:
            fname=create_backup("auto")
            print(f"âœ… Backup automÃ¡tico: {fname}")
        except Exception as e:
            print(f"âŒ Erro no backup: {e}")
        # Limpeza diÃ¡ria de sessÃµes expiradas (nÃ£o deixa o banco crescer indefinidamente)
        try:
            c=get_control_db()
            n=c.execute("DELETE FROM sessions WHERE expires_at < datetime('now')").rowcount
            c.commit(); c.close()
            if n>0: print(f"ðŸ§¹ {n} sessÃµes expiradas removidas")
        except Exception as e:
            print(f"âš ï¸  Cleanup de sessÃµes falhou: {e}")


# â”€â”€ FastAPI â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app_instance):
    """Inicializa banco e inicia scheduler de backup."""
    for company in COMPANY_DBS:
        init_db(company)
    # Limpeza de sessÃµes expiradas ao subir o serviÃ§o
    try:
        c=get_control_db()
        n=c.execute("DELETE FROM sessions WHERE expires_at < datetime('now')").rowcount
        c.commit(); c.close()
        if n>0: print(f"ðŸ§¹ Startup: {n} sessÃµes expiradas removidas")
    except Exception as e:
        print(f"âš ï¸  Startup cleanup falhou: {e}")
    # Criar backup inicial ao subir o servidor (apenas se nÃ£o houver backup hoje)
    try:
        today=datetime.now().strftime("%Y%m%d")
        existing=list(BACKUP_DIR.glob(f"bm_backup_{today}_*.*"))
        if not existing:
            create_backup("startup")
    except Exception as e:
        print(f"âš ï¸  Backup inicial falhou: {e}")
    # Iniciar thread de backup automÃ¡tico
    bk_thread=threading.Thread(target=backup_scheduler,daemon=True)
    bk_thread.start()
    motivation_thread=threading.Thread(target=motivation_scheduler,daemon=True)
    motivation_thread.start()
    yield

app=FastAPI(title="Menina dos Raios Ltda API", lifespan=lifespan)
# GZip: comprime respostas JSON grandes (ex: /api/sales ~600KB -> ~80KB).
# Reduz drasticamente o tempo de carregamento pela rede.
from fastapi.middleware.gzip import GZipMiddleware
app.add_middleware(GZipMiddleware, minimum_size=500)
# CORS: padrÃ£o estrito (apenas domÃ­nio de produÃ§Ã£o).
# Dev/local: setar variÃ¡vel de ambiente CORS_ORIGINS="*" ou lista separada por vÃ­rgula.
_default_origins = "https://sistema.meninadosraios.com.br,https://www.sistema.meninadosraios.com.br"
ALLOWED_ORIGINS = [o.strip() for o in os.getenv("CORS_ORIGINS", _default_origins).split(",") if o.strip()]
app.add_middleware(CORSMiddleware,
                   allow_origins=ALLOWED_ORIGINS,
                   allow_methods=["GET","POST","PUT","DELETE","OPTIONS"],
                   allow_headers=["Content-Type","x-token","x-company","x-app-token","x-real-ip","Authorization","Accept"])

@app.middleware("http")
async def company_context_middleware(request: Request, call_next):
    company=request.headers.get("x-company") or request.query_params.get("company") or "raios"
    if request.url.path.startswith(("/api/monteiro", "/api/paladar")):
        company="raios"
    token = CURRENT_COMPANY.set(_company_key(company))
    try:
        return await call_next(request)
    finally:
        CURRENT_COMPANY.reset(token)

# â”€â”€ Auth endpoints â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.post("/api/auth/login")
def login(body:LoginIn, request:Request):
    ip = _client_ip(request)
    allowed, retry_secs = _check_login_rate(ip)
    if not allowed:
        raise HTTPException(429, f"Muitas tentativas de login. Tente novamente em {retry_secs} segundos.")
    login_company=CURRENT_COMPANY.get()
    conn=get_control_db()
    user=conn.execute("SELECT * FROM users WHERE username=? AND active=1",
                      (body.username,)).fetchone()
    if not user or not verify_password(body.password, user["password_hash"]):
        legacy=find_legacy_company_user(body.username,body.password)
        if legacy:
            user=sync_legacy_user_to_control(conn,legacy)
        else:
            _record_login(ip, False)
            conn.close(); raise HTTPException(401,"Usuario ou senha incorretos.")
    _record_login(ip, True)
    if needs_rehash(user["password_hash"]):
        try:
            conn.execute("UPDATE users SET password_hash=? WHERE id=?",
                         (hash_password(body.password), user["id"]))
        except Exception: pass
    token=str(uuid.uuid4())
    expires=(datetime.now()+timedelta(hours=SESSION_HOURS)).strftime("%Y-%m-%d %H:%M:%S")
    try:
        conn.execute("INSERT INTO sessions(token,user_id,username,full_name,role,last_seen,expires_at) VALUES(?,?,?,?,?,datetime('now'),?)",
                     (token,user["id"],user["username"],user["full_name"],user["role"],expires))
    except sqlite3.OperationalError as e:
        if "last_seen" not in str(e).lower():
            raise
        conn.execute("INSERT INTO sessions(token,user_id,username,full_name,role,expires_at) VALUES(?,?,?,?,?,?)",
                     (token,user["id"],user["username"],user["full_name"],user["role"],expires))
    conn.commit(); conn.close()
    return {"token":token,"username":user["username"],"full_name":user["full_name"],"role":user["role"],"company":login_company}

@app.post("/api/auth/logout")
def logout(x_token:str=Header("")):
    conn=get_control_db(); conn.execute("DELETE FROM sessions WHERE token=?",(x_token,)); conn.commit(); conn.close()
    return {"ok":True}

# Vales solicitados pelo aplicativo Android (offline-first)
def _vale_key(nome:str)->str:
    return _normalize_name(str(nome or "").strip())

def _clean_mobile_vale(body:dict):
    client_id=str(body.get("client_id") or "").strip()[:100]
    if not client_id:
        raise HTTPException(400,"client_id obrigatorio.")
    solicitante_nome=re.sub(r"\s+"," ",str(body.get("solicitante_nome") or "").strip())[:180]
    if not solicitante_nome:
        raise HTTPException(400,"Informe o nome do solicitante.")
    try:
        amount=round(float(body.get("amount") or 0),2)
    except Exception:
        raise HTTPException(400,"Valor invalido.")
    if amount<=0:
        raise HTTPException(400,"O valor deve ser maior que zero.")
    request_date=str(body.get("request_date") or "").strip()[:10]
    try:
        datetime.strptime(request_date,"%Y-%m-%d")
    except Exception:
        raise HTTPException(400,"Data do vale invalida.")
    signature=str(body.get("signature_png_base64") or "").strip()
    if not signature:
        raise HTTPException(400,"Assinatura obrigatoria.")
    signature_format=str(body.get("signature_format") or "png").strip().lower()[:20] or "png"
    if signature_format!="png":
        raise HTTPException(400,"Formato de assinatura invalido.")
    source=str(body.get("source") or "android_app").strip()[:40] or "android_app"
    return client_id,solicitante_nome,amount,request_date,signature,signature_format,source

def _signature_png_bytes(signature:str)->bytes:
    raw=str(signature or "").strip()
    if "," in raw:
        raw=raw.split(",",1)[1]
    try:
        data=base64.b64decode(raw,validate=False)
    except Exception:
        data=b""
    if not data:
        raise HTTPException(404,"Assinatura nao encontrada.")
    return data

def _vale_dict(row):
    item=dict(row)
    item.pop("signature_png_base64",None)
    item["signature_url"]="/api/mobile/vales/"+str(item.get("id"))+"/signature"
    return item

@app.post("/api/mobile/vales")
def create_mobile_vale(body:dict,x_token:str=Header("")):
    sess=require_auth(x_token)
    client_id,solicitante_nome,amount,request_date,signature,signature_format,source=_clean_mobile_vale(body)
    now=datetime.now().isoformat(timespec="seconds")
    conn=get_app_notes_db()
    try:
        existing=conn.execute("SELECT * FROM app_vales WHERE client_id=?",(client_id,)).fetchone()
        if existing:
            return {"success":True,"duplicate":True,"vale":_vale_dict(existing)}
        vale_id=str(uuid.uuid4())
        conn.execute("""INSERT INTO app_vales
            (id,client_id,solicitante_nome,solicitante_key,amount,request_date,signature_png_base64,signature_format,
             registered_by_user_id,registered_by_username,registered_by_name,status,source,created_at,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (vale_id,client_id,solicitante_nome,_vale_key(solicitante_nome),amount,request_date,signature,signature_format,
             str(sess.get("user_id") or ""),str(sess.get("username") or ""),str(sess.get("full_name") or sess.get("username") or ""),
             "sincronizado",source,now,now))
        conn.commit()
        row=conn.execute("SELECT * FROM app_vales WHERE id=?",(vale_id,)).fetchone()
        return {"success":True,"duplicate":False,"vale":_vale_dict(row)}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

@app.get("/api/mobile/vales")
def list_mobile_vales(solicitante_nome:Optional[str]=None,date_from:Optional[str]=None,date_to:Optional[str]=None,status:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token)
    conn=get_app_notes_db(); where=[]; params=[]
    try:
        if solicitante_nome:
            where.append("solicitante_key LIKE ?"); params.append("%"+_vale_key(solicitante_nome)+"%")
        if date_from:
            where.append("date(request_date)>=date(?)"); params.append(str(date_from)[:10])
        if date_to:
            where.append("date(request_date)<=date(?)"); params.append(str(date_to)[:10])
        if status:
            where.append("status=?"); params.append(str(status)[:40])
        sql="SELECT * FROM app_vales"+(" WHERE "+" AND ".join(where) if where else "")+" ORDER BY request_date DESC, created_at DESC LIMIT 1000"
        rows=conn.execute(sql,params).fetchall()
        items=[_vale_dict(r) for r in rows]
        return {"success":True,"items":items,"count":len(items),"total":round(sum(float(i.get("amount") or 0) for i in items),2)}
    finally:
        conn.close()

@app.delete("/api/mobile/vales/{vale_id}")
def delete_mobile_vale(vale_id:str,x_token:str=Header("")):
    require_auth(x_token)
    vale_id=str(vale_id or "").strip()
    if not vale_id:
        raise HTTPException(400,"ID do vale obrigatorio.")
    conn=get_app_notes_db()
    try:
        row=conn.execute("SELECT id,solicitante_nome,amount,request_date FROM app_vales WHERE id=?",(vale_id,)).fetchone()
        if not row:
            raise HTTPException(404,"Vale nao encontrado.")
        conn.execute("DELETE FROM app_vales WHERE id=?",(vale_id,))
        conn.commit()
        return {"success":True,"deleted":True,"vale":dict(row)}
    except HTTPException:
        conn.rollback()
        raise
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

@app.get("/api/mobile/vales/{vale_id}/signature")
def mobile_vale_signature(vale_id:str,x_token:str=Header("")):
    require_auth(x_token)
    conn=get_app_notes_db()
    try:
        row=conn.execute("SELECT signature_png_base64 FROM app_vales WHERE id=?",(vale_id,)).fetchone()
        if not row:
            raise HTTPException(404,"Vale nao encontrado.")
        data=_signature_png_bytes(row["signature_png_base64"])
        return StreamingResponse(io.BytesIO(data),media_type="image/png",headers={"Cache-Control":"private, no-store"})
    finally:
        conn.close()

@app.get("/api/auth/me")
def me(x_token:str=Header("")):
    sess=require_auth(x_token)
    return {"username":sess["username"],"full_name":sess.get("full_name"),"role":sess["role"]}

@app.post("/api/auth/company-switch")
def switch_company_session(body:dict,x_token:str=Header("")):
    """Troca apenas a empresa ativa no frontend; usuario/sessao sao do grupo."""
    sess=require_auth(x_token)
    target=_company_key(body.get("company","raios"))
    return {"token":x_token,"username":sess["username"],"full_name":sess.get("full_name"),"role":sess["role"],"company":target}

@app.put("/api/auth/change-password")
def change_own_password(body:dict,x_token:str=Header("")):
    """Qualquer usuÃ¡rio logado pode alterar sua prÃ³pria senha."""
    sess=require_auth(x_token)
    old_pw=body.get("old_password","")
    new_pw=body.get("new_password","")
    validate_password(new_pw, sess.get("username",""))
    conn=get_control_db()
    user=conn.execute("SELECT * FROM users WHERE id=?",(sess["user_id"],)).fetchone()
    if not user or not verify_password(old_pw, user["password_hash"]):
        conn.close(); raise HTTPException(400,"Senha atual incorreta.")
    new_hash=hash_password(new_pw)
    conn.execute("UPDATE users SET password_hash=? WHERE id=?",(new_hash,sess["user_id"]))
    log_action(conn,sess,"CHANGE_PASSWORD","","","password","***","***","","Senha alterada pelo prÃ³prio usuÃ¡rio")
    conn.commit(); conn.close()
    return {"ok":True}

# â”€â”€ User management (admin only) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get("/api/users")
def list_users(x_token:str=Header("")):
    require_admin(x_token)
    conn=get_control_db()
    rows=conn.execute("SELECT id,username,full_name,role,active,created_at FROM users ORDER BY created_at").fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.post("/api/users")
def create_user(body:UserIn,x_token:str=Header("")):
    sess=require_admin(x_token)
    validate_password(body.password, body.username)
    pw_hash=hash_password(body.password)
    new_id=str(uuid.uuid4()); conn=get_control_db()
    try:
        conn.execute("INSERT INTO users(id,username,password_hash,full_name,role,created_by) VALUES(?,?,?,?,?,?)",
                     (new_id,body.username,pw_hash,body.full_name,body.role,sess["username"]))
        conn.commit()
    except Exception as e:
        conn.close(); raise HTTPException(400,f"UsuÃ¡rio jÃ¡ existe ou dados invÃ¡lidos: {e}")
    conn.close(); return {"ok":True,"id":new_id}

@app.put("/api/users/{user_id}")
def update_user(user_id:str,body:dict,x_token:str=Header("")):
    sess=require_admin(x_token); conn=get_control_db()
    target=conn.execute("SELECT * FROM users WHERE id=?",(user_id,)).fetchone()
    if not target: conn.close(); raise HTTPException(404,"UsuÃ¡rio nÃ£o encontrado.")
    if "role" in body and body["role"]!="admin" and target["role"]=="admin":
        admin_cnt=conn.execute("SELECT COUNT(*) FROM users WHERE role='admin' AND active=1").fetchone()[0]
        if admin_cnt<=1: conn.close(); raise HTTPException(400,"NÃ£o Ã© possÃ­vel remover o Ãºnico administrador.")
    if "password" in body and body["password"]:
        validate_password(body["password"], target["username"])
        conn.execute("UPDATE users SET password_hash=? WHERE id=?",(hash_password(body["password"]),user_id))
    if "role" in body: conn.execute("UPDATE users SET role=? WHERE id=?",(body["role"],user_id))
    if "active" in body: conn.execute("UPDATE users SET active=? WHERE id=?",(int(body["active"]),user_id))
    if "full_name" in body: conn.execute("UPDATE users SET full_name=? WHERE id=?",(body["full_name"],user_id))
    conn.commit(); conn.close(); return {"ok":True}

@app.delete("/api/users/{user_id}")
def delete_user(user_id:str,x_token:str=Header("")):
    sess=require_admin(x_token)
    if user_id==sess["user_id"]: raise HTTPException(400,"NÃ£o pode excluir sua prÃ³pria conta.")
    conn=get_control_db(); conn.execute("DELETE FROM users WHERE id=?",(user_id,)); conn.commit(); conn.close()
    return {"ok":True}

@app.get("/api/admin/messages")
def list_admin_messages(x_token:str=Header("")):
    require_admin(x_token)
    conn=get_control_db()
    rows=conn.execute("""
        SELECT m.id,m.user_id,m.title,m.body,m.created_at,m.expires_at,m.seen_at,m.removed_by_admin,m.created_by,
               u.username,u.full_name,u.role
        FROM admin_messages m
        LEFT JOIN users u ON u.id=m.user_id
        ORDER BY m.created_at DESC
        LIMIT 300
    """).fetchall()
    conn.close()
    out=[]
    now=datetime.now()
    for r in rows:
        d=dict(r)
        expired=False
        if d.get("expires_at"):
            try: expired=datetime.fromisoformat(str(d["expires_at"]).replace("Z","")) < now
            except Exception: expired=False
        if d.get("removed_by_admin"): status="removida"
        elif d.get("seen_at"): status="vista"
        elif expired: status="expirada"
        else: status="pendente"
        d["status"]=status
        out.append(d)
    return out

@app.post("/api/admin/messages")
def create_admin_message(body:AdminMessageIn,x_token:str=Header("")):
    sess=require_admin(x_token)
    title=(body.title or "").strip()
    text=(body.body or "").strip()
    uid=(body.user_id or "").strip()
    if not uid: raise HTTPException(400,"Selecione um usuÃ¡rio.")
    if not title: raise HTTPException(400,"Informe o tÃ­tulo da mensagem.")
    if not text: raise HTTPException(400,"Informe o texto da mensagem.")
    if len(title)>120: raise HTTPException(400,"TÃ­tulo muito longo.")
    if len(text)>2000: raise HTTPException(400,"Mensagem muito longa.")
    days=body.expires_days if body.expires_days is not None else 7
    try: days=int(days)
    except Exception: days=7
    if days<1 or days>90: raise HTTPException(400,"Validade deve ficar entre 1 e 90 dias.")
    expires=(datetime.now()+timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
    conn=get_control_db()
    user=conn.execute("SELECT id FROM users WHERE id=? AND active=1",(uid,)).fetchone()
    if not user:
        conn.close()
        raise HTTPException(404,"UsuÃ¡rio nÃ£o encontrado ou inativo.")
    mid=str(uuid.uuid4())
    conn.execute("""INSERT INTO admin_messages(id,user_id,title,body,expires_at,created_by)
                    VALUES(?,?,?,?,?,?)""",(mid,uid,title,text,expires,sess.get("username","admin")))
    conn.commit(); conn.close()
    return {"ok":True,"id":mid}

@app.delete("/api/admin/messages/{message_id}")
def remove_admin_message(message_id:str,x_token:str=Header("")):
    require_admin(x_token)
    conn=get_control_db()
    row=conn.execute("SELECT id FROM admin_messages WHERE id=?",(message_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404,"Mensagem nÃ£o encontrada.")
    conn.execute("UPDATE admin_messages SET removed_by_admin=1 WHERE id=?",(message_id,))
    conn.commit(); conn.close()
    return {"ok":True}

@app.get("/api/messages/pending")
def pending_admin_message(x_token:str=Header("")):
    sess=require_auth(x_token)
    conn=get_control_db()
    row=conn.execute("""
        SELECT m.id,m.title,m.body,m.created_at,m.expires_at,m.created_by
        FROM admin_messages m
        WHERE m.user_id=? AND m.seen_at IS NULL AND m.removed_by_admin=0
          AND (m.expires_at IS NULL OR m.expires_at>datetime('now'))
        ORDER BY m.created_at ASC
        LIMIT 1
    """,(sess["user_id"],)).fetchone()
    conn.close()
    return dict(row) if row else {}

@app.patch("/api/messages/{message_id}/seen")
def mark_admin_message_seen(message_id:str,x_token:str=Header("")):
    sess=require_auth(x_token)
    conn=get_control_db()
    row=conn.execute("SELECT id FROM admin_messages WHERE id=? AND user_id=?",(message_id,sess["user_id"])).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404,"Mensagem nÃ£o encontrada.")
    conn.execute("UPDATE admin_messages SET seen_at=datetime('now') WHERE id=? AND seen_at IS NULL",(message_id,))
    conn.commit(); conn.close()
    return {"ok":True}

# â”€â”€ Sales â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get("/api/sales")
def list_sales(sale_type:Optional[str]=None,month:Optional[int]=None,
               year:Optional[int]=None,search:Optional[str]=None,
               driver:Optional[str]=None,limit:Optional[int]=None,x_token:str=Header("")):
    require_any_tab_access(x_token,["consolidado","nf","pr","avulso","avaria","grafico","produtividade","clientes","pendentes"])
    conn=get_db(); sql="SELECT * FROM sales WHERE 1=1"; args=[]
    if sale_type: sql+=" AND sale_type=?"; args.append(sale_type)
    if year: sql+=" AND strftime('%Y',sale_date)=?"; args.append(str(year))
    if month: sql+=" AND strftime('%m',sale_date)=?"; args.append(f"{month:02d}")
    if driver: sql+=" AND delivery_person=?"; args.append(driver)
    if search:
        sql+=" AND (client LIKE ? OR product LIKE ? OR nf_number LIKE ? OR notes LIKE ? OR plate LIKE ?)"
        args+=[f"%{search}%"]*5
    sql+=" ORDER BY sale_date DESC, sale_time DESC"
    # Limite sÃ³ aplicado quando explicitamente solicitado (default era 2000 hardcoded)
    if limit is not None and limit > 0:
        sql+=" LIMIT ?"; args.append(int(limit))
    rows=conn.execute(sql,args).fetchall(); conn.close()
    return [dict(r) for r in rows]

@app.post("/api/sales")
def create_sale(sale:SaleIn,x_token:str=Header("")):
    sess=require_editor_tab_access(x_token,["consolidado","nf","pr","avulso","avaria","pendentes"])
    total=sale.total if sale.total is not None else sale.quantity*sale.unit_price
    new_id=str(uuid.uuid4()); now=datetime.now()
    # Normaliza nome do produto (unifica "Alho 250g"/"ALHO 250G", Macaxeira, etc).
    # ImportaÃ§Ãµes de Excel jÃ¡ aplicavam; criaÃ§Ã£o manual via form nÃ£o â€” corrigido aqui.
    product_norm = norm_p(sale.product, sale.sale_type) if sale.product else sale.product
    conn=get_db()
    conn.execute("""INSERT INTO sales(id,sale_type,sale_date,sale_time,client,product,nf_number,
        quantity,unit_price,total,notes,delivery_person,plate,source,created_by,created_at)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (new_id,sale.sale_type,sale.sale_date,sale.sale_time or now.strftime("%H:%M"),
         sale.client,product_norm,sale.nf_number,sale.quantity,sale.unit_price,total,
         sale.notes,sale.delivery_person,sale.plate,sale.source,sess["username"],now.isoformat()))
    import json as _json
    log_action(conn,sess,"CREATE_SALE","",product_norm or "",
               "venda_criada","",
               _json.dumps({"tipo":sale.sale_type,"data":sale.sale_date,"hora":sale.sale_time or "",
                 "cliente":sale.client or "","produto":product_norm or "",
                 "nf":sale.nf_number or "","placa":sale.plate or "",
                 "entregador":sale.delivery_person or "",
                 "qt":sale.quantity,"p_unit":sale.unit_price,"total":total},ensure_ascii=False),
               sale.sale_date,"")
    conn.commit(); clear_sales_cache(); row=conn.execute("SELECT * FROM sales WHERE id=?",(new_id,)).fetchone()
    conn.close(); return dict(row)

@app.put("/api/sales/bulk-delivered")
def bulk_delivery_status(body:dict,x_token:str=Header("")):
    """Marca mÃºltiplas notas como entregues de uma vez."""
    sess=require_editor_tab_access(x_token,["nf","pendentes","consolidado"]); conn=get_db()
    ids=body.get("ids",[])
    status=body.get("delivered","sim")
    delivered_at=body.get("delivered_at")
    if not ids: conn.close(); return {"ok":True,"updated":0}
    placeholders=",".join(["?"]*len(ids))
    if status is None or status=='':
        conn.execute(f"UPDATE sales SET delivered=NULL, delivered_at=NULL WHERE id IN ({placeholders})",ids)
    else:
        conn.execute(f"UPDATE sales SET delivered=?, delivered_at=? WHERE id IN ({placeholders})",[status,delivered_at]+ids)
    n=conn.execute("SELECT changes()").fetchone()[0]
    conn.commit(); clear_sales_cache(); conn.close(); return {"ok":True,"updated":n}

@app.get("/api/sales/delivery-sync")
def delivery_sync(x_token:str=Header("")):
    """Retorna apenas id + status de entrega de todas as vendas (endpoint leve para polling)."""
    require_any_tab_access(x_token,["nf","pendentes","consolidado"]); conn=get_db()
    rows=conn.execute("SELECT id, delivered, delivered_at FROM sales").fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.delete("/api/sales/clear-imports")
def clear_imports(x_token:str=Header("")):
    sess=require_editor_tab_access(x_token,["cfg_importar","consolidado"]); conn=get_db()
    conn.execute("DELETE FROM sales WHERE source='excel'"); conn.commit(); clear_sales_cache(); conn.close()
    return {"ok":True}

@app.put("/api/sales/{sale_id}")
def update_sale(sale_id:str,body:dict,x_token:str=Header("")):
    sess=require_editor_tab_access(x_token,["consolidado","nf","pr","avulso","avaria","pendentes"]); conn=get_db()
    try:
        old=conn.execute("SELECT * FROM sales WHERE id=?",(sale_id,)).fetchone()
        if not old: raise HTTPException(404,"Venda nÃ£o encontrada.")
        fields=["sale_date","sale_time","client","product","nf_number","quantity",
                "unit_price","total","notes","delivery_person","plate","sale_type",
                "delivered","delivered_at"]
        # Se produto estÃ¡ sendo alterado, normaliza com o sale_type vigente
        if "product" in body and body["product"]:
            st = body.get("sale_type", old["sale_type"]) or "NF"
            body["product"] = norm_p(body["product"], st)
        for f in fields:
            if f in body:
                conn.execute(f"UPDATE sales SET {f}=? WHERE id=?",(body[f],sale_id))
        log_action(conn,sess,"EDIT_SALE","","",str(old["product"]),"venda editada","",str(old["sale_date"]),"")
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    clear_sales_cache(); return {"ok":True}

@app.put("/api/sales/{sale_id}/delivered")
def update_delivery_status(sale_id:str,body:dict,x_token:str=Header("")):
    """Marca/desmarca nota como entregue."""
    sess=require_editor_tab_access(x_token,["nf","pendentes","consolidado"]); conn=get_db()
    status=body.get("delivered")  # 'sim', 'nao', ou None
    delivered_at=body.get("delivered_at")
    if status is None or status=='':
        conn.execute("UPDATE sales SET delivered=NULL, delivered_at=NULL WHERE id=?",(sale_id,))
    else:
        conn.execute("UPDATE sales SET delivered=?, delivered_at=? WHERE id=?",(status, delivered_at, sale_id))
    conn.commit(); clear_sales_cache(); conn.close(); return {"ok":True}

@app.delete("/api/sales/{sale_id}")
def delete_sale(sale_id:str,x_token:str=Header("")):
    sess=require_editor_tab_access(x_token,["consolidado","nf","pr","avulso","avaria","pendentes"]); conn=get_db()
    old=conn.execute("SELECT * FROM sales WHERE id=?",(sale_id,)).fetchone()
    if old:
        import json as _json
        detalhe=_json.dumps({
            "tipo":old["sale_type"],"data":old["sale_date"],"hora":old["sale_time"] or "",
            "cliente":old["client"] or "","produto":old["product"] or "",
            "nf":old["nf_number"] or "","placa":old["plate"] or "",
            "entregador":old["delivery_person"] or "",
            "qt":old["quantity"],"p_unit":old["unit_price"],"total":old["total"],
            "obs":old["notes"] or ""
        },ensure_ascii=False)
        log_action(conn,sess,"DELETE_SALE","",str(old["product"] or ""),
                   "venda_excluida",detalhe,"",str(old["sale_date"]),"")
    conn.execute("DELETE FROM sales WHERE id=?",(sale_id,)); conn.commit(); clear_sales_cache(); conn.close()
    return {"ok":True}

@lru_cache(maxsize=64)
def _summary_cached(company:str,year:int,bucket:int):
    conn=get_db(company)
    rows=conn.execute("""SELECT strftime('%m',sale_date) AS month,sale_type,
               SUM(total) AS total_val,SUM(quantity) AS total_qty
               FROM sales WHERE strftime('%Y',sale_date)=? GROUP BY month,sale_type ORDER BY month""",
               (str(year),)).fetchall()
    conn.close(); return tuple(tuple(dict(r).items()) for r in rows)

def summary(year:int=datetime.now().year,x_token:str=Header("")):
    require_any_tab_access(x_token,["consolidado","grafico","produtividade"])
    data=_summary_cached(_company_key(CURRENT_COMPANY.get()),int(year),_cache_bucket(60))
    return [dict(items) for items in data]

@lru_cache(maxsize=64)
def _consolidado_cached(company:str,year:int,bucket:int):
    conn=get_db(company)
    rows=conn.execute("""SELECT strftime('%m',sale_date) AS month,sale_type,product,
               SUM(total) AS total_val,SUM(quantity) AS total_qty,COUNT(*) AS count
               FROM sales WHERE strftime('%Y',sale_date)=?
               GROUP BY month,sale_type,product ORDER BY month,sale_type,product""",
               (str(year),)).fetchall()
    grouped={}
    for r in rows:
        product=norm_p(r["product"] or "", r["sale_type"] or "NF") if r["product"] else r["product"]
        key=(r["month"],r["sale_type"],product)
        if key not in grouped:
            grouped[key]={"month":r["month"],"sale_type":r["sale_type"],"product":product,"total_val":0,"total_qty":0,"count":0}
        grouped[key]["total_val"]+=float(r["total_val"] or 0)
        grouped[key]["total_qty"]+=float(r["total_qty"] or 0)
        grouped[key]["count"]+=int(r["count"] or 0)
    conn.close(); return tuple(tuple(row.items()) for row in grouped.values())

@app.get("/api/consolidado")
def consolidado(year:int=datetime.now().year,x_token:str=Header("")):
    require_any_tab_access(x_token,["consolidado","grafico","produtividade"])
    data=_consolidado_cached(_company_key(CURRENT_COMPANY.get()),int(year),_cache_bucket(60))
    return [dict(items) for items in data]

@app.get("/api/projecao/producao")
def projecao_producao(days:int=30,
                      product:str="",
                      source:str="",
                      view:str="consolidada",
                      start:Optional[str]=None,
                      end:Optional[str]=None,
                      x_token:str=Header("")):
    """Projecao de saida por produto com base no historico real de vendas."""
    require_auth(x_token)
    days = max(1, min(int(days or 30), 365))
    today = date.today()
    if start and end:
        try:
            start_dt = datetime.strptime(start[:10], "%Y-%m-%d").date()
            end_dt = datetime.strptime(end[:10], "%Y-%m-%d").date()
        except Exception:
            raise HTTPException(400, "Periodo personalizado invalido.")
        if end_dt < start_dt:
            start_dt, end_dt = end_dt, start_dt
        days = max(1, (end_dt - start_dt).days + 1)
    else:
        end_dt = today
        start_dt = today - timedelta(days=days-1)
    prev_end = start_dt - timedelta(days=1)
    prev_start = prev_end - timedelta(days=days-1)

    def source_clause(src):
        src = (src or "").strip().lower()
        if src == "cliente":
            return " AND COALESCE(client,'')<>''", []
        if src == "entregador":
            return " AND COALESCE(delivery_person,'')<>''", []
        if src == "tipo":
            return " AND COALESCE(sale_type,'')<>''", []
        if src == "origem":
            return " AND COALESCE(source,'')<>''", []
        return "", []

    conn=get_db()
    sql = """SELECT sale_date,sale_type,product,client,delivery_person,source,
                    SUM(COALESCE(quantity,0)) AS qty,
                    SUM(COALESCE(total,0)) AS total,
                    COUNT(*) AS regs
             FROM sales
             WHERE sale_type<>'AVARIA'
               AND date(sale_date) BETWEEN date(?) AND date(?)"""
    args = [start_dt.isoformat(), end_dt.isoformat()]
    extra, extra_args = source_clause(source)
    sql += extra
    args += extra_args
    if product.strip():
        sql += " AND product LIKE ?"
        args.append(f"%{product.strip()}%")
    sql += " GROUP BY sale_date,sale_type,product,client,delivery_person,source ORDER BY sale_date"
    rows = conn.execute(sql, args).fetchall()

    psql = """SELECT product,SUM(COALESCE(quantity,0)) AS qty
              FROM sales
              WHERE sale_type<>'AVARIA'
                AND date(sale_date) BETWEEN date(?) AND date(?)"""
    pargs = [prev_start.isoformat(), prev_end.isoformat()]
    pextra, pextra_args = source_clause(source)
    psql += pextra
    pargs += pextra_args
    if product.strip():
        psql += " AND product LIKE ?"
        pargs.append(f"%{product.strip()}%")
    psql += " GROUP BY product"
    prev_rows = conn.execute(psql, pargs).fetchall()
    conn.close()

    prev_by_product = {}
    for r in prev_rows:
        pname = norm_p(r["product"] or "", "NF") if r["product"] else "Sem produto"
        prev_by_product[pname] = prev_by_product.get(pname, 0.0) + float(r["qty"] or 0)

    phase_defs = {
        "pico": {"label": "Pico 1-5", "range": "dias 1 a 5"},
        "meio": {"label": "Meio 6-14", "range": "dias 6 a 14"},
        "fim": {"label": "Fim 15+", "range": "dia 15 em diante"},
    }

    def phase_for_date(day_value):
        if isinstance(day_value, date):
            day_num = day_value.day
        else:
            try:
                day_num = datetime.strptime(str(day_value)[:10], "%Y-%m-%d").date().day
            except Exception:
                day_num = 1
        if day_num <= 5:
            return "pico", phase_defs["pico"]["label"]
        if day_num <= 14:
            return "meio", phase_defs["meio"]["label"]
        return "fim", phase_defs["fim"]["label"]

    daily_product = {}
    product_sources = defaultdict(lambda: defaultdict(float))
    product_regs = defaultdict(int)
    product_value = defaultdict(float)
    daily = {}
    fontes = {}
    movement_days = set()
    for r in rows:
        pname = norm_p(r["product"] or "", r["sale_type"] or "NF") if r["product"] else "Sem produto"
        qty = float(r["qty"] or 0)
        val = float(r["total"] or 0)
        day = str(r["sale_date"] or "")[:10]
        if qty:
            movement_days.add(day)
        key = (pname, day)
        if key not in daily_product:
            daily_product[key] = {"produto": pname, "data": day, "quantidade": 0.0}
        daily_product[key]["quantidade"] += qty
        product_value[pname] += val
        product_regs[pname] += int(r["regs"] or 0)
        fonte = (r["client"] if source == "cliente" else
                 r["delivery_person"] if source == "entregador" else
                 r["sale_type"] if source == "tipo" else
                 r["source"] if source == "origem" else
                 (r["client"] or r["delivery_person"] or r["sale_type"] or r["source"] or "Geral"))
        product_sources[pname][fonte or "Geral"] += qty
        daily.setdefault(day, defaultdict(float))[pname] += qty
        fontes[fonte or "Geral"] = fontes.get(fonte or "Geral", 0.0) + qty

    products = {}
    for item in daily_product.values():
        pname = item["produto"]
        if pname not in products:
            products[pname] = {"produto": pname, "quantidade": 0.0, "dias_movimento": set()}
        products[pname]["quantidade"] += float(item["quantidade"] or 0)
        if float(item["quantidade"] or 0) > 0:
            products[pname]["dias_movimento"].add(item["data"])

    phase_stats = defaultdict(lambda: {
        "pico": {"qty": 0.0, "days": set()},
        "meio": {"qty": 0.0, "days": set()},
        "fim": {"qty": 0.0, "days": set()},
    })
    cycle_totals = {
        "pico": {"qty": 0.0, "days": set()},
        "meio": {"qty": 0.0, "days": set()},
        "fim": {"qty": 0.0, "days": set()},
    }
    for item in daily_product.values():
        qty = float(item["quantidade"] or 0)
        if qty <= 0:
            continue
        phase_key, _ = phase_for_date(item["data"])
        pname = item["produto"]
        phase_stats[pname][phase_key]["qty"] += qty
        phase_stats[pname][phase_key]["days"].add(item["data"])
        cycle_totals[phase_key]["qty"] += qty
        cycle_totals[phase_key]["days"].add(item["data"])

    movement_day_count = len(movement_days)
    period_days = max(1, days)
    current_phase_key, current_phase_label = phase_for_date(end_dt)
    product_list = []
    for pname, p in products.items():
        qtd = float(p["quantidade"])
        avg_day = qtd / period_days
        avg_week = avg_day * 7
        prev_qty = float(prev_by_product.get(pname, 0))
        variation = ((qtd - prev_qty) / prev_qty * 100) if prev_qty else (100.0 if qtd else 0.0)
        fonte_principal = max(product_sources[pname].items(), key=lambda x: x[1])[0] if product_sources[pname] else "Geral"
        risk = "estavel"
        if variation > 30 and avg_day > 0:
            risk = "ruptura"
        elif variation < -20:
            risk = "queda"
        special = pname.lower() in ("macaxeira a vacuo", "macaxeira a vÃ¡cuo", "macaxeira com casca (kg)", "macaxeira com casca kg")
        p_phases = phase_stats[pname]
        phase_avg = {}
        phase_payload = {}
        for phase_key in ("pico", "meio", "fim"):
            phase_qty = float(p_phases[phase_key]["qty"] or 0)
            phase_days = len(p_phases[phase_key]["days"])
            phase_avg[phase_key] = (phase_qty / phase_days) if phase_days else 0.0
            phase_payload[phase_key] = {
                "label": phase_defs[phase_key]["label"],
                "range": phase_defs[phase_key]["range"],
                "quantidade": round(phase_qty, 2),
                "dias": phase_days,
                "media": round(phase_avg[phase_key], 2),
            }

        def projected_by_cycle(total_days):
            projected = 0.0
            for offset in range(1, total_days + 1):
                future_day = end_dt + timedelta(days=offset)
                future_phase, _ = phase_for_date(future_day)
                projected += phase_avg.get(future_phase, 0.0) or avg_day
            return projected

        main_phase_key = max(("pico", "meio", "fim"), key=lambda key: (p_phases[key]["qty"], len(p_phases[key]["days"])))
        main_phase_label = phase_defs[main_phase_key]["label"] if p_phases[main_phase_key]["qty"] > 0 else "Sem perfil"
        current_phase_avg = phase_avg.get(current_phase_key, 0.0) or avg_day
        proj_7 = projected_by_cycle(7)
        proj_15 = projected_by_cycle(15)
        proj_30 = projected_by_cycle(30)
        product_list.append({
            "produto": pname,
            "quantidade_vendida": round(qtd, 2),
            "quantidade": round(qtd, 2),
            "valor": round(float(product_value[pname]), 2),
            "media_dia": round(avg_day, 2),
            "media_semana": round(avg_week, 2),
            "media_fase_atual": round(current_phase_avg, 2),
            "proj_7d": round(proj_7, 2),
            "proj_7": round(proj_7, 2),
            "proj_15d": round(proj_15, 2),
            "proj_15": round(proj_15, 2),
            "proj_30d": round(proj_30, 2),
            "proj_30": round(proj_30, 2),
            "comparacao_anterior": round(variation, 2),
            "fonte_principal": fonte_principal,
            "status": risk,
            "destaque": special,
            "dias_movimento": len(p["dias_movimento"]),
            "registros": int(product_regs[pname]),
            "fase_atual": current_phase_label,
            "fase_principal": main_phase_label,
            "perfil_compra": phase_payload
        })
    product_list.sort(key=lambda x: (0 if x["destaque"] else 1, -x["quantidade"], x["produto"]))

    top = sorted(product_list, key=lambda x: x["quantidade"], reverse=True)
    maior_variacao = max(product_list, key=lambda x: x["comparacao_anterior"], default=None)
    maior_risco = next((p for p in top if p["status"] == "ruptura"), top[0] if top else None)
    total_qty = sum(float(p["quantidade"] or 0) for p in product_list)
    total_value = sum(float(product_value[p["produto"]] or 0) for p in product_list)
    media_total_dia = total_qty / period_days
    total_proj_7 = sum(float(p["proj_7"] or 0) for p in product_list)
    total_proj_15 = sum(float(p["proj_15"] or 0) for p in product_list)
    total_proj_30 = sum(float(p["proj_30"] or 0) for p in product_list)
    ciclo_compra = {}
    for phase_key in ("pico", "meio", "fim"):
        phase_qty = float(cycle_totals[phase_key]["qty"] or 0)
        phase_days = len(cycle_totals[phase_key]["days"])
        ciclo_compra[phase_key] = {
            "label": phase_defs[phase_key]["label"],
            "range": phase_defs[phase_key]["range"],
            "quantidade": round(phase_qty, 2),
            "dias": phase_days,
            "media": round((phase_qty / phase_days) if phase_days else 0.0, 2),
            "atual": phase_key == current_phase_key,
        }
    mac_products = {"Macaxeira a VÃ¡cuo", "Macaxeira com Casca (KG)"}
    mac_rows = [p for p in product_list if p["produto"] in mac_products]
    mac_qty = sum(float(p["quantidade"] or 0) for p in mac_rows)
    mac_days = len(set(item["data"] for item in daily_product.values() if item["produto"] in mac_products and float(item["quantidade"] or 0) > 0))
    mac_avg = mac_qty / period_days
    mac_phase_avg = sum(float(p.get("media_fase_atual") or 0) for p in mac_rows)
    mac_proj_7 = sum(float(p.get("proj_7") or 0) for p in mac_rows)
    mac_proj_15 = sum(float(p.get("proj_15") or 0) for p in mac_rows)
    mac_proj_30 = sum(float(p.get("proj_30") or 0) for p in mac_rows)
    macaxeira_consolidado = {
        "produto": "Macaxeira consolidada",
        "quantidade_vendida": round(mac_qty, 2),
        "quantidade": round(mac_qty, 2),
        "dias_movimento": mac_days,
        "media_dia": round(mac_avg, 2),
        "media_semana": round(mac_avg * 7, 2),
        "media_fase_atual": round(mac_phase_avg, 2),
        "proj_7d": round(mac_proj_7, 2),
        "proj_7": round(mac_proj_7, 2),
        "proj_15d": round(mac_proj_15, 2),
        "proj_15": round(mac_proj_15, 2),
        "proj_30d": round(mac_proj_30, 2),
        "proj_30": round(mac_proj_30, 2),
        "produtos": mac_rows
    }
    resumo = {
        "periodo": {"inicio": start_dt.isoformat(), "fim": end_dt.isoformat(), "dias": max(1, days), "dias_movimento": movement_day_count,
                    "anterior_inicio": prev_start.isoformat(), "anterior_fim": prev_end.isoformat()},
        "total_vendido": round(total_qty, 2),
        "valor_total": round(total_value, 2),
        "media_dia": round(media_total_dia, 2),
        "media_semana": round(media_total_dia * 7, 2),
        "proj_7": round(total_proj_7, 2),
        "proj_15": round(total_proj_15, 2),
        "proj_30": round(total_proj_30, 2),
        "fase_atual": current_phase_label,
        "ciclo_compra": ciclo_compra,
        "produto_maior_saida": top[0] if top else None,
        "produto_maior_variacao": maior_variacao,
        "produto_maior_risco": maior_risco,
        "macaxeira_consolidado": macaxeira_consolidado,
        "fonte_principal": max(fontes.items(), key=lambda x: x[1])[0] if fontes else "Geral"
    }
    trend_days = sorted(daily.keys())
    trend = [{"data": d, "total": round(sum(daily[d].values()), 2)} for d in trend_days]
    return {"resumo": resumo, "produtos": product_list, "tendencia": trend, "view": view}

@app.get("/api/avarias/client-risk")
def avarias_client_risk(year:int=datetime.now().year,
                        month:Optional[int]=None,
                        client:str="",
                        x_token:str=Header("")):
    """Painel analÃ­tico de risco de avarias por cliente.

    MantÃ©m a regra de negÃ³cio atual: vendas = tipos diferentes de AVARIA;
    avarias = sale_type AVARIA, sempre tratadas em valor absoluto para anÃ¡lise.
    """
    require_auth(x_token)
    conn=get_db()
    ym = f"{int(year)}-{int(month):02d}" if month else None
    client_clean = (client or "").strip()

    where_period = "strftime('%Y', sale_date)=?"
    args = [str(int(year))]
    if ym:
        where_period = "strftime('%Y-%m', sale_date)=?"
        args = [ym]

    client_sql = ""
    if client_clean:
        client_sql = " AND TRIM(client) COLLATE NOCASE = ?"
        args.append(client_clean)

    av_rows = conn.execute(f"""
        SELECT TRIM(COALESCE(client,'')) AS client,
               COALESCE(product,'') AS product,
               sale_date,
               SUM(ABS(COALESCE(total,0))) AS total_avaria,
               SUM(ABS(COALESCE(quantity,0))) AS qtd_avaria,
               COUNT(*) AS ocorrencias
        FROM sales
        WHERE sale_type='AVARIA'
          AND client IS NOT NULL AND TRIM(client)!=''
          AND {where_period}{client_sql}
        GROUP BY TRIM(client) COLLATE NOCASE, COALESCE(product,''), sale_date
    """, args).fetchall()

    sale_rows = conn.execute(f"""
        SELECT TRIM(COALESCE(client,'')) AS client,
               SUM(ABS(COALESCE(total,0))) AS total_vendas,
               COUNT(*) AS vendas
        FROM sales
        WHERE sale_type!='AVARIA'
          AND client IS NOT NULL AND TRIM(client)!=''
          AND {where_period}{client_sql}
        GROUP BY TRIM(client) COLLATE NOCASE
    """, args).fetchall()

    sale_map = {r["client"]: {"total": float(r["total_vendas"] or 0), "count": int(r["vendas"] or 0)} for r in sale_rows}
    clients = {}
    product_totals = {}
    for r in av_rows:
        cli = r["client"] or "Sem cliente"
        product = norm_p(r["product"] or "Sem produto", "AVARIA")
        if cli not in clients:
            clients[cli] = {
                "client": cli, "avTotal": 0.0, "avQty": 0.0, "avCount": 0,
                "days": set(), "products": {}
            }
        c = clients[cli]
        val = float(r["total_avaria"] or 0)
        qty = float(r["qtd_avaria"] or 0)
        cnt = int(r["ocorrencias"] or 0)
        c["avTotal"] += val
        c["avQty"] += qty
        c["avCount"] += cnt
        if r["sale_date"]:
            c["days"].add(r["sale_date"])
        if product not in c["products"]:
            c["products"][product] = {"product": product, "total": 0.0, "qty": 0.0, "count": 0}
        c["products"][product]["total"] += val
        c["products"][product]["qty"] += qty
        c["products"][product]["count"] += cnt
        if product not in product_totals:
            product_totals[product] = {"product": product, "total": 0.0, "qty": 0.0, "count": 0}
        product_totals[product]["total"] += val
        product_totals[product]["qty"] += qty
        product_totals[product]["count"] += cnt

    out_clients = []
    for cli, c in clients.items():
        sales = sale_map.get(cli, {"total": 0.0, "count": 0})
        rate = (c["avTotal"] / sales["total"] * 100) if sales["total"] else 0.0
        # Risco padrÃ£o: taxa primeiro; valor e ocorrÃªncias entram como desempate.
        score = (rate * 1000000) + (c["avTotal"] * 10) + c["avCount"]
        if rate >= 12 or c["avTotal"] >= 3000 or c["avCount"] >= 8:
            level, label = "high", "Alto risco"
        elif rate >= 5 or c["avTotal"] >= 1000 or c["avCount"] >= 3:
            level, label = "med", "AtenÃ§Ã£o"
        else:
            level, label = "low", "Controlado"
        out_clients.append({
            "client": cli,
            "avTotal": round(c["avTotal"], 2),
            "avQty": round(c["avQty"], 3),
            "avCount": c["avCount"],
            "days": len(c["days"]),
            "saleTotal": round(sales["total"], 2),
            "saleCount": sales["count"],
            "rate": round(rate, 2),
            "score": round(score, 2),
            "level": level,
            "levelLabel": label,
            "products": sorted(c["products"].values(), key=lambda p: p["total"], reverse=True)
        })
    out_clients.sort(key=lambda r: r["score"], reverse=True)

    trend_rows = conn.execute("""
        SELECT strftime('%m',sale_date) AS month,
               SUM(ABS(COALESCE(total,0))) AS total_avaria,
               COUNT(*) AS ocorrencias
        FROM sales
        WHERE sale_type='AVARIA'
          AND strftime('%Y', sale_date)=?
        GROUP BY strftime('%m',sale_date)
        ORDER BY month
    """, (str(int(year)),)).fetchall()
    trend = [dict(r) for r in trend_rows]

    total_avaria = sum(c["avTotal"] for c in clients.values())
    total_vendas = sum(v["total"] for v in sale_map.values())
    conn.close()
    return {
        "summary": {
            "total_avaria": round(total_avaria, 2),
            "total_vendas": round(total_vendas, 2),
            "taxa_geral": round((total_avaria / total_vendas * 100) if total_vendas else 0, 2),
            "clientes_afetados": len(out_clients),
            "ocorrencias": sum(c["avCount"] for c in clients.values())
        },
        "clients": out_clients,
        "products": sorted(product_totals.values(), key=lambda p: p["total"], reverse=True),
        "trend": trend
    }

@app.get("/api/product-stats")
def product_stats(product:str,year:Optional[int]=None,x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    target=norm_p(product or "", "NF")
    sql="""SELECT strftime('%m',sale_date) AS month,sale_type,product,total,quantity,client,unit_price
           FROM sales WHERE product IS NOT NULL AND product!=''"""
    args=[]
    if year: sql+=" AND strftime('%Y',sale_date)=?"; args.append(str(year))
    rows=conn.execute(sql,args).fetchall()
    monthly={}
    total_val=qty=transactions=unit_sum=0
    clients=set()
    for r in rows:
        if norm_p(r["product"] or "", r["sale_type"] or "NF") != target:
            continue
        m=r["month"]
        if m not in monthly:
            monthly[m]={"month":m,"total_val":0,"qty":0,"transactions":0,"clients":set(),"unit_sum":0}
        v=float(r["total"] or 0); q=float(r["quantity"] or 0); up=float(r["unit_price"] or 0)
        monthly[m]["total_val"]+=v; monthly[m]["qty"]+=q; monthly[m]["transactions"]+=1; monthly[m]["unit_sum"]+=up
        if r["client"]: monthly[m]["clients"].add(r["client"]); clients.add(r["client"])
        total_val+=v; qty+=q; transactions+=1; unit_sum+=up
    out=[]
    for m in sorted(monthly):
        row=monthly[m]
        out.append({"month":m,"total_val":row["total_val"],"qty":row["qty"],"transactions":row["transactions"],
                    "clients":len(row["clients"]),"avg_price":(row["unit_sum"]/row["transactions"] if row["transactions"] else 0)})
    conn.close()
    return {"monthly":out,"totals":{"total_val":total_val,"qty":qty,"transactions":transactions,
            "clients":len(clients),"avg_price":(unit_sum/transactions if transactions else 0)}}

# â”€â”€ PreÃ§os â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get("/api/prices")
def get_prices(x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    rows=conn.execute("SELECT key,label,price,price_min,price_max,is_variable,updated_at FROM product_prices WHERE COALESCE(active,1)=1 ORDER BY label").fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.post("/api/prices")
def add_product(body:dict,x_token:str=Header("")):
    sess=require_prices_access(x_token)
    key=body.get("key","").upper(); label=body.get("label",""); price=float(body.get("price",0))
    is_var=int(body.get("is_variable",False)); pmin=body.get("price_min"); pmax=body.get("price_max")
    if not key or not label: raise HTTPException(400,"Chave e nome obrigatÃ³rios.")
    conn=get_db()
    conn.execute("INSERT OR REPLACE INTO product_prices(key,label,price,price_min,price_max,is_variable) VALUES(?,?,?,?,?,?)",
                 (key,label,price,pmin,pmax,is_var))
    try:
        row=conn.execute("SELECT value FROM settings WHERE key='deleted_product_keys'").fetchone()
        deleted=set(json.loads(row["value"])) if row and row["value"] else set()
        if key in deleted:
            deleted.remove(key)
            conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('deleted_product_keys',?)",(json.dumps(sorted(deleted)),))
    except Exception:
        pass
    log_action(conn,sess,"ADD_PRODUCT",key,label,"price","",str(price),"","Produto adicionado")
    conn.commit(); conn.close(); clear_price_cache(); return {"ok":True}

@app.put("/api/prices")
def update_price(update:PriceUpdate,x_token:str=Header("")):
    sess=require_prices_access(x_token); conn=get_db()
    try:
        old=conn.execute("SELECT price,label FROM product_prices WHERE key=?",(update.key,)).fetchone()
        old_price=old["price"] if old else 0; label=old["label"] if old else update.key
        today=datetime.now().strftime("%Y-%m-%d")
        conn.execute("UPDATE product_prices SET price=?,price_min=?,price_max=?,updated_at=datetime('now') WHERE key=?",
                     (update.price,update.price_min,update.price_max,update.key))
        conn.execute("INSERT INTO price_history(id,key,price,effective_date,note,changed_by) VALUES(?,?,?,?,?,?)",
                     (str(uuid.uuid4()),update.key,update.price,today,"Alterado via painel",sess["username"]))
        log_action(conn,sess,"PRICE_CHANGE",update.key,label,"price",str(old_price),str(update.price),today,"Alterado via painel")
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    try:
        clear_price_cache()
    except Exception:
        logger.exception("Falha ao limpar cache de precos apos atualizacao")
    return {"ok":True}

@app.put("/api/prices/label")
def update_label(body:dict,x_token:str=Header("")):
    sess=require_prices_access(x_token)
    key=body.get("key",""); label=body.get("label","")
    if not key or not label: raise HTTPException(400,"Dados invÃ¡lidos.")
    conn=get_db()
    old=conn.execute("SELECT label FROM product_prices WHERE key=?",(key,)).fetchone()
    conn.execute("UPDATE product_prices SET label=? WHERE key=?",(label,key))
    log_action(conn,sess,"RENAME_PRODUCT",key,label,"label",old["label"] if old else "","label",datetime.now().strftime("%Y-%m-%d"),"")
    conn.commit(); conn.close(); return {"ok":True}

@app.delete("/api/prices/{key}")
def delete_product(key:str,x_token:str=Header("")):
    """Remove produto da lista de preÃ§os de verdade e impede resemeadura automÃ¡tica."""
    sess=require_prices_access(x_token)
    conn=get_db()
    try:
        row=conn.execute("SELECT value FROM settings WHERE key='deleted_product_keys'").fetchone()
        deleted=set(json.loads(row["value"])) if row and row["value"] else set()
        deleted.add(key)
        conn.execute("DELETE FROM product_prices WHERE key=?",(key,))
        conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('deleted_product_keys',?)",(json.dumps(sorted(deleted)),))
        log_action(conn,sess,"DELETE_PRODUCT",key,"","active","1","deleted",datetime.now().strftime("%Y-%m-%d"),"Produto removido da lista de preÃ§os")
        conn.commit()
    except Exception as e:
        conn.rollback(); conn.close()
        raise HTTPException(500,f"Erro ao remover produto: {e}")
    conn.close(); clear_price_cache()
    return {"ok":True,"deleted":True}

@app.post("/api/avaria/parse")
def parse_avaria_preview(body:dict,x_token:str=Header("")):
    require_auth(x_token)
    text=body.get("text",""); prices=load_prices()
    val,items=parse_avaria_text(text,prices)
    return {"total":val,"items":items,"text":text}

# â”€â”€ HistÃ³rico de preÃ§os / Auditoria â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get("/api/prices/history")
def get_price_history(key:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    if key:
        rows=conn.execute("SELECT * FROM price_history WHERE key=? ORDER BY effective_date DESC",(key,)).fetchall()
    else:
        rows=conn.execute("SELECT * FROM price_history ORDER BY effective_date DESC,key LIMIT 300").fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.post("/api/prices/history")
def add_price_history(body:dict,x_token:str=Header("")):
    sess=require_prices_access(x_token)
    key=body.get("key",""); price=float(body.get("price",0))
    eff=body.get("effective_date",datetime.now().strftime("%Y-%m-%d")); note=body.get("note","")
    if not key or price<0: raise HTTPException(400,"Dados invÃ¡lidos.")
    old_price=get_price_on_date(key,eff)
    conn=get_db()
    try:
        label_row=conn.execute("SELECT label FROM product_prices WHERE key=?",(key,)).fetchone()
        label=label_row["label"] if label_row else key
        conn.execute("INSERT INTO price_history(id,key,price,effective_date,note,changed_by) VALUES(?,?,?,?,?,?)",
                     (str(uuid.uuid4()),key,price,eff,note,sess["username"]))
        if eff<=datetime.now().strftime("%Y-%m-%d"):
            conn.execute("UPDATE product_prices SET price=?,updated_at=datetime('now') WHERE key=?",(price,key))
        log_action(conn,sess,"PRICE_HISTORY",key,label,"price",str(old_price),str(price),eff,note)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    try:
        clear_price_cache()
    except Exception:
        logger.exception("Falha ao limpar cache de precos apos historico")
    return {"ok":True}

@app.delete("/api/prices/history/{entry_id}")
def delete_price_history(entry_id:str,x_token:str=Header("")):
    require_prices_access(x_token); conn=get_db()
    conn.execute("DELETE FROM price_history WHERE id=?",(entry_id,)); conn.commit(); conn.close(); clear_price_cache()
    return {"ok":True}

@app.get("/api/audit-log")
def get_audit_log(x_token:str=Header("")):
    require_admin(x_token); conn=get_db()
    rows=conn.execute("SELECT * FROM audit_log ORDER BY timestamp DESC LIMIT 1000").fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.get("/api/drivers/list")
def get_drivers_list(x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    rows=conn.execute("""
        SELECT name AS delivery_person
        FROM drivers
        WHERE active=1
        UNION
        SELECT DISTINCT delivery_person AS delivery_person
        FROM sales
        WHERE delivery_person IS NOT NULL AND TRIM(delivery_person)!=''
    """).fetchall()
    stats_rows=conn.execute("""
        SELECT delivery_person, COUNT(*) AS entregas, SUM(total) AS total_val
        FROM sales
        WHERE delivery_person IS NOT NULL AND TRIM(delivery_person)!=''
        GROUP BY delivery_person
    """).fetchall()
    conn.close()
    stats={r["delivery_person"]:dict(r) for r in stats_rows}
    merged={}
    for r in rows:
        name=(r["delivery_person"] or "").strip()
        if not name:
            continue
        item=merged.setdefault(name,{"delivery_person":name,"entregas":0,"total_val":0})
        if name in stats:
            item["entregas"]=stats[name].get("entregas") or 0
            item["total_val"]=stats[name].get("total_val") or 0
    return sorted(merged.values(), key=lambda r: str(r["delivery_person"] or "").lower())

@app.post("/api/drivers")
def create_driver(body:dict,x_token:str=Header("")):
    require_prices_access(x_token)
    name=(body.get("name") or "").strip()
    if not name:
        raise HTTPException(400,"Nome obrigatorio.")
    conn=get_db()
    try:
        existing=conn.execute("SELECT id FROM drivers WHERE lower(name)=lower(?)",(name,)).fetchone()
        if existing:
            conn.execute("UPDATE drivers SET name=?, active=1, updated_at=datetime('now') WHERE id=?",(name,existing["id"]))
        else:
            conn.execute("INSERT INTO drivers(id,name,active) VALUES(?,?,1)",(str(uuid.uuid4()),name))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.rollback()
        raise HTTPException(409,"Entregador ja cadastrado.")
    finally:
        conn.close()
    return {"ok":True,"driver":name}

@app.get("/api/vehicles")
def list_vehicles(x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    rows=conn.execute("SELECT * FROM vehicles WHERE active=1 ORDER BY name").fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.post("/api/vehicles")
def create_vehicle(body:dict,x_token:str=Header("")):
    require_prices_access(x_token)
    name=body.get("name","").strip(); plate=body.get("plate","").strip().upper()
    driver=body.get("driver","").strip(); notes=body.get("notes","").strip()
    if not name or not plate: raise HTTPException(400,"Nome e placa obrigatÃ³rios.")
    conn=get_db()
    conn.execute("INSERT INTO vehicles(id,name,plate,driver,notes) VALUES(?,?,?,?,?)",
                 (str(uuid.uuid4()),name,plate,driver or None,notes or None))
    conn.commit(); conn.close(); return {"ok":True}

@app.put("/api/vehicles/{vid}")
def update_vehicle(vid:str,body:dict,x_token:str=Header("")):
    require_prices_access(x_token); conn=get_db()
    for f in ["name","plate","driver","notes","active"]:
        if f in body:
            val=body[f].strip().upper() if f=="plate" and body[f] else body[f]
            conn.execute(f"UPDATE vehicles SET {f}=? WHERE id=?",(val,vid))
    conn.commit(); conn.close(); return {"ok":True}

@app.delete("/api/vehicles/{vid}")
def delete_vehicle(vid:str,x_token:str=Header("")):
    require_prices_access(x_token); conn=get_db()
    conn.execute("DELETE FROM vehicles WHERE id=?",(vid,)); conn.commit(); conn.close()
    return {"ok":True}


@app.post("/api/drivers/rename")
def rename_driver(body:dict,x_token:str=Header("")):
    require_prices_access(x_token)
    old_name=body.get("old_name","").strip()
    new_name=body.get("new_name","").strip()
    if not old_name or not new_name: raise HTTPException(400,"Nomes obrigatÃ³rios.")
    conn=get_db()
    try:
        existing=conn.execute("SELECT id FROM drivers WHERE lower(name)=lower(?)",(old_name,)).fetchone()
        if existing:
            conn.execute("UPDATE drivers SET name=?, active=1, updated_at=datetime('now') WHERE id=?",(new_name,existing["id"]))
        else:
            conn.execute("INSERT OR IGNORE INTO drivers(id,name,active) VALUES(?,?,1)",(str(uuid.uuid4()),new_name))
        n=conn.execute("UPDATE sales SET delivery_person=? WHERE delivery_person=?",(new_name,old_name)).rowcount
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return {"ok":True,"updated":n}


@app.delete("/api/drivers/{name}")
def delete_driver(name:str,x_token:str=Header("")):
    """Remove o entregador de todas as vendas (define delivery_person=NULL)."""
    require_prices_access(x_token)
    name = (name or "").strip()
    if not name: raise HTTPException(400,"Nome obrigatÃ³rio.")
    conn=get_db()
    try:
        conn.execute("UPDATE drivers SET active=0, updated_at=datetime('now') WHERE name=?",(name,))
        n=conn.execute("UPDATE sales SET delivery_person=NULL WHERE delivery_person=?",(name,)).rowcount
        conn.commit()
    except Exception as e:
        conn.rollback(); conn.close()
        raise HTTPException(500,f"Erro ao remover entregador: {e}")
    conn.close()
    return {"ok":True,"removed":n,"driver":name}


@app.get("/api/import-log")
def import_log(x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    rows=conn.execute("SELECT * FROM import_log ORDER BY imported_at DESC LIMIT 50").fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.get("/api/server-info")
def server_info():
    """InformaÃ§Ãµes do servidor para acesso em rede."""
    hostname=socket.gethostname()
    try: local_ip=socket.gethostbyname(hostname)
    except Exception: local_ip="127.0.0.1"
    port=int(os.environ.get("PORT",8765))
    return {"local_ip":local_ip,"port":port,"local_url":f"http://{local_ip}:{port}",
            "hostname":hostname}

# â”€â”€ Clientes â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get("/api/clients")
def list_clients(search:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    sql="SELECT * FROM clients WHERE active=1"
    args=[]
    if search:
        sql+=" AND (name LIKE ? OR cnpj LIKE ? OR phone LIKE ? OR email LIKE ?)"
        args+=[f"%{search}%"]*4
    sql+=" ORDER BY name"
    rows=conn.execute(sql,args).fetchall(); conn.close()
    return [dict(r) for r in rows]

@app.post("/api/clients")
def create_client(body:ClientIn,x_token:str=Header("")):
    sess=require_editor(x_token)
    new_id=str(uuid.uuid4()); conn=get_db()
    # Normalize CNPJ - remove formatting
    cnpj=re.sub(r'[^0-9]','',body.cnpj or '')
    conn.execute("""INSERT INTO clients(id,name,cnpj,cpf,phone,email,address,city,notes,created_by)
                    VALUES(?,?,?,?,?,?,?,?,?,?)""",
                 (new_id,body.name.strip(),cnpj or None,body.cpf or None,
                  body.phone or None,body.email or None,body.address or None,
                  body.city or None,body.notes or None,sess["username"]))
    conn.commit(); row=conn.execute("SELECT * FROM clients WHERE id=?",(new_id,)).fetchone()
    conn.close(); return dict(row)

@app.put("/api/clients/{client_id}")
def update_client(client_id:str,body:dict,x_token:str=Header("")):
    require_editor(x_token); conn=get_db()
    fields=["name","cnpj","cpf","phone","email","address","city","notes","active"]
    for f in fields:
        if f in body:
            val=re.sub(r'[^0-9]','',body[f]) if f=="cnpj" and body[f] else body[f]
            conn.execute(f"UPDATE clients SET {f}=? WHERE id=?",(val,client_id))
    conn.commit(); conn.close(); return {"ok":True}

@app.delete("/api/clients/{client_id}")
def delete_client(client_id:str,x_token:str=Header("")):
    require_editor(x_token); conn=get_db()
    conn.execute("UPDATE clients SET active=0 WHERE id=?",(client_id,))
    conn.commit(); conn.close(); return {"ok":True}

@app.get("/api/clients/{client_id}/stats")
def client_stats(client_id:str,x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    cl=conn.execute("SELECT * FROM clients WHERE id=?",(client_id,)).fetchone()
    if not cl: conn.close(); raise HTTPException(404,"Cliente nÃ£o encontrado.")
    # Match sales by client name
    name=cl["name"]
    stats=conn.execute("""SELECT COUNT(*) as total_sales,
        SUM(total) as total_value, AVG(total) as avg_ticket,
        MAX(sale_date) as last_purchase, MIN(sale_date) as first_purchase,
        COUNT(DISTINCT strftime('%Y-%m',sale_date)) as active_months
        FROM sales WHERE client = ? AND sale_type!='AVARIA'""",
        (name,)).fetchone()
    by_type=conn.execute("""SELECT sale_type, SUM(total) as val, COUNT(*) as cnt
        FROM sales WHERE client = ? AND sale_type!='AVARIA'
        GROUP BY sale_type""", (name,)).fetchall()
    raw_by_product=conn.execute("""SELECT sale_type, product, SUM(total) as val, SUM(quantity) as qty
        FROM sales WHERE client = ? AND sale_type!='AVARIA'
        GROUP BY sale_type, product""",
        (name,)).fetchall()
    by_product_map={}
    for r in raw_by_product:
        p=norm_p(r["product"] or "", r["sale_type"] or "NF")
        if p not in by_product_map:
            by_product_map[p]={"product":p,"val":0,"qty":0}
        by_product_map[p]["val"]+=float(r["val"] or 0)
        by_product_map[p]["qty"]+=float(r["qty"] or 0)
    by_product=sorted(by_product_map.values(), key=lambda x:x["val"], reverse=True)[:5]
    monthly=conn.execute("""SELECT strftime('%Y-%m',sale_date) as month,
        SUM(total) as val, COUNT(*) as cnt
        FROM sales WHERE client = ? AND sale_type!='AVARIA'
        GROUP BY month ORDER BY month DESC LIMIT 12""",
        (name,)).fetchall()
    conn.close()
    return {"client":dict(cl),"stats":dict(stats) if stats else {},
            "by_type":[dict(r) for r in by_type],
            "by_product":by_product,
            "monthly":[dict(r) for r in monthly]}

@app.get("/api/clients/{client_id}/sales")
def client_sales(client_id:str,limit:int=50,x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    cl=conn.execute("SELECT name FROM clients WHERE id=?",(client_id,)).fetchone()
    if not cl: conn.close(); raise HTTPException(404,"Cliente nÃ£o encontrado.")
    name=cl["name"]
    rows=conn.execute("""SELECT * FROM sales WHERE client = ?
        ORDER BY sale_date DESC, sale_time DESC LIMIT ?""",
        (name,limit)).fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.get("/api/clients/{client_id}/monthly-report")
def client_monthly_report(client_id:str,start_month:str,end_month:str,x_token:str=Header("")):
    require_auth(x_token)
    if not re.fullmatch(r"\d{4}-\d{2}",start_month or "") or not re.fullmatch(r"\d{4}-\d{2}",end_month or ""):
        raise HTTPException(400,"Informe os meses no formato AAAA-MM.")
    if start_month>end_month:
        raise HTTPException(400,"O mes inicial nao pode ser posterior ao mes final.")
    start_date=f"{start_month}-01"
    end_year,end_mon=map(int,end_month.split("-"))
    next_month=f"{end_year+1}-01-01" if end_mon==12 else f"{end_year}-{end_mon+1:02d}-01"
    conn=get_db()
    cl=conn.execute("SELECT * FROM clients WHERE id=? AND active=1",(client_id,)).fetchone()
    if not cl:
        conn.close(); raise HTTPException(404,"Cliente nao encontrado.")
    args=(cl["name"],start_date,next_month)
    rows=conn.execute("""SELECT sale_date,sale_time,sale_type,product,quantity,total,nf_number,delivery_person
        FROM sales WHERE client=? AND sale_type!='AVARIA' AND sale_date>=? AND sale_date<?
        ORDER BY sale_date,sale_time""",args).fetchall()
    monthly=conn.execute("""SELECT substr(sale_date,1,7) month, SUM(total) total,
        SUM(quantity) quantity, COUNT(*) records, COUNT(DISTINCT sale_date) sale_days
        FROM sales WHERE client=? AND sale_type!='AVARIA' AND sale_date>=? AND sale_date<?
        GROUP BY substr(sale_date,1,7) ORDER BY month""",args).fetchall()
    products=conn.execute("""SELECT product, SUM(quantity) quantity, SUM(total) total, COUNT(*) records
        FROM sales WHERE client=? AND sale_type!='AVARIA' AND sale_date>=? AND sale_date<?
        GROUP BY product ORDER BY total DESC""",args).fetchall()
    conn.close()
    return {"client":dict(cl),"start_month":start_month,"end_month":end_month,
            "summary":{"total_value":sum(float(r["total"] or 0) for r in rows),"records":len(rows),
                       "quantity":sum(float(r["quantity"] or 0) for r in rows),"active_months":len(monthly)},
            "monthly":[dict(r) for r in monthly],"products":[dict(r) for r in products],
            "sales":[dict(r) for r in rows]}


@app.post("/api/admin/migrate-products")
def migrate_products(x_token:str=Header("")):
    """Aplica norm_p a TODAS as vendas â€” unifica capitalizaÃ§Ã£o, acentos e
    sinÃ´nimos (Alho 250g/ALHO 250G, Macaxeira a VÃ¡cuo/vacuo, AbÃ³bora/Abobora, etc).
    Reescrito do hardcode anterior (que sÃ³ consertava macaxeira)."""
    require_editor(x_token)
    conn = get_db()
    rows = conn.execute("SELECT id, product, sale_type FROM sales").fetchall()
    changes = {}  # "antigo â†’ novo": qtd
    fixed = 0
    for r in rows:
        old = r["product"] or ""
        if not old:
            continue
        new = norm_p(old, r["sale_type"] or "NF")
        if new != old:
            conn.execute("UPDATE sales SET product=? WHERE id=?", (new, r["id"]))
            key = f"{old} â†’ {new}"
            changes[key] = changes.get(key, 0) + 1
            fixed += 1
    conn.commit(); conn.close()
    return {"ok": True, "fixed": fixed, "total": len(rows), "changes": changes}


@app.get("/api/export-excel")
def export_excel(year:Optional[int]=None,sale_type:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token)
    from openpyxl import Workbook; from openpyxl.styles import Font,PatternFill,Alignment
    conn=get_db(); sql="SELECT * FROM sales WHERE 1=1"; args=[]
    if year: sql+=" AND strftime('%Y',sale_date)=?"; args.append(str(year))
    if sale_type: sql+=" AND sale_type=?"; args.append(sale_type)
    sql+=" ORDER BY sale_date DESC,sale_time DESC"
    rows=conn.execute(sql,args).fetchall(); conn.close()
    wb=Workbook(); ws=wb.active; ws.title="Vendas"
    G="1B5E20"; W="FFFFFF"
    headers=["Data","Hora","Tipo","Cliente","Produto","NF","Placa","Entregador","Qt","P.Unit","Total","ObservaÃ§Ãµes","Origem"]
    for ci,h in enumerate(headers,1):
        c=ws.cell(1,ci,h); c.font=Font(bold=True,color=W,name="Arial",size=10)
        c.fill=PatternFill("solid",fgColor=G); c.alignment=Alignment(horizontal="center")
    for ri,row in enumerate(rows,2):
        d=dict(row)
        vals=[d.get("sale_date",""),d.get("sale_time",""),d.get("sale_type",""),
              d.get("client",""),d.get("product",""),d.get("nf_number",""),
              d.get("plate",""),d.get("delivery_person",""),d.get("quantity",0),
              d.get("unit_price",0),d.get("total",0),d.get("notes",""),d.get("source","")]
        for ci,v in enumerate(vals,1): ws.cell(ri,ci,v)
        ws.cell(ri,11).number_format='R$ #,##0.00'
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=max(len(str(col[0].value or "")),12)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fn=f"Menina_dos_Raios_{year or 'todos'}.xlsx"
    return StreamingResponse(buf,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition":f'attachment; filename="{fn}"'})

# â”€â”€ Backup endpoints â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get("/api/admin/backups")
def list_backups(x_token:str=Header("")):
    require_admin(x_token)
    bks=sorted(_backup_files(),key=lambda p:p.stat().st_mtime,reverse=True)
    result=[]
    for b in bks:
        stat=b.stat()
        n=b.name
        if   "_pre_restore" in n: label="pre_restore"
        elif "_uploaded"    in n: label="uploaded"
        elif "_auto"        in n: label="auto"
        elif "_manual"      in n: label="manual"
        else:                     label="startup"
        result.append({
            "filename":n,
            "size":stat.st_size,
            "size_mb":round(stat.st_size/1024/1024,2),
            "created_at":datetime.fromtimestamp(stat.st_mtime).isoformat(),
            "label":label,
            "kind":"pacote" if n.lower().endswith(".zip") else "banco",
            "databases":_backup_manifest_databases(b) if n.lower().endswith(".zip") else ["bm_monteiro.db"]
        })
    return result

@app.post("/api/admin/backup")
def manual_backup(x_token:str=Header("")):
    require_admin(x_token)
    fname=create_backup("manual")
    if not fname: raise HTTPException(500,"Banco nÃ£o encontrado.")
    return {"ok":True,"filename":fname,
            "kind":"pacote" if fname.lower().endswith(".zip") else "banco",
            "included_databases":[x["path"].name for x in _backup_db_sources()]}

@app.get("/api/admin/backup/{filename}")
def download_backup(filename:str,x_token:str=Header("")):
    require_admin(x_token)
    path=backup_path_for_filename(filename,BACKUP_DIR)
    if not path.exists(): raise HTTPException(404,"Backup nÃ£o encontrado.")
    return FileResponse(str(path),media_type="application/octet-stream",
                       headers={"Content-Disposition":f'attachment; filename="{filename}"'})

@app.delete("/api/admin/backup/{filename}")
def delete_backup(filename:str,x_token:str=Header("")):
    require_admin(x_token)
    path=backup_path_for_filename(filename,BACKUP_DIR)
    if path.exists(): path.unlink()
    return {"ok":True}

@app.get("/api/admin/backup-status")
def backup_status(x_token:str=Header("")):
    require_auth(x_token)
    bks=sorted(_backup_files(),key=lambda p:p.stat().st_mtime,reverse=True)
    last=None
    if bks:
        s=bks[0].stat()
        last={"filename":bks[0].name,"created_at":datetime.fromtimestamp(s.st_mtime).isoformat(),
              "size_mb":round(s.st_size/1024/1024,2),
              "kind":"pacote" if bks[0].suffix.lower()==".zip" else "banco",
              "databases":_backup_manifest_databases(bks[0]) if bks[0].suffix.lower()==".zip" else ["bm_monteiro.db"]}
    return {"total":len(bks),"last":last,"next_auto":"03:00 (diÃ¡rio)","max_kept":MAX_BACKUPS,
            "expected_databases":_backup_expected_databases(),
            "included_databases":[x["path"].name for x in _backup_db_sources()]}

# â”€â”€ Restore endpoints â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _safety_backup_before_restore()->str:
    return safety_backup_before_restore_with(create_backup)

def _restore_zip_backup(src:Path):
    return restore_zip_backup_with(src, BASE_DIR, DB_PATH, APP_NOTES_DB_PATH, COMPANY_DBS)

@app.post("/api/admin/backup/{filename}/restore")
def restore_backup(filename:str,x_token:str=Header("")):
    """Restaura um backup existente da pasta de backups por cima do banco atual.
    Cria automaticamente um backup de seguranÃ§a ('pre_restore') antes de substituir."""
    require_admin(x_token)
    import shutil
    src=backup_path_for_filename(filename,BACKUP_DIR)
    if not src.exists(): raise HTTPException(404,"Backup nÃ£o encontrado.")
    safety=_safety_backup_before_restore()
    try:
        if src.suffix.lower()==".zip":
            restored=_restore_zip_backup(src)
        else:
            if not _is_sqlite_file(src):
                raise HTTPException(400,"Arquivo de backup corrompido (nÃ£o Ã© um banco SQLite vÃ¡lido).")
            shutil.copy2(str(src),str(DB_PATH))
            restored=[DB_PATH.name]
    except Exception as e:
        if isinstance(e,HTTPException): raise e
        raise HTTPException(500,f"Falha ao restaurar: {e}")
    return {"ok":True,"restored_from":filename,"safety_backup":safety,
            "restored_databases":restored,
            "message":"Backup restaurado com sucesso. FaÃ§a login novamente."}

@app.post("/api/admin/backup/upload-restore")
async def upload_restore_backup(file:UploadFile=File(...),x_token:str=Header("")):
    """Recebe um arquivo .db enviado pelo usuÃ¡rio e restaura como banco principal.
    Ãštil para migrar de outra mÃ¡quina ou restaurar a partir de um download anterior."""
    require_admin(x_token)
    fname=(file.filename or "").lower()
    if not fname.endswith((".db",".zip")):
        raise HTTPException(400,"Apenas arquivos .db ou .zip sÃ£o aceitos.")
    content=await read_upload_limited(file, MAX_BACKUP_UPLOAD, "Backup")
    is_zip=content.startswith(b"PK")
    if not is_zip and not content.startswith(b"SQLite format 3"):
        raise HTTPException(400,"Arquivo enviado nÃ£o Ã© um banco SQLite nem pacote de backup vÃ¡lido.")
    # Backup de seguranÃ§a antes
    safety=_safety_backup_before_restore()
    # Salva uma cÃ³pia rotulada do arquivo enviado dentro de backups/ para histÃ³rico
    ts=datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_name=f"bm_backup_{ts}_uploaded.{ 'zip' if is_zip else 'db' }"
    archive_path=BACKUP_DIR/archive_name
    try:
        with open(archive_path,"wb") as f: f.write(content)
        if is_zip:
            restored=_restore_zip_backup(archive_path)
        else:
            with open(DB_PATH,"wb") as f: f.write(content)
            restored=[DB_PATH.name]
    except Exception as e:
        if isinstance(e,HTTPException): raise e
        raise HTTPException(500,f"Falha ao restaurar: {e}")
    # MantÃ©m o teto de backups
    all_bk=sorted(_backup_files(),key=lambda p:p.stat().st_mtime)
    for old in all_bk[:-MAX_BACKUPS]: old.unlink(missing_ok=True)
    clear_sales_cache(); clear_price_cache()
    return {"ok":True,"restored_from":file.filename,"archived_as":archive_name,
            "restored_databases":restored,
            "safety_backup":safety,
            "message":"Backup restaurado com sucesso. FaÃ§a login novamente."}



@app.get("/api/settings/tab-permissions")
def get_tab_permissions(x_token:str=Header("")):
    require_auth(x_token)
    conn=get_control_db()
    row=conn.execute("SELECT value FROM settings WHERE key='tab_permissions'").fetchone()
    conn.close()
    defaults={
        "viewer":["consolidado","nf","pr","avulso","avaria","projecao","grafico","clientes","produtividade","boletos","pendentes","produtos"],
        "editor":["consolidado","nf","pr","avulso","avaria","projecao","grafico","clientes","produtividade","boletos","pendentes","produtos","config"],
        "admin": ["consolidado","nf","pr","avulso","avaria","projecao","grafico","clientes","produtividade","boletos","pendentes","produtos","config"]
    }
    if row:
        import json as _j
        return _j.loads(row["value"])  # Return exactly what was saved â€” no auto-merge
    return defaults

@app.put("/api/settings/tab-permissions")
def set_tab_permissions(body:dict,x_token:str=Header("")):
    require_admin(x_token)
    import json as _j
    val=_j.dumps(body)
    conn=get_control_db()
    try:
        # Ensure settings table has correct structure
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)")
        except Exception: pass
        # Delete + insert to guarantee update regardless of existing constraints
        conn.execute("DELETE FROM settings WHERE key='tab_permissions'")
        conn.execute("INSERT INTO settings(key,value) VALUES('tab_permissions',?)",(val,))
        conn.commit()
        # Verify it was saved
        row=conn.execute("SELECT value FROM settings WHERE key='tab_permissions'").fetchone()
        if not row or row["value"]!=val:
            raise HTTPException(500,"Falha ao salvar permissÃµes no banco.")
        return {"ok":True,"saved":_j.loads(val)}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# â”€â”€ Ordem custom das abas (drag-and-drop / setas â–²â–¼ no painel Admin) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get("/api/settings/tab-order")
def get_tab_order(x_token:str=Header("")):
    """Retorna a ordem custom das abas salva (array), ou None se nunca foi
    configurada (frontend usa ordem default nesse caso)."""
    require_auth(x_token)
    conn=get_control_db()
    row=conn.execute("SELECT value FROM settings WHERE key='tab_order'").fetchone()
    conn.close()
    if row:
        import json as _j
        return _j.loads(row["value"])
    return None

@app.api_route("/api/settings/tab-order", methods=["POST","PUT"])
async def set_tab_order(request:Request,x_token:str=Header("")):
    """Salva a ordem custom das abas. Body: JSON array de strings (ids das abas).
    Aceita POST (frontend principal) e PUT (compatibilidade).
    Usa Request.json() porque FastAPI nÃ£o aceita JSON array diretamente como
    parÃ¢metro `body:list` â€” era a causa do erro [object Object] (422 com array
    de erros virava string '[object Object]' no toast do frontend)."""
    require_admin(x_token)
    body=await request.json()
    if not isinstance(body,list) or not all(isinstance(x,str) for x in body):
        raise HTTPException(400,"Esperado: array de strings com a ordem das abas.")
    import json as _j
    val=_j.dumps(body)
    conn=get_control_db()
    try:
        try:
            conn.execute("CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)")
        except Exception: pass
        conn.execute("DELETE FROM settings WHERE key='tab_order'")
        conn.execute("INSERT INTO settings(key,value) VALUES('tab_order',?)",(val,))
        conn.commit()
        row=conn.execute("SELECT value FROM settings WHERE key='tab_order'").fetchone()
        if not row or row["value"]!=val:
            raise HTTPException(500,"Falha ao salvar ordem das abas no banco.")
        return {"ok":True,"saved":body}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# â”€â”€ Config: quem pode lanÃ§ar pagamento Monteiro â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get("/api/settings/monteiro-payment-permission")
def get_monteiro_payment_permission(x_token:str=Header("")):
    require_auth(x_token)
    conn=get_db()
    row=conn.execute("SELECT value FROM settings WHERE key='monteiro_payment_perm'").fetchone()
    conn.close()
    if row:
        import json as _j
        return {"roles": _j.loads(row["value"])}
    return {"roles": ["admin"]}

@app.put("/api/settings/monteiro-payment-permission")
def set_monteiro_payment_permission(body:dict,x_token:str=Header("")):
    require_admin(x_token)
    roles=body.get("roles",["admin"])
    import json as _j
    conn=get_db()
    conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('monteiro_payment_perm',?)",(_j.dumps(roles),))
    conn.commit(); conn.close()
    return {"ok":True}

def _monteiro_calendar_roles():
    conn=get_db()
    row=conn.execute("SELECT value FROM settings WHERE key='monteiro_calendar_perm'").fetchone()
    conn.close()
    if row:
        import json as _j
        try: return _j.loads(row["value"])
        except Exception: pass
    return ["admin"]

def require_monteiro_calendar(x_token:str="")->dict:
    sess=require_auth(x_token)
    roles=_monteiro_calendar_roles()
    if sess.get("role") not in roles:
        raise HTTPException(403,"Seu perfil nao tem permissao para acessar o Calendario do app.")
    return sess

@app.get("/api/settings/monteiro-calendar-permission")
def get_monteiro_calendar_permission(x_token:str=Header("")):
    require_auth(x_token)
    return {"roles": _monteiro_calendar_roles()}

@app.put("/api/settings/monteiro-calendar-permission")
def set_monteiro_calendar_permission(body:dict,x_token:str=Header("")):
    require_admin(x_token)
    roles=body.get("roles",["admin"])
    roles=[r for r in roles if r in ("admin","editor","viewer")]
    if "admin" not in roles:
        roles.insert(0,"admin")
    import json as _j
    conn=get_db()
    conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('monteiro_calendar_perm',?)",(_j.dumps(roles),))
    conn.commit(); conn.close()
    return {"ok":True,"roles":roles}


@lru_cache(maxsize=64)
def _summary_array_cached(company:str,year:int,bucket:int):
    conn=get_db(company)
    rows=conn.execute("""SELECT strftime('%m',sale_date) AS month,
               sale_type, SUM(total) AS total_val, COUNT(*) as cnt
        FROM sales WHERE strftime('%Y',sale_date)=?
        GROUP BY month, sale_type ORDER BY month, sale_type""",
        (str(year),)).fetchall()
    conn.close(); return tuple(tuple(dict(r).items()) for r in rows)

@app.get("/api/summary")
def get_summary_array(year:Optional[int]=None,x_token:str=Header("")):
    """Monthly breakdown for chart - returns [{month, sale_type, total_val, cnt}]"""
    require_any_tab_access(x_token,["consolidado","grafico","produtividade"])
    y=year or datetime.now().year
    data=_summary_array_cached(_company_key(CURRENT_COMPANY.get()),int(y),_cache_bucket(60))
    return [dict(items) for items in data]


@app.get("/api/kpi-summary")
def get_kpi_summary(year:Optional[int]=None,month:Optional[int]=None,x_token:str=Header("")):
    require_auth(x_token)
    conn=get_db()
    y=year or datetime.now().year
    m=month or None
    where="strftime('%Y',sale_date)=?"
    args=[str(y)]
    if m:
        where+=" AND strftime('%m',sale_date)=?"
        args.append(f"{int(m):02d}")
    rows=conn.execute("""
        SELECT sale_type, SUM(total) as total, COUNT(*) as cnt,
               COUNT(DISTINCT nf_number) as nfs
        FROM sales WHERE """+where+"""
        GROUP BY sale_type
    """,args).fetchall()
    total_rows=conn.execute("""
        SELECT COUNT(*) as cnt FROM sales WHERE """+where+"""
    """,args).fetchone()
    conn.close()
    by={r["sale_type"]:{"total":r["total"],"cnt":r["cnt"],"nfs":r["nfs"]} for r in rows}
    nf =by.get("NF",{}).get("total",0)
    pr =by.get("PR",{}).get("total",0)
    av =by.get("AVULSO",{}).get("total",0)
    avr=by.get("AVARIA",{}).get("total",0)
    return {
        "year":y,
        "month":m,
        "total_records":total_rows["cnt"] if total_rows else 0,
        "NF":{"total":nf,"cnt":by.get("NF",{}).get("cnt",0),"nfs":by.get("NF",{}).get("nfs",0)},
        "PR":{"total":pr,"cnt":by.get("PR",{}).get("cnt",0)},
        "AVULSO":{"total":av,"cnt":by.get("AVULSO",{}).get("cnt",0)},
        "AVARIA":{"total":avr,"cnt":by.get("AVARIA",{}).get("cnt",0)},
        "receita_bruta":nf+pr+av,
        "receita_liquida":nf+pr+av-avr,
        "ticket_medio":(nf+pr+av)/(by.get("NF",{}).get("cnt",0)+by.get("PR",{}).get("cnt",0)+by.get("AVULSO",{}).get("cnt",0)) if (by.get("NF",{}).get("cnt",0)+by.get("PR",{}).get("cnt",0)+by.get("AVULSO",{}).get("cnt",0))>0 else 0
    }


# â”€â”€ DeduplicaÃ§Ã£o de Clientes â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.get("/api/clients/duplicates")
def find_duplicates(x_token:str=Header("")):
    """Find client name duplicates based on normalized names."""
    require_auth(x_token)
    conn = get_db()
    rows = conn.execute("""
        SELECT TRIM(client) as client, COUNT(*) as sales_cnt,
               SUM(total) as total_val, MAX(sale_date) as last_date
        FROM sales WHERE client IS NOT NULL AND TRIM(client)!=''
        GROUP BY TRIM(client) COLLATE NOCASE
        ORDER BY client
    """).fetchall()
    conn.close()

    # Group by normalized key
    groups = {}
    for r in rows:
        key = _normalize_client(r["client"])
        if key not in groups:
            groups[key] = []
        groups[key].append({
            "name": r["client"],
            "sales_cnt": r["sales_cnt"],
            "total_val": r["total_val"],
            "last_date": r["last_date"]
        })

    # Return only groups with more than 1 variant
    duplicates = [
        {"key": k, "variants": v}
        for k, v in groups.items()
        if len(v) > 1
    ]
    duplicates.sort(key=lambda x: -sum(v["sales_cnt"] for v in x["variants"]))
    return duplicates

@app.post("/api/clients/merge")
def merge_clients(body:dict, x_token:str=Header("")):
    """Rename all sales/boletos/clients from source names to target name."""
    require_editor(x_token)
    target = (body.get("target") or "").strip()
    sources = [s.strip() for s in (body.get("sources") or []) if s.strip() and s.strip() != target]
    if not target or not sources:
        raise HTTPException(400, "target e sources sÃ£o obrigatÃ³rios.")
    conn = get_db()
    stage = "init"
    try:
        total_updated = 0
        for src_name in sources:
            # 1) Sales â€” no unique constraint, safe to bulk update
            stage = f"sales:{src_name}"
            n = conn.execute(
                "UPDATE sales SET client=? WHERE TRIM(client)=?",
                (target, src_name)
            ).rowcount
            total_updated += n

            # 2) Clients table â€” keep target, remove duplicate source row
            stage = f"clients:{src_name}"
            try:
                tgt_exists = conn.execute(
                    "SELECT 1 FROM clients WHERE TRIM(name)=? LIMIT 1", (target,)
                ).fetchone()
                if tgt_exists:
                    conn.execute("DELETE FROM clients WHERE TRIM(name)=?", (src_name,))
                else:
                    conn.execute("UPDATE clients SET name=? WHERE TRIM(name)=?", (target, src_name))
            except Exception:
                pass  # clients table is optional metadata

            # 3) Boletos â€” UNIQUE(client, sale_date): resolve conflicts row by row
            stage = f"boletos:{src_name}"
            try:
                src_boletos = conn.execute(
                    "SELECT id, sale_date FROM boletos WHERE TRIM(client)=?",
                    (src_name,)
                ).fetchall()
                for b in src_boletos:
                    dup = conn.execute(
                        "SELECT id FROM boletos WHERE TRIM(client)=? AND sale_date=?",
                        (target, b["sale_date"])
                    ).fetchone()
                    if dup:
                        conn.execute("DELETE FROM boletos WHERE id=?", (b["id"],))
                    else:
                        conn.execute("UPDATE boletos SET client=? WHERE id=?", (target, b["id"]))
            except Exception:
                # If anything still conflicts, just remove the source boletos
                conn.execute("DELETE FROM boletos WHERE TRIM(client)=?", (src_name,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        raise HTTPException(500, f"Erro ao mesclar (etapa {stage}): {type(e).__name__}: {e}")
    conn.close()
    return {"ok": True, "updated": total_updated, "target": target, "merged": sources}


# â”€â”€ Boletos â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def init_boletos(conn):
    conn.execute("""CREATE TABLE IF NOT EXISTS boletos (
        id         TEXT PRIMARY KEY,
        client     TEXT NOT NULL,
        sale_date  TEXT NOT NULL,
        nf_number  TEXT,
        total_val  REAL NOT NULL DEFAULT 0,
        due_date   TEXT,
        status     TEXT NOT NULL DEFAULT 'pendente',
        paid_date  TEXT,
        notes      TEXT,
        created_at TEXT DEFAULT (datetime('now')),
        UNIQUE(client, sale_date, nf_number)
    )""")

@app.get("/api/boletos")
def list_boletos(period:str="",status:str="",search:str="",
                 year:Optional[int]=None, month:Optional[int]=None,
                 x_token:str=Header("")):
    """
    Lista boletos agrupados por cliente+data de venda.

    Filtros mutuamente compatÃ­veis:
      - period: 'month' (mÃªs corrente), 'week' (Ãºltimos 7 dias). Mantido para retrocompat.
      - year + month: ano/mÃªs especÃ­ficos (precedÃªncia sobre period quando ambos enviados).
      - year sÃ³: ano inteiro.
      - status, search: filtros pÃ³s-agregaÃ§Ã£o.
    """
    require_auth(x_token); conn=get_db()
    try:
        # Monta WHERE parametrizado (sem f-string para evitar SQL injection / mÃ¡ prÃ¡tica)
        where_parts = ["sale_type != 'AVARIA'", "client IS NOT NULL", "client != ''"]
        args = []

        # year + month tem precedÃªncia sobre period (request explÃ­cito do usuÃ¡rio)
        if year is not None and month is not None:
            where_parts.append("strftime('%Y-%m', sale_date) = ?")
            args.append(f"{int(year)}-{int(month):02d}")
        elif year is not None:
            where_parts.append("strftime('%Y', sale_date) = ?")
            args.append(str(int(year)))
        elif period == "month":
            where_parts.append("strftime('%Y-%m', sale_date) = ?")
            args.append(datetime.now().strftime('%Y-%m'))
        elif period == "week":
            where_parts.append("sale_date >= ?")
            args.append((datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'))

        where_sql = " AND ".join(where_parts)
        sales_grouped = conn.execute(
            f"""
            SELECT TRIM(client) as client, sale_date,
                   SUM(total) as total_val, COUNT(*) as item_count,
                   GROUP_CONCAT(DISTINCT product) as products,
                   nf_number as sample_nf
            FROM sales WHERE {where_sql}
            GROUP BY TRIM(client) COLLATE NOCASE, sale_date, nf_number
            ORDER BY sale_date DESC, TRIM(client) COLLATE NOCASE
            """, args
        ).fetchall()

        # Carrega metadata de boletos existentes
        boleto_meta = {}
        for row in conn.execute("SELECT * FROM boletos").fetchall():
            k = row['client'].strip().upper() + '|' + row['sale_date'] + '|' + (row['nf_number'] or '')
            boleto_meta[k] = dict(row)

        # Coleta boletos que precisam ser criados, faz INSERT em massa numa Ãºnica transaÃ§Ã£o
        new_rows = []
        for g in sales_grouped:
            cl = g['client'].strip(); dt = g['sale_date']
            nf = g['sample_nf'] or ''
            k = cl.upper() + '|' + dt + '|' + nf
            if k not in boleto_meta:
                new_rows.append((str(uuid.uuid4()), cl, dt, nf, g['total_val'], 'pendente'))

        if new_rows:
            try:
                conn.executemany(
                    "INSERT OR IGNORE INTO boletos(id,client,sale_date,nf_number,total_val,status) VALUES(?,?,?,?,?,?)",
                    new_rows
                )
                conn.commit()
            except Exception:
                conn.rollback()
                logger.exception("Falha ao gerar boletos automaticamente na listagem")
            else:
                try:
                    # Re-le boletos para pegar IDs autoritativos (resolve corrida com concorrentes)
                    for row in conn.execute("SELECT * FROM boletos").fetchall():
                        k = row['client'].strip().upper() + '|' + row['sale_date'] + '|' + (row['nf_number'] or '')
                        boleto_meta[k] = dict(row)
                except Exception:
                    logger.exception("Falha ao recarregar boletos apos geracao automatica")

        result = []
        today = datetime.now().strftime('%Y-%m-%d')
        for g in sales_grouped:
            cl = g['client'].strip(); dt = g['sale_date']; nf = g['sample_nf'] or ''
            k = cl.upper() + '|' + dt + '|' + nf
            meta = boleto_meta.get(k, {})
            bid     = meta.get('id') or str(uuid.uuid4())
            bstatus = meta.get('status', 'pendente')
            bdue    = meta.get('due_date')
            bpaid   = meta.get('paid_date')
            bnotes  = meta.get('notes')

            if status and bstatus != status: continue
            if search and search.lower() not in cl.lower(): continue

            is_overdue   = bool(bdue and bdue < today and bstatus == 'pendente')
            is_due_today = bool(bdue == today and bstatus == 'pendente')

            result.append({
                'id': bid, 'client': cl, 'sale_date': dt, 'nf_number': nf,
                'total_val': g['total_val'], 'item_count': g['item_count'],
                'products': g['products'], 'sample_nf': nf,
                'due_date': bdue, 'status': bstatus, 'paid_date': bpaid, 'notes': bnotes,
                'is_overdue': is_overdue, 'is_due_today': is_due_today
            })

        # Filtro por clientes habilitados na config de boletos
        try:
            import json as _j
            cfg_row = conn.execute("SELECT value FROM settings WHERE key='boleto_clients'").fetchone()
            if cfg_row:
                enabled = set(_j.loads(cfg_row["value"]))
                enabled_norm = set(_normalize_name(e) for e in enabled)
                before = len(result)
                result = [b for b in result if _normalize_name(b["client"]) in enabled_norm]
                print(f"[boletos] backend filtro: {before}->{len(result)} boletos")
            else:
                print(f"[boletos] backend: sem config de clientes, exibindo {len(result)} boletos")
        except Exception as ex:
            print(f"[boletos] backend erro no filtro: {ex}")

        return result
    finally:
        conn.close()

@app.get("/api/boletos/clients-config")
def get_boleto_clients(x_token:str=Header("")):
    require_auth(x_token)
    conn=get_db()
    row=conn.execute("SELECT value FROM settings WHERE key='boleto_clients'").fetchone()
    raw_clients=[r["client"] for r in conn.execute(
        "SELECT DISTINCT TRIM(client) as client FROM sales WHERE client IS NOT NULL AND TRIM(client)!='' AND sale_type!='AVARIA' ORDER BY client COLLATE NOCASE"
    ).fetchall()]
    import unicodedata
    def norm_name(s):
        return unicodedata.normalize("NFD",s.upper()).encode("ascii","ignore").decode()
    seen_norm={}
    all_clients=[]
    for c in raw_clients:
        n=norm_name(c)
        if n not in seen_norm:
            seen_norm[n]=c
            all_clients.append(c)
    conn.close()
    if row:
        import json as _j
        enabled=set(_j.loads(row["value"]))
        return {"configured": True, "clients":[{"name":c,"enabled":c in enabled} for c in all_clients]}
    return {"configured": False, "clients":[{"name":c,"enabled":True} for c in all_clients]}

@app.put("/api/boletos/clients-config")
def set_boleto_clients(body:dict,x_token:str=Header("")):
    require_auth(x_token)
    import json as _j
    enabled=body.get("enabled",[])
    conn=get_db()
    conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('boleto_clients',?)",(_j.dumps(enabled),))
    conn.commit(); conn.close()
    return {"ok":True}

@app.put("/api/boletos/{bid}")
def update_boleto(bid:str,body:dict,x_token:str=Header("")):
    # Era require_auth â€” viewer conseguia marcar/reverter pagamento. Corrigido.
    require_editor(x_token); conn=get_db()
    try:
        fields=[]
        vals=[]
        for f in ['due_date','status','paid_date','notes','total_val']:
            if f in body:
                fields.append(f+"=?")
                vals.append(body[f])
        if fields:
            vals.append(bid)
            conn.execute(f"UPDATE boletos SET {','.join(fields)} WHERE id=?",vals)
            conn.commit()
        return {"ok":True}
    except Exception:
        try:
            conn.rollback()
        finally:
            raise
    finally:
        conn.close()

@app.get("/api/boletos/{bid}/items")
def boleto_items(bid:str,x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    b=conn.execute("SELECT * FROM boletos WHERE id=?",(bid,)).fetchone()
    if not b: conn.close(); raise HTTPException(404,"Boleto nÃ£o encontrado.")
    rows=conn.execute("""SELECT * FROM sales WHERE TRIM(client) COLLATE NOCASE=?
        AND sale_date=? AND sale_type!='AVARIA'
        AND (nf_number=? OR (nf_number IS NULL AND ?=''))
        ORDER BY sale_time""",(b['client'],b['sale_date'],b['nf_number'] or '',b['nf_number'] or '')).fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.get("/api/boletos/summary")
def boleto_summary(x_token:str=Header("")):
    # Deriva os KPIs da MESMA fonte da lista (boletos originados das vendas +
    # filtro de clientes habilitados), evitando contar boletos "Ã³rfÃ£os" (sem
    # venda correspondente) que a lista nÃ£o consegue exibir. Assim o card
    # "Em atraso" sempre bate com o que aparece na lista.
    require_auth(x_token)
    allb = list_boletos(x_token=x_token)
    today = datetime.now().strftime('%Y-%m-%d')
    cur_month = datetime.now().strftime('%Y-%m')
    def _m(d): return (d or '')[:7]
    total_pendente = sum(float(b['total_val'] or 0) for b in allb if b['status'] == 'pendente')
    total_pago     = sum(float(b['total_val'] or 0) for b in allb if b['status'] == 'pago')
    total_pago_mes = sum(float(b['total_val'] or 0) for b in allb
                         if b['status'] == 'pago' and _m(b['paid_date'] or b['sale_date']) == cur_month)
    vencidos   = sum(1 for b in allb if b['status'] == 'pendente' and b['due_date'] and b['due_date'] < today)
    vence_hoje = sum(1 for b in allb if b['status'] == 'pendente' and b['due_date'] == today)
    return {'total_pendente': total_pendente, 'total_pago': total_pago,
            'total_pago_mes': total_pago_mes, 'vencidos': vencidos, 'vence_hoje': vence_hoje}

@app.post("/api/sebrae/verify-nf")
async def sebrae_verify_nf(file: UploadFile = File(...), source: str = "sebrae", x_token: str = Header("")):
    """Recebe PDF de conferencia de notas e compara com os lancamentos do sistema.
    source=sebrae usa o layout de NF-e do Sebrae; source=produtor_rural usa o
    Relatorio de Emissoes do produtor rural."""
    require_auth(x_token)
    try:
        import pdfplumber as _pdfplumber
        import io as _io
    except ImportError as _e:
        raise HTTPException(500, f"Dependencia ausente no servidor: {_e}. Execute: pip install pdfplumber")

    source = (source or "sebrae").strip().lower().replace("-", "_")
    if source not in ("sebrae", "produtor_rural", "pr"):
        raise HTTPException(400, "Tipo de relatorio invalido.")
    is_pr_report = source in ("produtor_rural", "pr")

    fname=(file.filename or "").lower()
    if fname and not fname.endswith(".pdf"):
        raise HTTPException(400, "Envie um PDF valido.")
    content = await read_upload_limited(file, MAX_VERIFY_PDF_UPLOAD, "PDF")
    full_text = ""
    try:
        with _pdfplumber.open(_io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"
    except Exception as e:
        raise HTTPException(400, f"Erro ao ler PDF: {e}")

    sebrae_entries = []
    if is_pr_report:
        # Layout do Relatorio de Emissoes do produtor rural:
        # 04/07/2026 09:46:21 40720001 0,00 0,00 0,00 0 120,00
        line_pattern = re.compile(
            r'^\s*(\d{2}/\d{2}/\d{4})\s+\d{2}:\d{2}:\d{2}\s+(\d+)\s+'
            r'[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+([\d.]+,\d{2})\s*$',
            re.MULTILINE
        )
        seen = set()
        for m in line_pattern.finditer(full_text):
            date_str, num, value_str = m.group(1), m.group(2), m.group(3)
            if num in seen:
                continue
            seen.add(num)
            value = 0.0
            try:
                value = float(value_str.replace(".", "").replace(",", "."))
            except ValueError:
                pass
            sebrae_entries.append({
                "nf_number": num, "serie": "", "nf_full": num,
                "status": "Emitida", "date": date_str,
                "value": value,
            })
    else:
        nf_pattern = re.compile(
            r'(\d{1,5})/(\d+)'
            r'(?:[^\n]*\n){0,4}'
            r'.*?(\d{2}/\d{2}/\d{4})'
            r'.*?\n.*?\n'
            r'(Emitida|Cancelada)',
            re.MULTILINE
        )
        # Extracao de valor robusta (janela por NF) - cobre os dois layouts do PDF.
        nf_starts = list(re.finditer(r'^(\d{1,5}/\d+)', full_text, re.MULTILINE))
        val_map = {}
        for i, ms in enumerate(nf_starts):
            nf_key = ms.group(1)
            win_start = ms.end()
            win_end = nf_starts[i + 1].start() if i + 1 < len(nf_starts) else len(full_text)
            window = full_text[win_start:win_end]
            vm = re.search(r'R[\$]\s*([\d.]+)', window)
            if not vm:
                continue
            inteiro = vm.group(1).replace(".", "")
            cm = re.search(r',(\d{2})', window[vm.end():])
            cents = cm.group(1) if cm else "00"
            try:
                val_map.setdefault(nf_key, float(f"{inteiro}.{cents}"))
            except ValueError:
                pass
        seen = set()
        for m in nf_pattern.finditer(full_text):
            num, serie, date_str, status = m.group(1), m.group(2), m.group(3), m.group(4)
            nf_full = f"{num}/{serie}"
            if nf_full in seen:
                continue
            seen.add(nf_full)
            sebrae_entries.append({
                "nf_number": num, "serie": serie, "nf_full": nf_full,
                "status": status, "date": date_str,
                "value": val_map.get(nf_full, val_map.get(num, 0.0)),
            })

    emitidas = [e for e in sebrae_entries if e["status"] == "Emitida"]
    if not emitidas:
        tipo = "Produtor Rural" if is_pr_report else "Sebrae"
        raise HTTPException(422, f"Nenhuma nota emitida encontrada no PDF. Verifique se e o arquivo correto de {tipo}.")

    conn = get_db()
    sale_types = ("PR",) if is_pr_report else ("NF", "AVULSO")
    placeholders = ",".join(["?"] * len(sale_types))
    rows = conn.execute(
        "SELECT nf_number FROM sales "
        f"WHERE sale_type IN ({placeholders}) AND nf_number IS NOT NULL AND nf_number != ''",
        sale_types
    ).fetchall()
    conn.close()

    system_nfs = set()
    for r in rows:
        nf = str(r["nf_number"]).strip()
        if not nf:
            continue
        system_nfs.add(nf)
        system_nfs.add(nf.split("/")[0])
        system_nfs.add(re.sub(r'\D+', '', nf))

    result_list = []
    for e in emitidas:
        num_digits = re.sub(r'\D+', '', e["nf_number"])
        e["in_system"] = (e["nf_full"] in system_nfs) or (e["nf_number"] in system_nfs) or (num_digits in system_nfs)
        result_list.append(e)
    missing = [e for e in result_list if not e["in_system"]]
    return {
        "source": "produtor_rural" if is_pr_report else "sebrae",
        "total_sebrae": len(emitidas),
        "canceladas_ignoradas": len(sebrae_entries) - len(emitidas),
        "found_in_system": len(result_list) - len(missing),
        "missing_count": len(missing),
        "entries": result_list,
    }


@app.get("/api/boletos/paid")
def boletos_paid(x_token:str=Header("")):
    """Retorna todos os boletos pagos, agregados por mÃªs de pagamento."""
    require_auth(x_token); conn=get_db()
    rows=conn.execute("""
        SELECT b.id, b.client, b.sale_date, b.total_val, b.due_date, b.paid_date, b.notes,
               strftime('%Y-%m', COALESCE(b.paid_date, b.sale_date)) as month
        FROM boletos b
        WHERE b.status='pago'
        ORDER BY COALESCE(b.paid_date, b.sale_date) DESC, b.client
    """).fetchall()
    # Agrega por mÃªs
    by_month={}
    for r in rows:
        m=r['month'] or 'sem-data'
        if m not in by_month:
            by_month[m]={'month':m,'total':0.0,'count':0,'boletos':[]}
        by_month[m]['total']+=float(r['total_val'] or 0)
        by_month[m]['count']+=1
        by_month[m]['boletos'].append({
            'id':r['id'],'client':r['client'],'sale_date':r['sale_date'],
            'total_val':r['total_val'],'due_date':r['due_date'],
            'paid_date':r['paid_date'],'notes':r['notes']
        })
    conn.close()
    # Ordena por mÃªs (mais recente primeiro)
    return sorted(by_month.values(), key=lambda x:x['month'], reverse=True)



# â”€â”€ Paladar (mÃ³dulo independente) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.get("/api/monteiro/products")
@app.get("/api/paladar/products")
def paladar_products(active:Optional[str]=None,search:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    where=[]; args=[]
    if active=='all':
        pass
    else:
        where.append("active=1")
    if search:
        where.append("name LIKE ?"); args.append(f"%{search}%")
    ws=" AND ".join(where) if where else "1"
    rows=conn.execute(f"SELECT * FROM paladar_products WHERE {ws} ORDER BY name COLLATE NOCASE",args).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/monteiro/products")
@app.post("/api/paladar/products")
def create_paladar_product(body:dict,x_token:str=Header("")):
    require_editor_tab_access(x_token,["produtos","cfg_precos"]); conn=get_db()
    name=body.get("name","").strip()
    if not name: conn.close(); return {"error":"Nome obrigatÃ³rio"}
    price=float(body.get("suggested_price",0))
    try:
        conn.execute("INSERT INTO paladar_products(name,suggested_price) VALUES(?,?)",(name,price))
        pid=conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.commit(); conn.close()
        return {"id":pid,"name":name,"suggested_price":price,"message":"ok"}
    except sqlite3.IntegrityError:
        conn.close(); return {"error":"Produto jÃ¡ existe"}

@app.put("/api/monteiro/products/{pid}")
@app.put("/api/paladar/products/{pid}")
def update_paladar_product(pid:str,body:dict,x_token:str=Header("")):
    require_editor_tab_access(x_token,["produtos","cfg_precos"]); conn=get_db()
    name=body.get("name","").strip()
    price=float(body.get("suggested_price",0))
    active=body.get("active",1)
    if name:
        conn.execute("UPDATE paladar_products SET name=?, suggested_price=?, active=? WHERE id=?",(name,price,active,pid))
    else:
        conn.execute("UPDATE paladar_products SET suggested_price=?, active=? WHERE id=?",(price,active,pid))
    conn.commit(); conn.close()
    return {"message":"ok"}

@app.delete("/api/monteiro/products/{pid}")
@app.delete("/api/paladar/products/{pid}")
def delete_paladar_product(pid:str,x_token:str=Header("")):
    require_editor_tab_access(x_token,["produtos","cfg_precos"]); conn=get_db()
    conn.execute("DELETE FROM paladar_products WHERE id=?",(pid,))
    conn.commit(); conn.close()
    return {"ok":True}

@app.get("/api/monteiro/summary")
@app.get("/api/paladar/summary")
def paladar_summary(period:Optional[str]=None,month:Optional[str]=None,year:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    where, args = _pal_period_where(period, month, year)
    ws=" AND ".join(where) if where else "1"
    row=conn.execute(f"""SELECT
        COALESCE(SUM(total),0) as total_receita,
        COUNT(DISTINCT sale_group) as total_vendas,
        CASE WHEN COUNT(DISTINCT sale_group)>0 THEN COALESCE(SUM(total),0)/COUNT(DISTINCT sale_group) ELSE 0 END as ticket_medio,
        COUNT(DISTINCT saledate) as dias_produtivos
        FROM paladar_sales WHERE {ws}""",args).fetchone()
    # Melhor dia
    melhor=conn.execute(f"""SELECT saledate as date, SUM(total) as total, SUM(quantity) as qty, COUNT(DISTINCT sale_group) as count
        FROM paladar_sales WHERE {ws} GROUP BY saledate ORDER BY total DESC LIMIT 1""",args).fetchone()
    # Por produto
    prod_raw=conn.execute(f"""SELECT product, SUM(total) as total, SUM(quantity) as qty, COUNT(DISTINCT sale_group) as count
        FROM paladar_sales WHERE {ws} GROUP BY product ORDER BY total DESC""",args).fetchall()
    prod_map={}
    for r in prod_raw:
        p=norm_p(r["product"] or "", "NF")
        if p not in prod_map:
            prod_map[p]={"product":p,"total":0,"qty":0,"count":0}
        prod_map[p]["total"]+=float(r["total"] or 0)
        prod_map[p]["qty"]+=float(r["qty"] or 0)
        prod_map[p]["count"]+=int(r["count"] or 0)
    prod=sorted(prod_map.values(), key=lambda x:x["total"], reverse=True)
    # Por dia
    dia=conn.execute(f"""SELECT saledate as date, SUM(total) as total, SUM(quantity) as qty, COUNT(DISTINCT sale_group) as count
        FROM paladar_sales WHERE {ws} GROUP BY saledate ORDER BY saledate""",args).fetchall()
    conn.close()
    return {
        "total_receita": float(row["total_receita"]),
        "total_vendas": row["total_vendas"],
        "ticket_medio": float(row["ticket_medio"]),
        "dias_produtivos": row["dias_produtivos"],
        "melhor_dia": dict(melhor) if melhor else {"date":None,"total":0,"qty":0,"count":0},
        "por_produto": prod,
        "por_dia": [dict(r) for r in dia]
    }

@app.get("/api/monteiro/sales")
@app.get("/api/paladar/sales")
def paladar_sales(period:Optional[str]=None,month:Optional[str]=None,year:Optional[str]=None,
                  product:Optional[str]=None,
                  nf_number:Optional[str]=None,driver:Optional[str]=None,
                  vehicle_plate:Optional[str]=None,client:Optional[str]=None,
                  x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    where, args = _pal_period_where(period, month, year)
    if product: where.append("product like ?"); args.append(f"%{product}%")
    if nf_number: where.append("nf_number like ?"); args.append(f"%{nf_number}%")
    if driver: where.append("driver like ?"); args.append(f"%{driver}%")
    if vehicle_plate: where.append("(vehicle like ? OR plate like ?)"); args.append(f"%{vehicle_plate}%"); args.append(f"%{vehicle_plate}%")
    if client: where.append("client like ?"); args.append(f"%{client}%")
    ws=" AND ".join(where) if where else "1"
    rows=conn.execute(f"SELECT * FROM paladar_sales WHERE {ws} ORDER BY saledate DESC, sale_group",args).fetchall()
    # Agrupar por sale_group
    groups={}
    for r in rows:
        sg=r["sale_group"]
        if sg not in groups:
            groups[sg]={
                "sale_group":sg,"saledate":r["saledate"],"client":r["client"],
                "nf_number":r["nf_number"],"driver":r["driver"],"vehicle":r["vehicle"],
                "plate":r["plate"],"notes":r["notes"],"total_group":0,"items":[],
                "invoice_file_path":r["invoice_file_path"],"invoice_original_name":r["invoice_original_name"],
                "invoice_mime":r["invoice_mime"],"invoice_uploaded_at":r["invoice_uploaded_at"]
            }
        groups[sg]["items"].append(dict(r))
        groups[sg]["total_group"]=round(groups[sg]["total_group"]+(r["total"] or 0),2)
    conn.close()
    return list(groups.values())

@app.post("/api/monteiro/sales")
@app.post("/api/paladar/sales")
def create_paladar_sale(body:dict,x_token:str=Header("")):
    require_editor_tab_access(x_token,["consolidado","nf","avulso","pendentes"]); conn=get_db()
    try:
        saledate=body.get("saledate")
        items=body.get("items")
        # Campos de cabeÃ§alho (cÃ³pia para todos os itens do grupo)
        nf_number=body.get("nf_number","") or ""
        driver=body.get("driver","") or ""
        vehicle=body.get("vehicle","") or ""
        plate=body.get("plate","") or ""
        client=body.get("client","") or ""
        # Suporte legado: body antigo de item Ãºnico
        if not items:
            items=[{"product":body.get("product"),"quantity":body.get("quantity",1),
                    "unitprice":body.get("unitprice",0),"total":body.get("total",0),
                    "notes":body.get("notes","")}]
        if len(items) > 20:
            raise HTTPException(400, "MÃ¡ximo de 20 itens por venda.")
        import uuid
        # Agrupar por NF: se jÃ¡ existir grupo com mesma NF+data+cliente, reaproveitar
        group=None
        if nf_number:
            existing=conn.execute("SELECT sale_group FROM paladar_sales WHERE nf_number=? AND saledate=? AND client=? LIMIT 1",
                                  (nf_number,saledate,client)).fetchone()
            if existing:
                group=existing["sale_group"]
                # Atualizar campos de cabeÃ§alho no grupo existente
                conn.execute("UPDATE paladar_sales SET driver=?,vehicle=?,plate=?,notes=? WHERE sale_group=? AND id=(SELECT MIN(id) FROM paladar_sales WHERE sale_group=?)",
                             (driver,vehicle,plate,notes,group,group))
        if not group:
            group=str(uuid.uuid4())[:8]
        ids=[]
        for it in items:
            conn.execute("""INSERT INTO paladar_sales(sale_group,saledate,product,quantity,unitprice,total,notes,
                nf_number,driver,vehicle,plate,client)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                (group,saledate,it.get("product"),
                 float(it.get("quantity",1)),float(it.get("unitprice",0)),
                 float(it.get("total",0)),it.get("notes",""),
                 nf_number,driver,vehicle,plate,client))
            ids.append(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
        conn.commit()
        return {"ids":ids,"group":group,"message":"ok"}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def _monteiro_invoice_row(conn, group_id: str):
    return conn.execute("""SELECT sale_group, invoice_file_path, invoice_original_name, invoice_mime
        FROM paladar_sales WHERE sale_group=? LIMIT 1""", (group_id,)).fetchone()

@app.post("/api/monteiro/sales/group/{group_id}/invoice")
@app.post("/api/paladar/sales/group/{group_id}/invoice")
async def upload_monteiro_invoice(group_id: str, file: UploadFile = File(...), x_token: str = Header("")):
    require_editor_tab_access(x_token,["nf","pendentes","consolidado"])
    gid = (group_id or "").strip()
    if not gid:
        raise HTTPException(400, "Grupo invÃ¡lido.")
    original = file.filename or "nota_fiscal"
    ext = Path(original).suffix.lower()
    allowed = {".pdf",".xml",".jpg",".jpeg",".png",".webp"}
    if ext not in allowed:
        raise HTTPException(400, "Envie PDF, XML ou imagem da nota fiscal.")
    content = await read_upload_limited(file, MAX_NF_UPLOAD, "Nota fiscal")
    conn = get_db()
    row = _monteiro_invoice_row(conn, gid)
    if not row:
        conn.close()
        raise HTTPException(404, "LanÃ§amento nÃ£o encontrado.")
    old_path = row["invoice_file_path"]
    safe_gid = re.sub(r"[^A-Za-z0-9_-]+", "_", gid)
    fname = f"{safe_gid}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
    dest = MONTEIRO_NOTES_DIR / fname
    with open(dest, "wb") as f:
        f.write(content)
    if old_path:
        try:
            old = Path(old_path)
            if old.exists() and old.resolve().parent == MONTEIRO_NOTES_DIR.resolve():
                old.unlink()
        except Exception:
            pass
    conn.execute("""UPDATE paladar_sales
        SET invoice_file_path=?, invoice_original_name=?, invoice_mime=?, invoice_uploaded_at=datetime('now')
        WHERE sale_group=?""", (str(dest), original, file.content_type or "application/octet-stream", gid))
    conn.commit(); conn.close()
    return {"ok": True, "filename": original}

@app.get("/api/monteiro/sales/group/{group_id}/invoice")
@app.get("/api/paladar/sales/group/{group_id}/invoice")
def view_monteiro_invoice(group_id: str, x_token: str = Header("")):
    require_auth(x_token)
    conn = get_db()
    row = _monteiro_invoice_row(conn, group_id)
    conn.close()
    if not row or not row["invoice_file_path"]:
        raise HTTPException(404, "Nota fiscal nÃ£o anexada.")
    path = Path(row["invoice_file_path"])
    try:
        resolved = path.resolve()
        if resolved.parent != MONTEIRO_NOTES_DIR.resolve():
            raise HTTPException(403, "Arquivo invÃ¡lido.")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(404, "Arquivo nÃ£o encontrado.")
    if not resolved.exists():
        raise HTTPException(404, "Arquivo nÃ£o encontrado.")
    return FileResponse(str(resolved), media_type=row["invoice_mime"] or "application/octet-stream",
                        filename=row["invoice_original_name"] or resolved.name)

@app.delete("/api/monteiro/sales/group/{group_id}/invoice")
@app.delete("/api/paladar/sales/group/{group_id}/invoice")
def delete_monteiro_invoice(group_id: str, x_token: str = Header("")):
    require_editor_tab_access(x_token,["nf","pendentes","consolidado"])
    conn = get_db()
    row = _monteiro_invoice_row(conn, group_id)
    if not row:
        conn.close()
        raise HTTPException(404, "LanÃ§amento nÃ£o encontrado.")
    path_txt = row["invoice_file_path"]
    conn.execute("""UPDATE paladar_sales
        SET invoice_file_path=NULL, invoice_original_name=NULL, invoice_mime=NULL, invoice_uploaded_at=NULL
        WHERE sale_group=?""", (group_id,))
    conn.commit(); conn.close()
    if path_txt:
        try:
            p = Path(path_txt)
            if p.exists() and p.resolve().parent == MONTEIRO_NOTES_DIR.resolve():
                p.unlink()
        except Exception:
            pass
    return {"ok": True}

@app.post("/api/monteiro/invoices/zip")
@app.post("/api/paladar/invoices/zip")
def download_monteiro_invoices_zip(body: dict, x_token: str = Header("")):
    """Gera um ZIP com as notas fiscais (PDF/imagem/XML) dos grupos informados.
    body: {"groups": ["sale_group", ...]}. Nomeia cada arquivo como data_NF_cliente."""
    require_auth(x_token)
    import io as _io, zipfile
    groups = body.get("groups") or []
    groups = [str(g).strip() for g in groups if str(g).strip()]
    if not groups:
        raise HTTPException(400, "Nenhuma nota selecionada para download.")
    conn = get_db()
    ph = ",".join("?" * len(groups))
    rows = conn.execute(f"""SELECT DISTINCT sale_group, saledate, client, nf_number,
        invoice_file_path, invoice_original_name
        FROM paladar_sales
        WHERE sale_group IN ({ph}) AND invoice_file_path IS NOT NULL
        ORDER BY saledate DESC""", groups).fetchall()
    conn.close()
    if not rows:
        raise HTTPException(404, "Nenhuma nota fiscal anexada nos lanÃ§amentos selecionados.")
    buf = _io.BytesIO()
    used = set(); included = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in rows:
            try:
                p = Path(r["invoice_file_path"])
                if p.resolve().parent != MONTEIRO_NOTES_DIR.resolve() or not p.exists():
                    continue
            except Exception:
                continue
            ext = p.suffix or Path(r["invoice_original_name"] or "").suffix or ""
            cli = re.sub(r"[^A-Za-z0-9]+", "_", (r["client"] or "").strip())[:30].strip("_")
            base = f"{r['saledate'] or 'sem_data'}_{r['nf_number'] or r['sale_group']}"
            if cli:
                base += "_" + cli
            base = re.sub(r"[^A-Za-z0-9._-]+", "_", base)
            name = base + ext
            i = 1
            while name in used:
                name = f"{base}_{i}{ext}"; i += 1
            used.add(name)
            zf.write(str(p), arcname=name); included += 1
    if included == 0:
        raise HTTPException(404, "Os arquivos das notas nÃ£o foram encontrados no servidor.")
    buf.seek(0)
    fname = f"notas_monteiro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return StreamingResponse(buf, media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})

@app.get("/api/monteiro/drivers")
@app.get("/api/paladar/drivers")
def paladar_drivers_list(x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    rows=conn.execute("""
        SELECT name AS delivery_person
        FROM drivers
        WHERE active=1
        UNION
        SELECT DISTINCT delivery_person AS delivery_person
        FROM sales
        WHERE delivery_person IS NOT NULL AND TRIM(delivery_person)!=''
        ORDER BY delivery_person
    """).fetchall()
    conn.close(); return [r["delivery_person"] for r in rows]

@app.get("/api/monteiro/vehicles")
@app.get("/api/paladar/vehicles")
def paladar_vehicles_list(x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    rows=conn.execute("SELECT id,name,plate,driver FROM vehicles WHERE active=1 ORDER BY name").fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.put("/api/monteiro/sales/{pid}")
@app.put("/api/paladar/sales/{pid}")
def update_paladar_sale(pid:str,body:dict,x_token:str=Header("")):
    require_editor_tab_access(x_token,["consolidado","nf","avulso","pendentes"]); conn=get_db()
    product=(body.get("product") or "").strip()
    try:
        quantity=float(body.get("quantity",0))
        unitprice=float(body.get("unitprice",0))
    except Exception:
        conn.close()
        raise HTTPException(400, "Quantidade ou preÃ§o invÃ¡lido.")
    notes=body.get("notes",None)
    if not product:
        conn.close()
        raise HTTPException(400, "Informe o produto.")
    if quantity <= 0:
        conn.close()
        raise HTTPException(400, "Quantidade deve ser maior que zero.")
    if unitprice < 0:
        conn.close()
        raise HTTPException(400, "PreÃ§o nÃ£o pode ser negativo.")
    row=conn.execute("SELECT id FROM paladar_sales WHERE id=?",(pid,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, "Item nÃ£o encontrado.")
    total=round(quantity*unitprice,2)
    if notes is None:
        conn.execute("UPDATE paladar_sales SET product=?, quantity=?, unitprice=?, total=? WHERE id=?",
                     (product,quantity,unitprice,total,pid))
    else:
        conn.execute("UPDATE paladar_sales SET product=?, quantity=?, unitprice=?, total=?, notes=? WHERE id=?",
                     (product,quantity,unitprice,total,notes,pid))
    conn.commit(); conn.close()
    return {"ok":True,"total":total}

@app.put("/api/monteiro/sales/group/{group_id}")
@app.put("/api/paladar/sales/group/{group_id}")
def update_paladar_sale_group(group_id:str,body:dict,x_token:str=Header("")):
    require_editor_tab_access(x_token,["consolidado","nf","avulso","pendentes"]); conn=get_db()
    client=(body.get("client") or "").strip()
    row=conn.execute("SELECT sale_group FROM paladar_sales WHERE sale_group=? LIMIT 1",(group_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, "LanÃ§amento nÃ£o encontrado.")
    conn.execute("UPDATE paladar_sales SET client=? WHERE sale_group=?",(client,group_id))
    conn.commit(); conn.close()
    return {"ok":True,"client":client}

@app.delete("/api/monteiro/sales/{pid}")
@app.delete("/api/paladar/sales/{pid}")
def delete_paladar_sale(pid:str,x_token:str=Header("")):
    require_editor_tab_access(x_token,["consolidado","nf","avulso","pendentes"]); conn=get_db()
    conn.execute("DELETE FROM paladar_sales WHERE id=?",(pid,))
    conn.commit(); conn.close()
    return {"ok":True}

@app.delete("/api/monteiro/sales/group/{group_id}")
@app.delete("/api/paladar/sales/group/{group_id}")
def delete_paladar_sale_group(group_id:str,x_token:str=Header("")):
    require_editor_tab_access(x_token,["consolidado","nf","avulso","pendentes"]); conn=get_db()
    row=_monteiro_invoice_row(conn, group_id)
    path_txt=row["invoice_file_path"] if row else None
    conn.execute("DELETE FROM paladar_sales WHERE sale_group=?",(group_id,))
    conn.commit(); conn.close()
    if path_txt:
        try:
            p=Path(path_txt)
            if p.exists() and p.resolve().parent == MONTEIRO_NOTES_DIR.resolve():
                p.unlink()
        except Exception:
            pass
    return {"ok":True}

@app.get("/api/monteiro/clients/list")
def monteiro_clients_list(period:Optional[str]=None,month:Optional[str]=None,year:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    where, args = _pal_period_where(period, month, year)
    ws=" AND ".join(where) if where else "1"
    rows=conn.execute(f"SELECT DISTINCT client FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {ws} ORDER BY client",args).fetchall()
    conn.close(); return [r["client"] for r in rows]

@app.get("/api/monteiro/clients/summary")
def monteiro_clients_summary(period:Optional[str]=None,month:Optional[str]=None,year:Optional[str]=None,client_filter:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    where, args = _pal_period_where(period, month, year)
    if client_filter: where.append("client like ?"); args.append(f"%{client_filter}%")
    ws=" AND ".join(where) if where else "1"
    # Resumo geral
    row=conn.execute(f"""SELECT
        COALESCE(SUM(total),0) as total_vendido,
        COUNT(DISTINCT client) as total_clientes,
        CASE WHEN COUNT(DISTINCT sale_group) > 0 THEN COALESCE(SUM(total),0)/COUNT(DISTINCT sale_group) ELSE 0 END as ticket_medio
        FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {ws}""",args).fetchone()
    # Cliente lÃ­der
    lider=conn.execute(f"""SELECT client, SUM(total) as total, COUNT(DISTINCT sale_group) as count
        FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {ws}
        GROUP BY client ORDER BY total DESC LIMIT 1""",args).fetchone()
    # Vendas por cliente
    clientes=conn.execute(f"""SELECT client, COUNT(DISTINCT sale_group) as num_vendas, SUM(quantity) as qty_total,
        SUM(total) as total, CASE WHEN COUNT(DISTINCT sale_group)>0 THEN ROUND(SUM(total)/COUNT(DISTINCT sale_group),2) ELSE 0 END as media_venda, MAX(saledate) as ultima_compra
        FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {ws}
        GROUP BY client ORDER BY total DESC""",args).fetchall()
    conn.close()
    return {
        "total_vendido": float(row["total_vendido"]),
        "total_clientes": row["total_clientes"],
        "ticket_medio": float(row["ticket_medio"]),
        "cliente_lider": dict(lider) if lider else None,
        "clientes": [dict(r) for r in clientes]
    }

@app.get("/api/monteiro/clients/history")
def monteiro_clients_history(period:Optional[str]=None,month:Optional[str]=None,year:Optional[str]=None,client_filter:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    where, args = _pal_period_where(period, month, year)
    if client_filter: where.append("client like ?"); args.append(f"%{client_filter}%")
    ws=" AND ".join(where) if where else "1"
    rows=conn.execute(f"""SELECT strftime('%Y-%m',saledate) as mes, COUNT(DISTINCT sale_group) as num_vendas,
        SUM(quantity) as qty_total, SUM(total) as total
        FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {ws}
        GROUP BY mes ORDER BY mes""",args).fetchall()
    conn.close(); return [dict(r) for r in rows]

# â”€â”€ Monteiro: relatÃ³rio detalhado por cliente â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get("/api/monteiro/clients/report")
def monteiro_client_report(client:str="",period:Optional[str]=None,month:Optional[str]=None,year:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token)
    if not client.strip():
        raise HTTPException(400, "Cliente obrigatÃ³rio.")
    conn=get_db()
    where, args = _pal_period_where(period, month, year)
    where.append("client=?")
    args.append(client.strip())
    ws=" AND ".join(where)
    # Resumo
    row=conn.execute(f"""SELECT COALESCE(SUM(total),0) as total_vendido,
        COALESCE(SUM(quantity),0) as qty_total,
        COUNT(DISTINCT sale_group) as num_vendas,
        CASE WHEN COUNT(DISTINCT sale_group)>0 THEN ROUND(SUM(total)/COUNT(DISTINCT sale_group),2) ELSE 0 END as ticket_medio,
        MAX(saledate) as ultima_compra
        FROM paladar_sales WHERE {ws}""",args).fetchone()
    # Produtos consolidados
    prods_raw=conn.execute(f"""SELECT product,
        SUM(quantity) as qty, SUM(total) as total,
        CASE WHEN SUM(quantity)>0 THEN ROUND(SUM(total)/SUM(quantity),2) ELSE 0 END as preco_medio,
        COUNT(DISTINCT sale_group) as ocorrencias
        FROM paladar_sales WHERE {ws}
        GROUP BY product ORDER BY qty DESC""",args).fetchall()
    prod_map={}
    for p in prods_raw:
        name=norm_p(p["product"] or "", "NF")
        if name not in prod_map:
            prod_map[name]={"product":name,"qty":0,"total":0,"ocorrencias":0}
        prod_map[name]["qty"]+=float(p["qty"] or 0)
        prod_map[name]["total"]+=float(p["total"] or 0)
        prod_map[name]["ocorrencias"]+=int(p["ocorrencias"] or 0)
    prods=sorted(prod_map.values(), key=lambda x:x["qty"], reverse=True)
    for p in prods:
        p["preco_medio"]=round((p["total"]/p["qty"]) if p["qty"] else 0,2)
    # Notas
    notas=conn.execute(f"""SELECT sale_group, saledate,
        MAX(nf_number) as nf,
        GROUP_CONCAT(product, ', ') as produtos,
        SUM(quantity) as qty, SUM(total) as total,
        COUNT(*) as itens, MAX(driver) as entregador,
        MAX(vehicle) as veiculo, MAX(notes) as observacoes
        FROM paladar_sales WHERE {ws}
        GROUP BY sale_group ORDER BY saledate DESC""",args).fetchall()
    conn.close()
    r={**dict(row),
        "produto_mais_comprado": prods[0]["product"] if prods else None,
        "produtos": prods,
        "notas": [dict(n) for n in notas],
        "client": client.strip(),
        "period": period or "mensal",
        "emissao": datetime.now().strftime("%d/%m/%Y %H:%M")}
    return r

@app.get("/api/monteiro/clients/report/excel")
def monteiro_client_excel(client:str="",period:Optional[str]=None,month:Optional[str]=None,year:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token)
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    # Busca dados
    conn=get_db()
    where, args = _pal_period_where(period, month, year)
    cli_filter=client.strip()
    if cli_filter:
        where.append("client=?")
        args.append(cli_filter)
    ws_=" AND ".join(where) if where else "1"
    # Produtos
    prods=conn.execute(f"""SELECT client, product, SUM(quantity) as qty, SUM(total) as total,
        CASE WHEN SUM(quantity)>0 THEN ROUND(SUM(total)/SUM(quantity),2) ELSE 0 END as preco_medio,
        COUNT(DISTINCT sale_group) as ocorrencias
        FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {ws_}
        GROUP BY client, product ORDER BY client, qty DESC""",args).fetchall()
    # Notas
    notas=conn.execute(f"""SELECT client, sale_group, saledate, MAX(nf_number) as nf,
        GROUP_CONCAT(product, ', ') as produtos, SUM(quantity) as qty, SUM(total) as total,
        COUNT(*) as itens, MAX(driver) as entregador, MAX(vehicle) as veiculo
        FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {ws_}
        GROUP BY client, sale_group ORDER BY client, saledate DESC""",args).fetchall()
    # Resumo por cliente
    resumo=conn.execute(f"""SELECT client, COALESCE(SUM(total),0) as total_vendido,
        COALESCE(SUM(quantity),0) as qty_total, COUNT(DISTINCT sale_group) as num_vendas
        FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {ws_}
        GROUP BY client ORDER BY total_vendido DESC""",args).fetchall()
    conn.close()
    wb=Workbook()
    # Estilos
    hdr=Font(bold=True,color="FFFFFF",size=11)
    hdr_fill=PatternFill("solid",fgColor="4F46E5")
    bdr=Border(bottom=Side(style='thin',color='E5E7EB'))
    money_fmt='#.##0,00'
    # â”€â”€ Sheet 1: Resumo â”€â”€
    ws1=wb.active; ws1.title="Resumo"
    ws1.append(["Cliente","Total Vendido","Qtd Itens","NÂº Vendas","Ticket MÃ©dio"])
    ws1.cell(row=1,column=1).font=hdr; ws1.cell(row=1,column=1).fill=hdr_fill
    ws1.cell(row=1,column=2).font=hdr; ws1.cell(row=1,column=2).fill=hdr_fill
    ws1.cell(row=1,column=3).font=hdr; ws1.cell(row=1,column=3).fill=hdr_fill
    ws1.cell(row=1,column=4).font=hdr; ws1.cell(row=1,column=4).fill=hdr_fill
    ws1.cell(row=1,column=5).font=hdr; ws1.cell(row=1,column=5).fill=hdr_fill
    for r in resumo:
        tm=round(r["total_vendido"]/r["num_vendas"],2) if r["num_vendas"] else 0
        ws1.append([r["client"], round(r["total_vendido"],2), round(r["qty_total"],1), r["num_vendas"], tm])
    ws1.column_dimensions['A'].width=30; ws1.column_dimensions['B'].width=18
    ws1.column_dimensions['C'].width=12; ws1.column_dimensions['D'].width=12; ws1.column_dimensions['E'].width=16
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, max_col=5):
        for c in row: c.border=bdr
    # â”€â”€ Sheet 2: Produtos â”€â”€
    ws2=wb.create_sheet("Produtos")
    ws2.append(["Cliente","Produto","Quantidade","Valor Total","PreÃ§o MÃ©dio","OcorrÃªncias"])
    for i in range(1,7):
        ws2.cell(row=1,column=i).font=hdr; ws2.cell(row=1,column=i).fill=hdr_fill
    for p in prods:
        ws2.append([p["client"],p["product"],round(p["qty"],1),round(p["total"],2),p["preco_medio"],p["ocorrencias"]])
    ws2.column_dimensions['A'].width=25; ws2.column_dimensions['B'].width=25
    ws2.column_dimensions['C'].width=12; ws2.column_dimensions['D'].width=16
    ws2.column_dimensions['E'].width=14; ws2.column_dimensions['F'].width=12
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=6):
        for c in row: c.border=bdr
    # â”€â”€ Sheet 3: Notas â”€â”€
    ws3=wb.create_sheet("Notas")
    ws3.append(["Cliente","Data","NF","Produtos","Qtd Itens","Valor","Entregador","VeÃ­culo"])
    for i in range(1,9):
        ws3.cell(row=1,column=i).font=hdr; ws3.cell(row=1,column=i).fill=hdr_fill
    for n in notas:
        ws3.append([n["client"],n["saledate"],n["nf"] or "",n["produtos"] or "",
                    round(n["qty"],1),round(n["total"],2),n["entregador"] or "",n["veiculo"] or ""])
    ws3.column_dimensions['A'].width=25; ws3.column_dimensions['B'].width=12; ws3.column_dimensions['C'].width=16
    ws3.column_dimensions['D'].width=40; ws3.column_dimensions['E'].width=10
    ws3.column_dimensions['F'].width=14; ws3.column_dimensions['G'].width=20; ws3.column_dimensions['H'].width=16
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, max_col=8):
        for c in row: c.border=bdr
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=monteiro_clientes.xlsx"})

def _monteiro_report_data(client, period, month, year):
    from datetime import datetime as _dt
    _t0=_dt.now()
    print(f"[_monteiro_report_data] iniciando: client={client!r} period={period!r} month={month!r} year={year!r}")
    try:
        conn = get_db()
        where, args = _pal_period_where(period, month, year)
        cli_filter = client.strip()
        if cli_filter:
            where.append("client=?")
            args.append(cli_filter)
        ws = " AND ".join(where) if where else "1"
        sql=f"SELECT client, product, sale_group, saledate, nf_number, quantity, total FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {ws} ORDER BY client, saledate"
        print(f"[_monteiro_report_data] SQL: {sql}")
        print(f"[_monteiro_report_data] args: {args}")
        rows = conn.execute(sql, args).fetchall()
        conn.close()
        _elapsed=(_dt.now()-_t0).total_seconds()
        print(f"[_monteiro_report_data] OK: {len(rows)} linhas em {_elapsed:.3f}s")
        return [dict(r) for r in rows]
    except Exception as e:
        import traceback
        _elapsed=(_dt.now()-_t0).total_seconds()
        print(f"[_monteiro_report_data] *** EXCECAO *** apos {_elapsed:.3f}s: {e}")
        traceback.print_exc()
        raise

def _print_page(title, body, period_str=""):
    """Retorna pÃ¡gina HTML completa para impressÃ£o com CSS e auto-print."""
    from datetime import datetime
    agora = datetime.now().strftime('%d/%m/%Y %H:%M')
    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head><meta charset="utf-8">
<title>{title}</title>
<style>
  @page {{ size: A4; margin: 12mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: Arial, Helvetica, sans-serif; font-size: 11px; color: #222; margin: 0; padding: 20px; }}
  h1 {{ font-size: 17px; color: #4F46E5; margin: 0 0 3px; }}
  .sub {{ font-size: 10px; color: #888; margin-bottom: 14px; }}
  .ch {{ font-size: 13px; font-weight: 700; color: #4F46E5; background: #EEF2FF; padding: 6px 10px; border-radius: 4px; margin: 14px 0 6px; }}
  .sm {{ font-size: 10px; color: #555; margin-bottom: 8px; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 12px; font-size: 10px; }}
  th, td {{ padding: 4px 6px; border: 1px solid #d0d0d0; text-align: left; }}
  th {{ background: #4F46E5; color: #fff; font-weight: 600; }}
  tr:nth-child(even) {{ background: #f8f8ff; }}
  .r {{ text-align: right; font-family: 'Courier New', monospace; }}
  .s {{ font-size: 12px; font-weight: 700; margin: 6px 0; }}
  .s0 {{ color: #16a34a; }} .s1 {{ color: #dc2626; }}
  .nd {{ color: #999; font-style: italic; padding: 8px 0; }}
  .ft {{ font-size: 8px; color: #bbb; margin-top: 20px; padding-top: 6px; border-top: 1px solid #eee; text-align: center; }}
</style>
</head><body>
<h1>{title}</h1>
<div class="sub">PerÃ­odo: {_safe_txt(period_str)} â€” Gerado em {agora}</div>
{body}
<div class="ft">Menina dos Raios Ltda Â· {agora}</div>
<script>window.onload=function(){{setTimeout(function(){{window.print();window.close();}},500);}};</script>
</body></html>"""

@app.get("/api/monteiro/clients/report/print")
def monteiro_client_print(client:str="",period:Optional[str]=None,month:Optional[str]=None,year:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token)
    from collections import defaultdict, OrderedDict
    from datetime import datetime as _dt
    print(f"[monteiro_client_print] cl={client!r} per={period!r} m={month!r} a={year!r}")
    period_str = f"{period or 'mensal'}" + (f" {month}/{year}" if month or year else "")
    try:
        rows = _monteiro_report_data(client, period, month, year)
    except Exception as e:
        import traceback; traceback.print_exc()
        rows = []
    try:
        if not rows:
            body = '<div class="nd">Nenhum cliente encontrado no perÃ­odo.</div>'
        else:
            clientes = OrderedDict()
            for r in rows:
                c = _safe_txt(r.get("client") or "", "SEM NOME")
                if c not in clientes:
                    clientes[c] = {"prods": defaultdict(lambda: {"q": 0.0, "t": 0.0}), "grps": set()}
                clientes[c]["grps"].add(str(r.get("sale_group") or ""))
                p = clientes[c]["prods"][_safe_txt(r.get("product"), "SEM NOME")]
                p["q"] += float(r.get("quantity") or 0); p["t"] += float(r.get("total") or 0)
            parts = []
            for nome in sorted(clientes.keys()):
                d = clientes[nome]
                prods = d["prods"]
                total_v = round(sum(p["t"] for p in prods.values()), 2)
                qty_t = round(sum(p["q"] for p in prods.values()), 1)
                grps = d["grps"] - {""}
                num_v = len(grps) if grps else 1
                tm = round(total_v / num_v, 2) if num_v else 0
                parts.append(f'<div class="ch">{nome}</div>')
                parts.append(f'<div class="sm">Total: R$ {total_v:,.2f} &nbsp;|&nbsp; Itens: {qty_t:.0f} &nbsp;|&nbsp; Vendas: {num_v} &nbsp;|&nbsp; Ticket MÃ©dio: R$ {tm:,.2f}</div>')
                if prods:
                    parts.append('<table><tr><th>Produto</th><th class="r">Qtd</th><th class="r">Total R$</th><th class="r">MÃ©dio R$</th></tr>')
                    for pnome in sorted(prods.keys()):
                        p = prods[pnome]
                        pm = round(p["t"] / p["q"], 2) if p["q"] else 0
                        parts.append(f'<tr><td>{_safe_txt(pnome[:45])}</td><td class="r">{p["q"]:.1f}</td><td class="r">{p["t"]:,.2f}</td><td class="r">{pm:,.2f}</td></tr>')
                    parts.append('</table>')
                else:
                    parts.append('<div class="nd">Nenhum produto encontrado.</div>')
            body = "".join(parts)
        html = _print_page("Monteiro â€” Clientes", body, period_str)
        return HTMLResponse(content=html, status_code=200)
    except Exception as e:
        import traceback; traceback.print_exc()
        body = f'<div class="nd">Erro ao gerar relatÃ³rio. Detalhe: {_safe_txt(str(e)[:150])}</div>'
        html = _print_page("Monteiro â€” Clientes", body, period_str)
        return HTMLResponse(content=html, status_code=200)

@app.get("/api/monteiro/payments/report/print")
def monteiro_payments_print(client:str="",period:Optional[str]=None,month:str="",year:str="",x_token:str=Header("")):
    require_auth(x_token)
    from collections import defaultdict
    from datetime import datetime as _dt
    print(f"[monteiro_payments_print] cl={client!r} per={period!r} m={month!r} a={year!r}")
    period_str = f"{period or 'mensal'}" + (f" {month}/{year}" if month or year else "")
    try:
        pm, py = _pay_period_map(period, month, year)
        sw, sa = _pal_period_where(period, pm, py)
        if client.strip(): sw.append("client=?"); sa.append(client.strip())
        sws = " AND ".join(sw) if sw else "1"
        pw, pa = [], []
        if client.strip(): pw.append("client=?"); pa.append(client.strip())
        if pm: pw.append("month=?"); pa.append(pm)
        if py: pw.append("year=?"); pa.append(py)
        pws = " AND ".join(pw) if pw else "1"
        conn = get_db()
        vendas_raw = conn.execute(f"""SELECT client, sale_group, saledate, nf_number, SUM(quantity) as qty, SUM(total) as total
            FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {sws}
            GROUP BY client, sale_group ORDER BY client, saledate""", sa).fetchall()
        pags_raw = conn.execute(f"""SELECT * FROM monteiro_payments WHERE {pws} ORDER BY client, payment_date""", pa).fetchall()
        conn.close()
        vendas = [dict(r) for r in vendas_raw]
        pags = [dict(r) for r in pags_raw]
    except Exception as e:
        import traceback; traceback.print_exc()
        vendas = []; pags = []
    try:
        cli_notas = defaultdict(list); cli_pags = defaultdict(list)
        for v in vendas: cli_notas[_safe_txt(v["client"], "SEM NOME")].append(v)
        for p in pags: cli_pags[_safe_txt(p["client"], "SEM NOME")].append(p)
        todos = sorted(set(list(cli_notas.keys()) + list(cli_pags.keys())))
        if not todos:
            body = '<div class="nd">Nenhum dado encontrado no perÃ­odo.</div>'
        else:
            parts = []
            for cli in todos:
                nts = cli_notas.get(cli, [])
                pp = cli_pags.get(cli, [])
                total_notas = sum(float(n.get("total") or 0) for n in nts)
                total_pago = sum(float(p.get("amount") or 0) for p in pp)
                saldo = round(total_notas - total_pago, 2)
                sclass = "s0" if saldo <= 0 else "s1"
                parts.append(f'<div class="ch">{cli}</div>')
                parts.append(f'<div class="sm">Total Vendido: R$ {total_notas:,.2f} &nbsp;|&nbsp; Total Recebido: R$ {total_pago:,.2f}</div>')
                parts.append(f'<div class="s {sclass}">Saldo: R$ {saldo:,.2f}</div>')
                if nts:
                    parts.append('<table><tr><th>Data</th><th>NF</th><th class="r">Qtd</th><th class="r">Valor</th></tr>')
                    for n in nts:
                        parts.append(f'<tr><td>{_safe_txt(n.get("saledate"))}</td><td>{_safe_txt(n.get("nf_number"))}</td><td class="r">{float(n.get("qty") or 0):.1f}</td><td class="r">{float(n.get("total") or 0):.2f}</td></tr>')
                    parts.append('</table>')
                else:
                    parts.append('<div class="nd">Nenhuma nota no perÃ­odo.</div>')
                if pp:
                    parts.append('<table><tr><th>Data</th><th class="r">Valor</th><th>Tipo</th><th>Obs</th></tr>')
                    for p in pp:
                        parts.append(f'<tr><td>{_safe_txt(p.get("payment_date"))}</td><td class="r">{float(p.get("amount") or 0):.2f}</td><td>{_safe_txt(p.get("payment_type"))}</td><td>{_safe_txt(p.get("notes",""))[:40]}</td></tr>')
                    parts.append('</table>')
                else:
                    parts.append('<div class="nd">Nenhum pagamento registrado.</div>')
            body = "".join(parts)
        html = _print_page("Monteiro â€” Pagamentos", body, period_str)
        return HTMLResponse(content=html, status_code=200)
    except Exception as e:
        import traceback; traceback.print_exc()
        body = f'<div class="nd">Erro ao gerar relatÃ³rio. Detalhe: {_safe_txt(str(e)[:150])}</div>'
        html = _print_page("Monteiro â€” Pagamentos", body, period_str)
        return HTMLResponse(content=html, status_code=200)

# â”€â”€ wkhtmltopdf: helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_WKHTMLTOPDF_BIN = ""  # descoberto automaticamente na primeira chamada

def _wkhtmltopdf_path():
    global _WKHTMLTOPDF_BIN
    if _WKHTMLTOPDF_BIN:
        return _WKHTMLTOPDF_BIN
    candidatos = [
        "wkhtmltopdf",
        "/usr/bin/wkhtmltopdf",
        "/usr/local/bin/wkhtmltopdf",
        "/usr/local/sbin/wkhtmltopdf",
        "/opt/wkhtmltopdf/wkhtmltopdf",
        "/snap/bin/wkhtmltopdf",
        "/usr/bin/wkhtmltox/bin/wkhtmltopdf",
        "/opt/wkhtmltox/bin/wkhtmltopdf",
    ]
    import shutil
    for c in candidatos:
        p = shutil.which(c) or (c if os.path.isfile(c) else None)
        if p and os.access(p, os.X_OK):
            _WKHTMLTOPDF_BIN = p
            print(f"[wkhtmltopdf] binario encontrado: {p}")
            return p
    print("[wkhtmltopdf] BINARIO NAO ENCONTRADO em nenhum caminho comum")
    return ""

def _html_to_pdf_wk_unlocked(html: str) -> Optional[bytes]:
    """Converte HTML em PDF via wkhtmltopdf com validacao do arquivo gerado.
    Retorna bytes do PDF, ou None se falhar em qualquer etapa."""
    import subprocess, tempfile, os, traceback
    bin_path = _wkhtmltopdf_path()
    if not bin_path:
        print("[wkhtmltopdf] sem binario â€” pulando conversao")
        return None
    html_path = pdf_path = None
    tmp_prefix = "monteiro_pdf_"
    print(f"[wkhtmltopdf] bin={bin_path} html_len={len(html)}")
    try:
        # â”€â”€ escreve HTML temporario â”€â”€
        hf = tempfile.NamedTemporaryFile(suffix=".html", prefix=tmp_prefix, delete=False, mode="w", encoding="utf-8")
        hf.write(html); hf.close()
        html_path = hf.name
        # â”€â”€ prepara PDF temporario â”€â”€
        pf = tempfile.NamedTemporaryFile(suffix=".pdf", prefix=tmp_prefix, delete=False)
        pf.close()
        pdf_path = pf.name
        print(f"[wkhtmltopdf] html_tmp={html_path} pdf_tmp={pdf_path}")
        # â”€â”€ comando wkhtmltopdf (apenas flags essenciais) â”€â”€
        cmd = [
            bin_path,
            "--encoding", "UTF-8",
            "--page-size", "A4",
            "--margin-top", "10mm",
            "--margin-bottom", "10mm",
            "--margin-left", "10mm",
            "--margin-right", "10mm",
            "--enable-local-file-access",
            html_path,
            pdf_path,
        ]
        print(f"[wkhtmltopdf] cmd={' '.join(cmd)}")
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        print(f"[wkhtmltopdf] returncode={r.returncode}")
        print(f"[wkhtmltopdf] stdout={r.stdout.strip()[:500]!r}" if r.stdout.strip() else "[wkhtmltopdf] stdout=(vazio)")
        if r.stderr.strip():
            for line in r.stderr.strip().split("\n"):
                print(f"[wkhtmltopdf] stderr: {line}")
        else:
            print("[wkhtmltopdf] stderr=(vazio)")
        if r.returncode != 0:
            print(f"[wkhtmltopdf] ERRO: codigo {r.returncode}")
            return None
        # â”€â”€ valida existencia e tamanho â”€â”€
        if not os.path.isfile(pdf_path):
            print("[wkhtmltopdf] ERRO: arquivo PDF nao foi criado")
            return None
        pdf_size = os.path.getsize(pdf_path)
        if pdf_size == 0:
            print("[wkhtmltopdf] ERRO: PDF tem 0 bytes")
            return None
        # â”€â”€ valida header %PDF â”€â”€
        with open(pdf_path, "rb") as f:
            header = f.read(4)
            data = f.read()
        print(f"[wkhtmltopdf] pdf_size={pdf_size} pdf_header={header!r}")
        if header != b"%PDF":
            print(f"[wkhtmltopdf] ERRO: header invalido ({header!r}), nao e um PDF")
            return None
        print(f"[wkhtmltopdf] OK: {len(data)+4} bytes, header valido")
        return header + data
    except subprocess.TimeoutExpired:
        print("[wkhtmltopdf] TIMEOUT: processo excedeu 60s")
        return None
    except Exception as e:
        print(f"[wkhtmltopdf] EXCECAO: {e}")
        traceback.print_exc()
        return None
    finally:
        for p in [html_path, pdf_path]:
            if p:
                try:
                    if os.path.isfile(p): os.remove(p)
                except Exception:
                    pass

def _html_to_pdf_wk(html: str) -> Optional[bytes]:
    with pdf_generation_slot():
        return _html_to_pdf_wk_unlocked(html)

# â”€â”€ Monteiro: PDF via wkhtmltopdf â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.get("/api/monteiro/clients/report/pdf")
def monteiro_client_pdf_wk(client:str="",period:Optional[str]=None,month:Optional[str]=None,year:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token)
    from collections import defaultdict, OrderedDict
    from datetime import datetime as _dt
    _t0 = _dt.now()
    print(f"[monteiro_client_pdf] === ROTA CHAMADA === {_t0.isoformat()}")
    print(f"[monteiro_client_pdf] params: client={client!r} period={period!r} month={month!r} year={year!r}")
    period_str = f"{period or 'mensal'}" + (f" {month}/{year}" if month or year else "")
    try:
        rows = _monteiro_report_data(client, period, month, year)
        print(f"[monteiro_client_pdf] dados brutos: {len(rows)} linhas")
    except Exception as e:
        import traceback
        print(f"[monteiro_client_pdf] ERRO ao buscar dados: {e}")
        traceback.print_exc()
        rows = []
    try:
        # â”€â”€ Monta HTML (mesma logica do print) â”€â”€
        if not rows:
            body = '<div class="nd">Nenhum cliente encontrado no perÃ­odo.</div>'
        else:
            clientes = OrderedDict()
            for r in rows:
                c = _safe_txt(r.get("client") or "", "SEM NOME")
                if c not in clientes:
                    clientes[c] = {"prods": defaultdict(lambda: {"q": 0.0, "t": 0.0}), "grps": set()}
                clientes[c]["grps"].add(str(r.get("sale_group") or ""))
                p = clientes[c]["prods"][_safe_txt(r.get("product"), "SEM NOME")]
                p["q"] += float(r.get("quantity") or 0); p["t"] += float(r.get("total") or 0)
            parts = []
            for nome in sorted(clientes.keys()):
                d = clientes[nome]
                prods = d["prods"]
                total_v = round(sum(p["t"] for p in prods.values()), 2)
                qty_t = round(sum(p["q"] for p in prods.values()), 1)
                grps = d["grps"] - {""}
                num_v = len(grps) if grps else 1
                tm = round(total_v / num_v, 2) if num_v else 0
                parts.append(f'<div class="ch">{nome}</div>')
                parts.append(f'<div class="sm">Total: R$ {total_v:,.2f} | Itens: {qty_t:.0f} | Vendas: {num_v} | Ticket Medio: R$ {tm:,.2f}</div>')
                if prods:
                    parts.append('<table><tr><th>Produto</th><th class="r">Qtd</th><th class="r">Total R$</th><th class="r">Medio R$</th></tr>')
                    for pnome in sorted(prods.keys()):
                        p = prods[pnome]
                        pm = round(p["t"] / p["q"], 2) if p["q"] else 0
                        parts.append(f'<tr><td>{_safe_txt(pnome[:45])}</td><td class="r">{p["q"]:.1f}</td><td class="r">{p["t"]:,.2f}</td><td class="r">{pm:,.2f}</td></tr>')
                    parts.append('</table>')
                else:
                    parts.append('<div class="nd">Nenhum produto encontrado.</div>')
            body = "".join(parts)
        html = _print_page("Monteiro â€” Clientes", body, period_str)
        print(f"[monteiro_client_pdf] HTML gerado: {len(html)} bytes")
        # â”€â”€ Converte para PDF â”€â”€
        pdf_bytes = _html_to_pdf_wk(html)
        if pdf_bytes:
            _elapsed = (_dt.now() - _t0).total_seconds()
            print(f"[monteiro_client_pdf] PDF OK: {len(pdf_bytes)}b em {_elapsed:.1f}s")
            return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf",
                headers={"Content-Disposition": "attachment; filename=monteiro_clientes.pdf"})
        # â”€â”€ Fallback: tentar weasyprint â”€â”€
        print("[monteiro_client_pdf] wkhtmltopdf falhou, tentando weasyprint...")
        try:
            from weasyprint import HTML as WPHTML
            with pdf_generation_slot():
                pdf_bytes = WPHTML(string=html).write_pdf()
            if pdf_bytes and len(pdf_bytes) > 0:
                _elapsed = (_dt.now() - _t0).total_seconds()
                print(f"[monteiro_client_pdf] WeasyPrint OK: {len(pdf_bytes)}b em {_elapsed:.1f}s")
                return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=monteiro_clientes.pdf"})
        except Exception as e2:
            print(f"[monteiro_client_pdf] WeasyPrint tambem falhou: {e2}")
        # â”€â”€ Fallback final: JSON com erro (NUNCA retornar HTML como se fosse PDF) â”€â”€
        _elapsed = (_dt.now() - _t0).total_seconds()
        print(f"[monteiro_client_pdf] AMBAS conversoes falharam ({_elapsed:.1f}s)")
        print(f"[monteiro_client_pdf] Dica: verifique se wkhtmltopdf esta instalado e funcional no servidor")
        return JSONResponse({
            "detail": "Nao foi possivel gerar o PDF. Verifique se wkhtmltopdf esta instalado no servidor.",
            "erro": "wkhtmltopdf e weasyprint indisponiveis"
        }, status_code=500)
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"[monteiro_client_pdf] EXCECAO GERAL: {e}")
        traceback.print_exc()
        _elapsed = (_dt.now() - _t0).total_seconds()
        return JSONResponse({
            "detail": f"Erro interno ao gerar PDF: {_safe_txt(str(e)[:200])}",
            "erro": "excecao_geral"
        }, status_code=500)

@app.get("/api/monteiro/payments/report/pdf")
def monteiro_payments_pdf_wk(client:str="",period:Optional[str]=None,month:str="",year:str="",x_token:str=Header("")):
    require_auth(x_token)
    from collections import defaultdict
    from datetime import datetime as _dt
    _t0 = _dt.now()
    print(f"[monteiro_payments_pdf] === ROTA CHAMADA === {_t0.isoformat()}")
    print(f"[monteiro_payments_pdf] params: client={client!r} period={period!r} month={month!r} year={year!r}")
    period_str = f"{period or 'mensal'}" + (f" {month}/{year}" if month or year else "")
    try:
        pm, py = _pay_period_map(period, month, year)
        sw, sa = _pal_period_where(period, pm, py)
        if client.strip(): sw.append("client=?"); sa.append(client.strip())
        sws = " AND ".join(sw) if sw else "1"
        pw, pa = [], []
        if client.strip(): pw.append("client=?"); pa.append(client.strip())
        if pm: pw.append("month=?"); pa.append(pm)
        if py: pw.append("year=?"); pa.append(py)
        pws = " AND ".join(pw) if pw else "1"
        conn = get_db()
        vendas_raw = conn.execute(f"""SELECT client, sale_group, saledate, nf_number, SUM(quantity) as qty, SUM(total) as total
            FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {sws}
            GROUP BY client, sale_group ORDER BY client, saledate""", sa).fetchall()
        pags_raw = conn.execute(f"""SELECT * FROM monteiro_payments WHERE {pws} ORDER BY client, payment_date""", pa).fetchall()
        conn.close()
        # Converter sqlite3.Row para dict (sqlite3.Row nao suporta .get())
        vendas = [dict(r) for r in vendas_raw]
        pags = [dict(r) for r in pags_raw]
        print(f"[monteiro_payments_pdf] vendas={len(vendas)} pagamentos={len(pags)} vendas_type={type(vendas_raw).__name__ if vendas_raw else 'empty'}")
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"[monteiro_payments_pdf] ERRO SQL: {e}")
        vendas = []; pags = []
    try:
        cli_notas = defaultdict(list); cli_pags = defaultdict(list)
        for v in vendas: cli_notas[_safe_txt(v.get("client"), "SEM NOME")].append(v)
        for p in pags: cli_pags[_safe_txt(p.get("client"), "SEM NOME")].append(p)
        todos = sorted(set(list(cli_notas.keys()) + list(cli_pags.keys())))
        print(f"[monteiro_payments_pdf] clientes={len(todos)}")
        def _money_br(v):
            return f"R$ {float(v or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        def _date_obj(s):
            try:
                return _dt.strptime(str(s or "")[:10], "%Y-%m-%d")
            except Exception:
                return None
        def _date_short(s):
            d = _date_obj(s) if not hasattr(s, "strftime") else s
            return d.strftime("%d/%m") if d else "-"
        def _range_label(a, b):
            aa = _date_short(a); bb = _date_short(b or a)
            return aa if aa == bb else f"{aa} a {bb}"
        def _paid_bands_html(nts, pp):
            notes = []
            for n in sorted(nts, key=lambda x: str(x.get("saledate") or "")):
                d = _date_obj(n.get("saledate"))
                total = float(n.get("total") or 0)
                if d and total > 0:
                    notes.append({"date": d, "raw": n.get("saledate"), "nf": n.get("nf_number") or "", "total": total, "left": total})
            payments = []
            for p in sorted(pp, key=lambda x: str(x.get("payment_date") or "")):
                d = _date_obj(p.get("payment_date"))
                amount = float(p.get("amount") or 0)
                if d and amount > 0:
                    payments.append({"date": d, "raw": p.get("payment_date"), "amount": amount, "left": amount, "type": p.get("payment_type") or ""})
            bands = []
            idx = 0
            for pg in payments:
                left = pg["amount"]
                band = {"kind": "paid", "payment": pg["raw"], "amount": pg["amount"], "allocated": 0.0, "start": None, "end": None, "nfs": [], "partial": False, "credit": 0.0}
                while left > 0 and idx < len(notes):
                    n = notes[idx]
                    if n["left"] <= 0:
                        idx += 1
                        continue
                    use = min(left, n["left"])
                    band["start"] = band["start"] or n["date"]
                    band["end"] = n["date"]
                    band["allocated"] += use
                    left -= use
                    n["left"] -= use
                    if n["nf"] and n["nf"] not in band["nfs"]:
                        band["nfs"].append(str(n["nf"]))
                    if n["left"] > 0:
                        band["partial"] = True
                        break
                    idx += 1
                band["credit"] = left
                if band["allocated"] > 0 or band["credit"] > 0:
                    bands.append(band)
            remaining = [n for n in notes if n["left"] > 0]
            if remaining:
                bands.append({
                    "kind": "open",
                    "amount": sum(n["left"] for n in remaining),
                    "start": remaining[0]["date"],
                    "end": remaining[-1]["date"],
                    "nfs": [str(n["nf"]) for n in remaining if n["nf"]],
                    "count": len(remaining)
                })
            if not bands:
                return '<div class="nd">Sem notas ou pagamentos para montar cobertura.</div>'
            out = ['<div class="sm"><b>Dias cobertos por pagamento</b></div><table><tr><th>Status</th><th>PerÃ­odo coberto</th><th class="r">Valor</th><th>Detalhe</th></tr>']
            for b in bands:
                if b["kind"] == "open":
                    ref_nf = ", ".join(b.get("nfs", [])[:8])
                    detalhe = "Valor pendente no perÃ­odo" + (f" Â· ref. NF {ref_nf}" if ref_nf else "")
                    out.append(f'<tr><td style="color:#991b1b;font-weight:700">Em aberto</td><td>{_range_label(b["start"], b["end"])}</td><td class="r" style="color:#991b1b;font-weight:700">{_money_br(b["amount"])}</td><td>{_safe_txt(detalhe)}</td></tr>')
                else:
                    status = "Pago parcialmente" if b.get("partial") else "Pago"
                    detalhe = f'Pagamento em {_date_short(b.get("payment"))} Â· recebido {_money_br(b.get("amount"))}'
                    if b.get("credit", 0) > 0:
                        detalhe += f' Â· crÃ©dito sobrando {_money_br(b.get("credit"))}'
                    if b.get("nfs"):
                        detalhe += " Â· NF " + ", ".join(b["nfs"][:8])
                    color = "#92400e" if b.get("partial") else "#166534"
                    out.append(f'<tr><td style="color:{color};font-weight:700">{status}</td><td>{_range_label(b.get("start"), b.get("end"))}</td><td class="r" style="color:{color};font-weight:700">{_money_br(b.get("allocated") or b.get("amount"))}</td><td>{_safe_txt(detalhe)}</td></tr>')
            out.append("</table>")
            return "".join(out)
        if not todos:
            body = '<div class="nd">Nenhum dado encontrado no periodo.</div>'
        else:
            parts = []
            for cli in todos:
                nts = cli_notas.get(cli, [])
                pp = cli_pags.get(cli, [])
                total_notas = sum(float(n.get("total") or 0) for n in nts)
                total_pago = sum(float(p.get("amount") or 0) for p in pp)
                saldo = round(total_notas - total_pago, 2)
                sclass = "s0" if saldo <= 0 else "s1"
                parts.append(f'<div class="ch">{cli}</div>')
                parts.append(f'<div class="sm">Total Vendido: {_money_br(total_notas)} | Total Recebido: {_money_br(total_pago)} | Notas: {len(nts)} | Pagamentos: {len(pp)}</div>')
                parts.append(f'<div class="s {sclass}">Saldo: {_money_br(max(saldo, 0))}</div>')
                parts.append(_paid_bands_html(nts, pp))
                if nts:
                    parts.append('<table><tr><th>Data</th><th>NF</th><th class="r">Qtd</th><th class="r">Valor</th></tr>')
                    for n in nts:
                        parts.append(f'<tr><td>{_safe_txt(n.get("saledate"))}</td><td>{_safe_txt(n.get("nf_number"))}</td><td class="r">{float(n.get("qty") or 0):.1f}</td><td class="r">{_money_br(n.get("total"))}</td></tr>')
                    parts.append('</table>')
                else:
                    parts.append('<div class="nd">Nenhuma nota no periodo.</div>')
                if pp:
                    parts.append('<table><tr><th>Data</th><th class="r">Valor</th><th>Tipo</th><th>Obs</th></tr>')
                    for p in pp:
                        parts.append(f'<tr><td>{_safe_txt(p.get("payment_date"))}</td><td class="r">{_money_br(p.get("amount"))}</td><td>{_safe_txt(p.get("payment_type"))}</td><td>{_safe_txt(p.get("notes",""))[:80]}</td></tr>')
                    parts.append('</table>')
                else:
                    parts.append('<div class="nd">Nenhum pagamento registrado.</div>')
            body = "".join(parts)
        html = _print_page("Monteiro â€” Pagamentos", body, period_str)
        print(f"[monteiro_payments_pdf] HTML gerado: {len(html)} bytes")
        # â”€â”€ Converte para PDF â”€â”€
        pdf_bytes = _html_to_pdf_wk(html)
        if pdf_bytes:
            _elapsed = (_dt.now() - _t0).total_seconds()
            print(f"[monteiro_payments_pdf] PDF OK: {len(pdf_bytes)}b em {_elapsed:.1f}s")
            return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf",
                headers={"Content-Disposition": "attachment; filename=monteiro_pagamentos.pdf"})
        # â”€â”€ Fallback: weasyprint â”€â”€
        print("[monteiro_payments_pdf] wkhtmltopdf falhou, tentando weasyprint...")
        try:
            from weasyprint import HTML as WPHTML
            with pdf_generation_slot():
                pdf_bytes = WPHTML(string=html).write_pdf()
            if pdf_bytes and len(pdf_bytes) > 0:
                _elapsed = (_dt.now() - _t0).total_seconds()
                print(f"[monteiro_payments_pdf] WeasyPrint OK: {len(pdf_bytes)}b em {_elapsed:.1f}s")
                return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=monteiro_pagamentos.pdf"})
        except Exception as e2:
            print(f"[monteiro_payments_pdf] WeasyPrint tambem falhou: {e2}")
        # â”€â”€ Fallback final: JSON com erro (NUNCA retornar HTML como se fosse PDF) â”€â”€
        _elapsed = (_dt.now() - _t0).total_seconds()
        print(f"[monteiro_payments_pdf] AMBAS conversoes falharam ({_elapsed:.1f}s)")
        return JSONResponse({
            "detail": "Nao foi possivel gerar o PDF. Verifique se wkhtmltopdf esta instalado no servidor.",
            "erro": "wkhtmltopdf e weasyprint indisponiveis"
        }, status_code=500)
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"[monteiro_payments_pdf] EXCECAO GERAL: {e}")
        traceback.print_exc()
        return JSONResponse({
            "detail": f"Erro interno ao gerar PDF: {_safe_txt(str(e)[:200])}",
            "erro": "excecao_geral"
        }, status_code=500)

# â”€â”€ Monteiro: Pagamentos â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get("/api/monteiro/payments")
def list_monteiro_payments(client:str="",period:Optional[str]=None,month:str="",year:str="",x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    where=[]; args=[]
    if client.strip(): where.append("client=?"); args.append(client.strip())
    pm,py=_pay_period_map(period,month,year)
    if pm: where.append("month=?"); args.append(pm)
    if py: where.append("year=?"); args.append(py)
    ws=" AND ".join(where) if where else "1"
    rows=conn.execute(f"SELECT * FROM monteiro_payments WHERE {ws} ORDER BY payment_date DESC, id DESC",args).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def _check_payment_perm(x_token):
    sess=require_auth(x_token)
    import json as _j
    conn=get_db()
    row=conn.execute("SELECT value FROM settings WHERE key='monteiro_payment_perm'").fetchone()
    conn.close()
    allowed= _j.loads(row["value"]) if row else ["admin"]
    if not payment_role_allowed(sess, allowed):
        raise HTTPException(403,"Seu perfil nÃ£o tem permissÃ£o para lanÃ§ar pagamentos.")
    return sess

@app.post("/api/monteiro/payments")
def create_monteiro_payment(body: dict, x_token: str = Header(...)):
    _check_payment_perm(x_token)
    client=body.get("client","").strip()
    if not client: raise HTTPException(400,"Cliente obrigatÃ³rio.")
    pdate=body.get("payment_date","")
    amount=float(body.get("amount",0))
    if not pdate or amount<=0: raise HTTPException(400,"Data e valor obrigatÃ³rios.")
    m=body.get("month","") or pdate[5:7]
    y=body.get("year","") or pdate[:4]
    conn=get_db()
    try:
        conn.execute("INSERT INTO monteiro_payments (client, payment_date, amount, month, year, payment_type, notes) VALUES (?,?,?,?,?,?,?)",
                     [client,pdate,amount,m.zfill(2),str(y),body.get("payment_type","repasse"),body.get("notes","")])
        conn.commit()
        return {"ok":True}
    except Exception:
        try:
            conn.rollback()
        finally:
            raise
    finally:
        conn.close()

@app.delete("/api/monteiro/payments/{pid}")
def delete_monteiro_payment(pid: int, x_token: str = Header(...)):
    _check_payment_perm(x_token)
    conn=get_db()
    try:
        conn.execute("DELETE FROM monteiro_payments WHERE id=?",[pid])
        conn.commit()
        return {"ok":True}
    except Exception:
        try:
            conn.rollback()
        finally:
            raise
    finally:
        conn.close()

@app.put("/api/monteiro/payments/{pid}")
def update_monteiro_payment(pid:int,body:dict,x_token:str=Header(...)):
    _check_payment_perm(x_token)
    client=body.get("client","").strip()
    pdate=body.get("payment_date","")
    amount=float(body.get("amount",0))
    if not client or not pdate or amount<=0:
        raise HTTPException(400,"Cliente, data e valor obrigatÃ³rios.")
    m=body.get("month","") or pdate[5:7]
    y=body.get("year","") or pdate[:4]
    conn=get_db()
    try:
        conn.execute("""UPDATE monteiro_payments SET client=?, payment_date=?, amount=?,
            month=?, year=?, payment_type=?, notes=? WHERE id=?""",
            [client,pdate,amount,m.zfill(2),str(y),body.get("payment_type","repasse"),body.get("notes",""),pid])
        conn.commit()
        return {"ok":True}
    except Exception:
        try:
            conn.rollback()
        finally:
            raise
    finally:
        conn.close()

@app.get("/api/monteiro/payments/summary")
def monteiro_payments_summary(client:str="",period:Optional[str]=None,month:str="",year:str="",x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    pm,py=_pay_period_map(period,month,year)
    # Sales: use _pal_period_where que entende quinzenal, mensal, anual
    sw, sa = _pal_period_where(period, pm, py)
    if client.strip(): sw.append("client=?"); sa.append(client.strip())
    sws=" AND ".join(sw) if sw else "1"
    # Payments: filtro por month/year + client
    pw, pa = [], []
    if client.strip(): pw.append("client=?"); pa.append(client.strip())
    if pm: pw.append("month=?"); pa.append(pm)
    if py: pw.append("year=?"); pa.append(py)
    pws=" AND ".join(pw) if pw else "1"
    srow=conn.execute(f"""SELECT COALESCE(SUM(total),0) as total_vendido,
        COUNT(DISTINCT sale_group) as num_notas
        FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {sws}""",sa).fetchone()
    prow=conn.execute(f"""SELECT COALESCE(SUM(amount),0) as total_pago,
        COUNT(*) as num_pagamentos, MAX(payment_date) as ultimo_pagamento
        FROM monteiro_payments WHERE {pws}""",pa).fetchone()
    conn.close()
    tv=float(srow["total_vendido"]); tp=float(prow["total_pago"])
    return {
        "total_vendido": tv, "num_notas": srow["num_notas"],
        "total_pago": tp, "num_pagamentos": prow["num_pagamentos"],
        "ultimo_pagamento": prow["ultimo_pagamento"],
        "saldo": round(tv-tp,2)
    }

@app.get("/api/monteiro/payments/report")
def monteiro_payments_report(client:str="",period:Optional[str]=None,month:str="",year:str="",x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    pm,py=_pay_period_map(period,month,year)
    sw, sa = _pal_period_where(period, pm, py)
    if client.strip(): sw.append("client=?"); sa.append(client.strip())
    sws=" AND ".join(sw) if sw else "1"
    pw, pa = [], []
    if client.strip(): pw.append("client=?"); pa.append(client.strip())
    if pm: pw.append("month=?"); pa.append(pm)
    if py: pw.append("year=?"); pa.append(py)
    pws=" AND ".join(pw) if pw else "1"
    vendas=conn.execute(f"""SELECT sale_group as id, saledate, nf_number, SUM(quantity) as qty, SUM(total) as total, MAX(driver) as entregador
        FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {sws}
        GROUP BY sale_group ORDER BY saledate""",sa).fetchall()
    pags=conn.execute(f"""SELECT * FROM monteiro_payments WHERE {pws} ORDER BY payment_date""",pa).fetchall()
    conn.close()
    extrato=[]
    for v in vendas:
        extrato.append({"tipo":"nota","data":v["saledate"],"descricao":f"NF {v['nf_number'] or 'â€”'}","valor":round(float(v["total"]),2)})
    total_vendas=sum(float(v["total"]) for v in vendas)
    extrato.append({"tipo":"total_notas","data":"","descricao":"Total das notas","valor":round(total_vendas,2)})
    for p in pags:
        extrato.append({"tipo":"pagamento","data":p["payment_date"],"descricao":f"Pagamento: {p['payment_type']}","valor":-round(float(p["amount"]),2)})
    total_pago=sum(float(p["amount"]) for p in pags)
    saldo=round(total_vendas-total_pago,2)
    return {
        "client": client.strip() or "Todos",
        "period": f"{pm}/{py}" if pm or py else "Todo perÃ­odo",
        "vendas": [dict(v) for v in vendas],
        "pagamentos": [dict(p) for p in pags],
        "extrato": extrato,
        "total_vendas": round(total_vendas,2),
        "total_pago": round(total_pago,2),
        "saldo": saldo
    }

@app.get("/api/monteiro/payments/report/excel")
def monteiro_payments_excel(client:str="",period:Optional[str]=None,month:str="",year:str="",x_token:str=Header("")):
    require_auth(x_token)
    import io; from openpyxl import Workbook; from openpyxl.styles import Font,PatternFill,Border,Side
    pm,py=_pay_period_map(period,month,year)
    sw, sa = _pal_period_where(period, pm, py)
    if client.strip(): sw.append("client=?"); sa.append(client.strip())
    sws=" AND ".join(sw) if sw else "1"
    pw, pa = [], []
    if client.strip(): pw.append("client=?"); pa.append(client.strip())
    if pm: pw.append("month=?"); pa.append(pm)
    if py: pw.append("year=?"); pa.append(py)
    pws=" AND ".join(pw) if pw else "1"
    conn=get_db()
    vendas=conn.execute(f"""SELECT client, sale_group, saledate, nf_number, SUM(quantity) as qty, SUM(total) as total
        FROM paladar_sales WHERE client IS NOT NULL AND client!='' AND {sws}
        GROUP BY client, sale_group ORDER BY client, saledate""",sa).fetchall()
    pags=conn.execute(f"""SELECT * FROM monteiro_payments WHERE {pws} ORDER BY client, payment_date""",pa).fetchall()
    conn.close()
    wb=Workbook(); hdr=Font(bold=True,color="FFFFFF",size=11); hf=PatternFill("solid",fgColor="4F46E5"); bdr=Border(bottom=Side(style='thin',color='E5E7EB'))
    ws=wb.active; ws.title="Pagamentos"
    ws.append(["Cliente","Data","NF","Valor Nota","Data Pagamento","Valor Pago","Tipo","Obs"])
    for i in range(1,9): ws.cell(row=1,column=i).font=hdr; ws.cell(row=1,column=i).fill=hf
    from collections import defaultdict
    notas_map=defaultdict(list); pags_map=defaultdict(list)
    for v in vendas: notas_map[(v["client"],v["saledate"][5:7],v["saledate"][:4])].append(v)
    for p in pags: pags_map[(p["client"],p["month"],p["year"])].append(p)
    for (cli,mm,yy), nts in sorted(notas_map.items()):
        total_notas=sum(float(n["total"]) for n in nts)
        for n in nts: ws.append([cli,n["saledate"],n["nf_number"] or "",round(float(n["total"]),2),"","","",""])
        ws.append(["","","","",f"Total client: R$ {total_notas:.2f}","","",""])
        pags_cli=pags_map.get((cli,mm,yy),[])
        for p in pags_cli: ws.append(["","","","",p["payment_date"],round(float(p["amount"]),2),p["payment_type"],p.get("notes","")])
        total_pago=sum(float(p["amount"]) for p in pags_cli)
        saldo=round(total_notas-total_pago,2)
        ws.append(["","","","","Saldo:",round(total_pago,2),"",f"R$ {saldo:.2f}"])
        ws.append([])
    ws.column_dimensions['A'].width=25; ws.column_dimensions['B'].width=14; ws.column_dimensions['C'].width=16
    ws.column_dimensions['D'].width=14; ws.column_dimensions['E'].width=16; ws.column_dimensions['F'].width=14
    ws.column_dimensions['G'].width=14; ws.column_dimensions['H'].width=30
    for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=8):
        for c in row: c.border=bdr
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":"attachment;filename=monteiro_pagamentos.xlsx"})

# â”€â”€ Analytics de Clientes â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@app.get("/api/analytics/clients")
def client_analytics(days_inactive:int=30,period:str="all",sale_type:str="",
                     x_token:str=Header("")):
    require_auth(x_token); conn=get_db()
    from datetime import date, timedelta
    import unicodedata
    today=date.today()

    def ckey(name):
        """Chave normalizada: maiÃºsculo, sem acento, separadores -/â€“â€” viram espaÃ§o."""
        s=unicodedata.normalize('NFD',(name or '').strip().upper()).encode('ascii','ignore').decode()
        s=re.sub(r'[-/\u2013\u2014]',' ',s)
        s=re.sub(r'\s+',' ',s).strip()
        return s

    # Base filter
    where="sale_type!='AVARIA'"; args=[]
    if sale_type: where+=" AND sale_type=?"; args.append(sale_type)
    if period=="month":
        where+=" AND strftime('%Y-%m',sale_date)=?"; args.append(today.strftime('%Y-%m'))
    elif period=="week":
        where+=" AND sale_date>=?"; args.append((today-timedelta(days=7)).isoformat())

    # Buscar todas as vendas com cliente (agrupar em Python por chave normalizada)
    rows=conn.execute(f"""
        SELECT TRIM(client) AS client, sale_date, total
        FROM sales
        WHERE client IS NOT NULL AND TRIM(client)!='' AND {where}
    """, args).fetchall()

    # Agregar por chave normalizada
    groups={}  # key â†’ dict
    for r in rows:
        k=ckey(r["client"])
        if not k: continue
        g=groups.get(k)
        if not g:
            g=groups[k]={"names":{}, "total_val":0.0, "order_cnt":0,
                         "last_purchase":None, "first_purchase":None, "months":set()}
        # Conta a frequÃªncia de cada variante de nome para escolher o display
        g["names"][r["client"]]=g["names"].get(r["client"],0)+1
        g["total_val"]+=float(r["total"] or 0)
        g["order_cnt"]+=1
        sd=r["sale_date"]
        if sd:
            if g["last_purchase"] is None or sd>g["last_purchase"]: g["last_purchase"]=sd
            if g["first_purchase"] is None or sd<g["first_purchase"]: g["first_purchase"]=sd
            g["months"].add(sd[:7])

    def display_name(g):
        # Escolhe a variante mais frequente; em empate, a mais longa (mais completa)
        return sorted(g["names"].items(), key=lambda x:(-x[1], -len(x[0])))[0][0]

    ranking=[]
    for k,g in groups.items():
        cnt=g["order_cnt"]
        ranking.append({
            "client":display_name(g),
            "total_val":round(g["total_val"],2),
            "order_cnt":cnt,
            "avg_ticket":round(g["total_val"]/cnt,2) if cnt else 0,
            "last_purchase":g["last_purchase"],
            "first_purchase":g["first_purchase"],
            "active_months":len(g["months"]),
        })
    ranking.sort(key=lambda x:-x["total_val"])

    # Inativos: usa o MAX da data entre TODAS as variantes e tipos
    cutoff=(today-timedelta(days=days_inactive)).isoformat()
    inactive=[{"client":r["client"],"last_purchase":r["last_purchase"],
               "total_val":r["total_val"],"order_cnt":r["order_cnt"]}
              for r in ranking if r["last_purchase"] and r["last_purchase"]<cutoff]
    inactive.sort(key=lambda x:x["last_purchase"])

    # Queda nas compras: Ãºltimos 30d vs 30d anteriores (tambÃ©m por chave normalizada)
    p1_start=(today-timedelta(days=30)).isoformat()
    p2_start=(today-timedelta(days=60)).isoformat()
    p2_end  =(today-timedelta(days=31)).isoformat()
    def agg_period(d1,d2=None):
        q="SELECT TRIM(client) AS client,sale_date,total FROM sales WHERE client IS NOT NULL AND sale_type!='AVARIA' AND sale_date>=?"
        a=[d1]
        if d2: q+=" AND sale_date<=?"; a.append(d2)
        m={}
        for r in conn.execute(q,a).fetchall():
            k=ckey(r["client"])
            m[k]=m.get(k,0)+float(r["total"] or 0)
        return m
    recent_map=agg_period(p1_start)
    prev_map=agg_period(p2_start,p2_end)
    # nome de exibiÃ§Ã£o por chave
    keyname={k:display_name(g) for k,g in groups.items()}
    declining=[]
    for k,prev_val in prev_map.items():
        cur_val=recent_map.get(k,0)
        if prev_val>0 and cur_val<prev_val*0.7:
            declining.append({"client":keyname.get(k,k),"prev_val":prev_val,"cur_val":cur_val,
                              "drop_pct":round((1-cur_val/prev_val)*100,1)})
    declining.sort(key=lambda x:-x["drop_pct"])

    # Ticket baixo
    if ranking:
        avg_tickets=[r["avg_ticket"] for r in ranking if r["avg_ticket"]]
        median_ticket=sorted(avg_tickets)[len(avg_tickets)//2] if avg_tickets else 0
        low_ticket=[r for r in ranking
                    if r["avg_ticket"] and r["avg_ticket"]<median_ticket*0.6 and r["order_cnt"]>=3]
    else:
        low_ticket=[]

    conn.close()
    return {
        "ranking":ranking,
        "inactive":inactive,
        "declining":declining[:10],
        "low_ticket":low_ticket[:10],
        "summary":{
            "total_clients":len(ranking),
            "inactive_count":len(inactive),
            "declining_count":len(declining),
            "low_ticket_count":len(low_ticket),
            "top_client":ranking[0] if ranking else None,
            "bottom_client":ranking[-1] if ranking else None,
        }
    }


# â”€â”€ Config (persistÃªncia genÃ©rica) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get('/api/config/{key}')
def get_config(key: str, x_token: str = Header(...)):
    require_auth(x_token)
    conn = get_db()
    row = conn.execute('SELECT value FROM app_config WHERE key=?', [key]).fetchone()
    conn.close()
    return {'key': key, 'value': row['value'] if row else None}

@app.put('/api/config/{key}')
def set_config(key: str, body: dict, x_token: str = Header(...)):
    require_admin_or_editor(x_token)
    value = str(body.get('value', ''))
    conn = get_db()
    conn.execute('''INSERT INTO app_config (key, value) VALUES (?, ?)
                    ON CONFLICT(key) DO UPDATE SET value=excluded.value''', [key, value])
    conn.commit()
    conn.close()
    return {'ok': True}

# â”€â”€ Frontend â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _quote_totals(items, discount=0):
    try:
        return quote_totals_from_items(items, discount)
    except QuoteItemsLimitError:
        raise HTTPException(400,"O orÃ§amento permite no mÃ¡ximo 20 itens.")

@app.get("/api/orcamentos/company")
def quote_company(x_token:str=Header("")):
    require_auth(x_token)
    return {"default":"estrada","companies":list(_quote_companies().values())}

@app.get("/api/orcamentos/products")
def list_quote_products(active:Optional[int]=1, search:Optional[str]=None, x_token:str=Header("")):
    require_auth(x_token)
    conn=get_db(); sql="SELECT * FROM quote_products WHERE 1=1"; args=[]
    if active is not None:
        sql+=" AND active=?"; args.append(int(active))
    if search:
        sql+=" AND (name LIKE ? OR code LIKE ? OR description LIKE ?)"; args += [f"%{search}%"]*3
    sql+=" ORDER BY name COLLATE NOCASE"
    rows=conn.execute(sql,args).fetchall(); conn.close()
    return [dict(r) for r in rows]

@app.post("/api/orcamentos/products")
def create_quote_product(body:dict, x_token:str=Header("")):
    require_admin_or_editor(x_token)
    name=str(body.get("name") or "").strip()
    if not name: raise HTTPException(400,"Nome do produto obrigatorio.")
    code=str(body.get("code") or "").strip()
    unit=str(body.get("unit") or "UND").strip().upper()
    price=float(body.get("default_price") or body.get("price") or 0)
    desc=str(body.get("description") or "").strip()
    conn=get_db()
    try:
        cur=conn.execute("""INSERT INTO quote_products(name,code,unit,default_price,description,active,updated_at)
            VALUES(?,?,?,?,?,1,datetime('now'))""",(name,code,unit,price,desc))
        conn.commit()
        row=conn.execute("SELECT * FROM quote_products WHERE id=?",(cur.lastrowid,)).fetchone()
    except sqlite3.IntegrityError:
        conn.close(); raise HTTPException(400,"Produto ja cadastrado em Orcamentos.")
    conn.close(); return dict(row)

@app.put("/api/orcamentos/products/{pid}")
def update_quote_product(pid:int, body:dict, x_token:str=Header("")):
    require_admin_or_editor(x_token)
    name=str(body.get("name") or "").strip()
    if not name: raise HTTPException(400,"Nome do produto obrigatorio.")
    code=str(body.get("code") or "").strip()
    unit=str(body.get("unit") or "UND").strip().upper()
    price=float(body.get("default_price") or body.get("price") or 0)
    desc=str(body.get("description") or "").strip()
    active=int(body.get("active",1))
    conn=get_db()
    conn.execute("""UPDATE quote_products SET name=?,code=?,unit=?,default_price=?,description=?,active=?,updated_at=datetime('now')
                   WHERE id=?""",(name,code,unit,price,desc,active,pid))
    conn.commit()
    row=conn.execute("SELECT * FROM quote_products WHERE id=?",(pid,)).fetchone()
    conn.close()
    if not row: raise HTTPException(404,"Produto nao encontrado.")
    return dict(row)

@app.delete("/api/orcamentos/products/{pid}")
def delete_quote_product(pid:int, x_token:str=Header("")):
    require_admin_or_editor(x_token)
    conn=get_db()
    conn.execute("UPDATE quote_products SET active=0,updated_at=datetime('now') WHERE id=?",(pid,))
    conn.commit(); conn.close()
    return {"ok":True}

@app.get("/api/orcamentos")
def list_quotes(search:Optional[str]=None, x_token:str=Header("")):
    require_auth(x_token)
    conn=get_db(); sql="SELECT * FROM quotes WHERE 1=1"; args=[]
    if search:
        sql+=" AND (client_name LIKE ? OR client_cnpj LIKE ? OR CAST(quote_number AS TEXT) LIKE ?)"; args += [f"%{search}%"]*3
    sql+=" ORDER BY issue_date DESC, id DESC LIMIT 300"
    rows=conn.execute(sql,args).fetchall(); conn.close()
    return [dict(r) for r in rows]

@app.get("/api/orcamentos/{qid}")
def get_quote(qid:int, x_token:str=Header("")):
    require_auth(x_token)
    conn=get_db()
    q=conn.execute("SELECT * FROM quotes WHERE id=?",(qid,)).fetchone()
    if not q:
        conn.close(); raise HTTPException(404,"Orcamento nao encontrado.")
    items=conn.execute("SELECT * FROM quote_items WHERE quote_id=? ORDER BY item_order,id",(qid,)).fetchall()
    conn.close()
    qd=dict(q)
    return {"company":_quote_company(qd.get("company_key")),"quote":qd,"items":[dict(r) for r in items]}

@app.post("/api/orcamentos")
def create_quote(body:dict, x_token:str=Header("")):
    sess=require_admin_or_editor(x_token)
    items, subtotal, total = _quote_totals(body.get("items") or [], body.get("discount") or 0)
    if not items: raise HTTPException(400,"Adicione pelo menos um produto ao orcamento.")
    client=str(body.get("client_name") or "").strip()
    if not client: raise HTTPException(400,"Nome do cliente obrigatorio.")
    company_key=str(body.get("company_key") or "estrada").strip().lower()
    if company_key not in _quote_companies(): company_key="estrada"
    now=datetime.now()
    conn=get_db()
    try:
        next_no=(conn.execute("SELECT COALESCE(MAX(quote_number),0)+1 FROM quotes").fetchone()[0] or 1)
        cur=conn.execute("""INSERT INTO quotes(quote_number,company_key,client_name,attention,client_cnpj,client_ie,client_phone,client_email,
            client_address,client_district,client_city,client_state,client_zip,issue_date,issue_time,validity_days,delivery_deadline,
            payment_terms,observations,discount,subtotal,total,status,created_by,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))""",
            (next_no,company_key,client,body.get("attention"),body.get("client_cnpj"),body.get("client_ie"),body.get("client_phone"),
             body.get("client_email"),body.get("client_address"),body.get("client_district"),body.get("client_city"),
             body.get("client_state"),body.get("client_zip"),body.get("issue_date") or now.strftime("%Y-%m-%d"),
             body.get("issue_time") or now.strftime("%H:%M:%S"),int(body.get("validity_days") or 3),body.get("delivery_deadline"),
             body.get("payment_terms"),body.get("observations"),float(body.get("discount") or 0),subtotal,total,
             body.get("status") or "emitido",sess["username"]))
        qid=cur.lastrowid
        for it in items:
            conn.execute("""INSERT INTO quote_items(quote_id,product_id,item_order,code,description,quantity,unit,unit_price,discount,subtotal)
                VALUES(?,?,?,?,?,?,?,?,?,?)""",(qid,it.get("product_id"),it["item_order"],it.get("code"),it["description"],
                it["quantity"],it["unit"],it["unit_price"],it["discount"],it["subtotal"]))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return get_quote(qid,x_token)

@app.put("/api/orcamentos/{qid}")
def update_quote(qid:int, body:dict, x_token:str=Header("")):
    require_admin_or_editor(x_token)
    items, subtotal, total = _quote_totals(body.get("items") or [], body.get("discount") or 0)
    if not items: raise HTTPException(400,"Adicione pelo menos um produto ao orcamento.")
    client=str(body.get("client_name") or "").strip()
    if not client: raise HTTPException(400,"Nome do cliente obrigatorio.")
    company_key=str(body.get("company_key") or "estrada").strip().lower()
    if company_key not in _quote_companies(): company_key="estrada"
    conn=get_db()
    try:
        conn.execute("""UPDATE quotes SET company_key=?,client_name=?,attention=?,client_cnpj=?,client_ie=?,client_phone=?,client_email=?,
            client_address=?,client_district=?,client_city=?,client_state=?,client_zip=?,issue_date=?,issue_time=?,validity_days=?,
            delivery_deadline=?,payment_terms=?,observations=?,discount=?,subtotal=?,total=?,status=?,updated_at=datetime('now')
            WHERE id=?""",(company_key,client,body.get("attention"),body.get("client_cnpj"),body.get("client_ie"),body.get("client_phone"),
            body.get("client_email"),body.get("client_address"),body.get("client_district"),body.get("client_city"),
            body.get("client_state"),body.get("client_zip"),body.get("issue_date"),body.get("issue_time"),
            int(body.get("validity_days") or 3),body.get("delivery_deadline"),body.get("payment_terms"),body.get("observations"),
            float(body.get("discount") or 0),subtotal,total,body.get("status") or "emitido",qid))
        conn.execute("DELETE FROM quote_items WHERE quote_id=?",(qid,))
        for it in items:
            conn.execute("""INSERT INTO quote_items(quote_id,product_id,item_order,code,description,quantity,unit,unit_price,discount,subtotal)
                VALUES(?,?,?,?,?,?,?,?,?,?)""",(qid,it.get("product_id"),it["item_order"],it.get("code"),it["description"],
                it["quantity"],it["unit"],it["unit_price"],it["discount"],it["subtotal"]))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return get_quote(qid,x_token)

@app.delete("/api/orcamentos/{qid}")
def delete_quote(qid:int, x_token:str=Header("")):
    require_admin_or_editor(x_token)
    conn=get_db()
    try:
        conn.execute("DELETE FROM quote_items WHERE quote_id=?",(qid,))
        conn.execute("DELETE FROM quotes WHERE id=?",(qid,))
        conn.commit()
        return {"ok":True}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def find_index():
    for p in [STATIC_DIR/"index.html",STATIC_DIR/"client"/"index.html"]:
        if p.exists(): return p
    return None

assets_dir=next((d for d in [STATIC_DIR/"assets",STATIC_DIR/"client"/"assets"] if d.exists()),None)
if assets_dir: app.mount("/assets",StaticFiles(directory=str(assets_dir)),name="assets")

# â”€â”€ WhatsApp Notifications â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def wa_send(phone: str, message: str, cfg: dict) -> dict:
    try:
        import httpx as _hlib
        def _post(url, data=None, json=None, headers=None, timeout=10):
            r = _hlib.post(url, data=data, json=json, headers=headers, timeout=timeout)
            return r
    except ImportError:
        try:
            import requests as _rlib
            class _Resp:
                def __init__(self, r):
                    self.status_code = r.status_code
                    self.text        = r.text
            def _post(url, data=None, json=None, headers=None, timeout=10):
                return _Resp(_rlib.post(url, data=data, json=json, headers=headers, timeout=timeout))
        except ImportError:
            return {"ok": False, "response": "Nenhuma biblioteca HTTP. Execute: pip install httpx"}
    provider = cfg.get("provider", "ultramsg")
    api_url  = (cfg.get("api_url") or "").strip().rstrip("/")
    token    = cfg.get("api_token", "")
    instance = cfg.get("instance_id", "")
    if not api_url or not token:
        return {"ok": False, "response": "API n\u00e3o configurada."}
    try:
        if provider == "ultramsg":
            url  = f"{api_url}/instance{instance}/messages/chat"
            data = {"token": token, "to": phone, "body": message}
            r = _post(url, data=data, timeout=10)
        elif provider == "zapi":
            url  = f"{api_url}/instances/{instance}/token/{token}/send-text"
            r = _post(url, json={"phone": phone, "message": message}, timeout=10)
        elif provider == "evolution":
            url  = f"{api_url}/message/sendText/{instance}"
            r = _post(url, json={"number": phone, "textMessage": {"text": message}},
                      headers={"apikey": token, "Content-Type": "application/json"}, timeout=10)
        elif provider == "baileys":
            # Baileys local (server.js rodando em 127.0.0.1:3001)
            # api_url = http://127.0.0.1:3001  |  token = API_KEY do .env
            url = f"{api_url}/send"
            r = _post(url, json={"phone": phone, "message": message},
                      headers={"x-api-key": token, "Content-Type": "application/json"}, timeout=15)
        else:
            return {"ok": False, "response": "Provider desconhecido."}
        # HTTP 2xx Ã© necessÃ¡rio, mas nÃ£o suficiente: UltraMsg/Z-API/Evolution
        # devolvem {"error": "..."} ou {"sent": "false"} dentro do body com
        # status 200 quando a URL/token/instance estÃ£o errados. Sem essa
        # checagem, marcÃ¡vamos como 'sent' uma mensagem que nunca saiu.
        ok = r.status_code in (200, 201)
        text = r.text or ""
        if ok and text.strip().startswith(("{", "[")):
            try:
                j = json.loads(text)
                if isinstance(j, dict):
                    if j.get("error") or j.get("err"):
                        ok = False
                    elif str(j.get("sent", "")).lower() == "false":
                        ok = False
                    elif j.get("status") and str(j["status"]).lower() in ("error","fail","failed"):
                        ok = False
            except Exception:
                pass
        return {"ok": ok, "response": text[:300]}
    except Exception as e:
        return {"ok": False, "response": str(e)}

def _motivation_now():
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("America/Manaus"))
    except Exception:
        return datetime.utcnow()-timedelta(hours=4)

def _motivation_templates(conn):
    rows=conn.execute("""SELECT id,name,content,category FROM whatsapp_templates
        WHERE TRIM(COALESCE(content,''))!='' AND
        (LOWER(COALESCE(category,'')) LIKE '%motiv%' OR LOWER(COALESCE(name,'')) LIKE '%motiv%')
        ORDER BY created_at,id""").fetchall()
    if not rows:
        stored=conn.execute("""SELECT id,name,content,category FROM whatsapp_templates
            WHERE TRIM(COALESCE(content,''))!='' ORDER BY created_at,id""").fetchall()
        if len(stored)==60:
            rows=stored
    return [dict(r) for r in rows]

def send_daily_motivation(force:bool=False):
    now=_motivation_now();today=now.strftime("%Y-%m-%d")
    conn=get_db()
    try:
        cfg={r["key"]:r["value"] for r in conn.execute("SELECT key,value FROM whatsapp_config").fetchall()}
        if not force and cfg.get("motivation_enabled","1")!="1": return {"ok":False,"reason":"disabled"}
        send_time=cfg.get("motivation_time","07:00") or "07:00"
        if not force:
            try:
                hh,mm=[int(part) for part in send_time.split(":",1)]
                scheduled=now.replace(hour=hh,minute=mm,second=0,microsecond=0)
            except Exception:
                scheduled=now.replace(hour=7,minute=0,second=0,microsecond=0)
            if now<scheduled: return {"ok":False,"reason":"before_time"}
            if now>=scheduled+timedelta(hours=2): return {"ok":False,"reason":"after_time"}
        if not force and cfg.get("motivation_last_success","")==today: return {"ok":True,"reason":"already_sent"}
        last_attempt=cfg.get("motivation_last_attempt","")
        if not force and last_attempt.startswith(today):
            try:
                attempted=datetime.fromisoformat(last_attempt)
                if attempted.tzinfo is None: attempted=attempted.replace(tzinfo=now.tzinfo)
                if (now-attempted).total_seconds()<3600: return {"ok":False,"reason":"retry_wait"}
            except Exception: pass
        templates=_motivation_templates(conn)
        if not templates:
            print("[motivacao] Nenhum template motivacional encontrado no banco.")
            return {"ok":False,"reason":"no_templates","count":0}
        contacts=[dict(r) for r in conn.execute("SELECT * FROM whatsapp_contacts WHERE active=1 ORDER BY name").fetchall()]
        if not contacts: return {"ok":False,"reason":"no_contacts","count":len(templates)}
        item=templates[(now.toordinal()-1)%len(templates)]
        content=(item.get("content") or "").strip()
        conn.execute("INSERT OR REPLACE INTO whatsapp_config(key,value) VALUES('motivation_last_attempt',?)",[now.isoformat()])
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    sent=0;failed=0;results=[]
    for contact in contacts:
        contact_name=(contact.get("name") or "").strip()
        personalized=content.replace("{cliente}",contact_name).replace("{CLIENTE}",contact_name)
        message=f"*Mensagem do dia*\n\n{personalized}\n\n_Um excelente dia para todos nos._"
        result=wa_send(contact["phone"],message,cfg)
        results.append((contact,result,message))
        if result["ok"]: sent+=1
        else: failed+=1
    conn=get_db()
    try:
        for contact,result,message in results:
            conn.execute("INSERT INTO whatsapp_log(id,phone,contact,event_type,message,status,response) VALUES(?,?,?,?,?,?,?)",
                         [str(uuid.uuid4()),contact["phone"],contact["name"],"motivacao",message,
                          "sent" if result["ok"] else "error",_wa_log_response(result,cfg)])
        if sent>0:
            conn.execute("INSERT OR REPLACE INTO whatsapp_config(key,value) VALUES('motivation_last_success',?)",[today])
        conn.commit()
        print(f"[motivacao] {today}: template {item.get('name')} | enviados={sent} falhas={failed}")
        return {"ok":sent>0,"sent":sent,"failed":failed,"templates":len(templates),"message":content}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def motivation_scheduler():
    import time
    time.sleep(8)
    while True:
        try: send_daily_motivation()
        except Exception as e: print(f"[motivacao] Erro no agendador: {type(e).__name__}: {e}")
        time.sleep(300)

@app.get("/api/whatsapp/contacts")
def list_wa_contacts(x_token: str = Header(...)):
    require_auth(x_token)
    conn = get_db()
    rows = conn.execute("SELECT * FROM whatsapp_contacts ORDER BY name").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/whatsapp/contacts")
def create_wa_contact(body: dict, x_token: str = Header(...)):
    sess  = require_admin(x_token)
    name  = (body.get("name") or "").strip()
    phone = re.sub(r'\D', '', body.get("phone") or "")
    if not name or not phone:
        raise HTTPException(400, "Nome e telefone obrigat\u00f3rios.")
    if not (10 <= len(phone) <= 15):
        raise HTTPException(400, "Telefone inv\u00e1lido. Use DDD + n\u00famero (ex: 95999999999).")
    if not phone.startswith("55"):
        phone = "55" + phone
    conn = get_db()
    conn.execute("INSERT INTO whatsapp_contacts (id, name, phone, created_by) VALUES (?,?,?,?)",
                 [str(uuid.uuid4()), name, phone, sess["username"]])
    conn.commit()
    conn.close()
    return {"ok": True}

@app.put("/api/whatsapp/contacts/{cid}")
def update_wa_contact(cid: str, body: dict, x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    try:
        for f in ["name", "phone", "active"]:
            if f not in body: continue
            if f == "phone":
                val = re.sub(r'\D', '', str(body[f]))
                if val and not val.startswith("55"): val = "55" + val
            else: val = body[f]
            conn.execute(f"UPDATE whatsapp_contacts SET {f}=? WHERE id=?", [val, cid])
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        finally:
            raise
    finally:
        conn.close()
    return {"ok": True}

@app.delete("/api/whatsapp/contacts/{cid}")
def delete_wa_contact(cid: str, x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    conn.execute("DELETE FROM whatsapp_contacts WHERE id=?", [cid])
    conn.commit()
    conn.close()
    return {"ok": True}

@app.get("/api/whatsapp/config")
def get_wa_config(x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    rows = conn.execute("SELECT key, value FROM whatsapp_config").fetchall()
    conn.close()
    return {r["key"]: r["value"] for r in rows}

@app.put("/api/whatsapp/config")
def save_wa_config(body: dict, x_token: str = Header(...)):
    require_admin(x_token)
    allowed = {"provider", "api_url", "api_token", "instance_id",
               "notify_boleto", "notify_avaria", "notify_inativo",
               "avaria_min", "inativo_dias", "auto_period"}
    conn = get_db()
    try:
        for k, v in body.items():
            if k in allowed:
                conn.execute("INSERT OR REPLACE INTO whatsapp_config (key, value) VALUES (?,?)", [k, str(v)])
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        conn.close()
    return {"ok": True}

@app.post("/api/whatsapp/check-triggers")
def check_wa_triggers(body: dict, x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    try:
        cfg_rows = conn.execute("SELECT key, value FROM whatsapp_config").fetchall()
        cfg = {r["key"]: r["value"] for r in cfg_rows}
        send = body.get("send", False)
        contact_ids = body.get("contact_ids")
        selected_contact_ids = None
        if isinstance(contact_ids, list):
            selected_contact_ids = [str(cid) for cid in contact_ids if str(cid or "").strip()]
        results = {}
        messages = []

        if cfg.get("notify_boleto", "1") == "1":
            auto_period_b = int(cfg.get("auto_period", "7"))
            # COALESCE protege contra total_val NULL em registros antigos/migrados
            overdue = conn.execute(
                """SELECT client, COALESCE(total_val,0) AS total_val, due_date, nf_number
                   FROM boletos
                   WHERE status='pendente'
                     AND due_date IS NOT NULL
                     AND due_date < date('now')
                     AND due_date >= date('now', '-' || ? || ' days')
                   ORDER BY due_date""",
                [auto_period_b]
            ).fetchall()
            if overdue:
                lines = [f"\u2022 {(b['client'] or '?')}: R$ {(b['total_val'] or 0):.2f} (venc {b['due_date']})" for b in overdue]
                total = sum((b['total_val'] or 0) for b in overdue)
                limit = 25
                if len(lines) > limit:
                    lines = lines[:limit] + [f"...e mais {len(overdue) - limit} boleto(s)"]
                msg = f"\U0001f4c4 *Boletos Vencidos*\n\n" + "\n".join(lines) + f"\n\nTotal: R$ {total:.2f}"
                messages.append({"type": "boleto", "message": msg, "count": len(overdue)})
                results["boleto"] = {"found": len(overdue), "total_value": round(total, 2)}

        if cfg.get("notify_avaria", "1") == "1":
            avaria_rate_min = float(cfg.get("avaria_min", "25"))
            from datetime import datetime as _dt_av
            _now_av = _dt_av.now()
            _mes_av = str(_now_av.month).zfill(2)
            _ano_av = str(_now_av.year)
            # Busca o mesmo perÃ­odo do ranking da tela (mÃªs corrente)
            # para garantir que os percentuais do WhatsApp batam com os exibidos.
            base_map = {}
            for r in conn.execute(
                """SELECT client, SUM(ABS(COALESCE(total,0))) AS total_vendas
                   FROM sales
                   WHERE sale_type != 'AVARIA'
                     AND strftime('%m',sale_date)=?
                     AND strftime('%Y',sale_date)=?
                     AND client IS NOT NULL AND TRIM(client) != ''
                   GROUP BY client""",
                [_mes_av, _ano_av]
            ).fetchall():
                base_map[r["client"]] = r["total_vendas"] or 0
            avarias_rows = conn.execute(
                """SELECT client, SUM(ABS(COALESCE(total,0))) AS total_avaria, COUNT(*) AS qtd
                   FROM sales
                   WHERE sale_type='AVARIA'
                     AND strftime('%m',sale_date)=?
                     AND strftime('%Y',sale_date)=?
                     AND client IS NOT NULL AND TRIM(client) != ''
                   GROUP BY client""",
                [_mes_av, _ano_av]
            ).fetchall()
            clientes = []
            for r in avarias_rows:
                cli = r["client"]
                val = r["total_avaria"] or 0
                base = base_map.get(cli, 0)
                taxa = (val / base) * 100 if base > 0 else 0
                clientes.append({"client": cli, "valor": val, "qtd": r["qtd"] or 0, "taxa": taxa})
            clientes.sort(key=lambda c: c["taxa"], reverse=True)
            clientes_filtrados = [c for c in clientes if c["taxa"] > avaria_rate_min]
            top5 = clientes_filtrados[:5]
            if top5:
                total_geral = sum(c["valor"] for c in clientes)
                lines = []
                for c in top5:
                    risco = "\U0001f534 ALT\u00cdSSIMO" if c["taxa"] > 40 else "\U0001f7e0 Alto"
                    lines.append(f"\u2022 {c['client']}: {c['taxa']:.1f}% ({risco}) - R$ {c['valor']:.2f} - {c['qtd']} ocorr\u00eancia(s)")
                taxa_media = total_geral / sum(base_map.values()) * 100 if sum(base_map.values()) > 0 else 0
                _nome_mes = ["Janeiro","Fevereiro","Mar\u00e7o","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
                msg = (
                    f"\u26a0\ufe0f *Avarias Altas* (>{avaria_rate_min:.0f}%)\n\n"
                    f"Per\u00edodo: {_nome_mes[_now_av.month-1]} {_ano_av}\n"
                    f"Total de avarias: R$ {total_geral:.2f}\n"
                    f"Taxa m\u00e9dia: {taxa_media:.1f}%\n\n"
                    f"*Top {len(top5)} clientes com maior taxa:*\n" + "\n".join(lines)
                )
                messages.append({"type": "avaria", "message": msg, "count": len(clientes_filtrados)})
                results["avaria"] = {
                    "found": len(clientes_filtrados),
                    "total_value": round(total_geral, 2),
                    "rate_threshold": avaria_rate_min,
                    "top5": [{"client": c["client"], "taxa": round(c["taxa"], 1),
                              "value": round(c["valor"], 2), "qtd": c["qtd"]} for c in top5]
                }

        if cfg.get("notify_inativo", "1") == "1":
            inativo_dias = int(cfg.get("inativo_dias", "30"))
            inativos = conn.execute(
                """SELECT s.client, MAX(s.sale_date) AS last_sale
                   FROM sales s
                   WHERE s.client IS NOT NULL AND TRIM(s.client) != ''
                   GROUP BY s.client
                   HAVING MAX(s.sale_date) < date('now', '-' || ? || ' days')
                   ORDER BY last_sale""",
                [inativo_dias]
            ).fetchall()
            if inativos:
                lines = [f"\u2022 {c['client']} (\u00faltima venda: {c['last_sale']})" for c in inativos[:20]]
                remaining = len(inativos) - 20
                if remaining > 0:
                    lines.append(f"...e mais {remaining} cliente(s)")
                msg = f"\U0001f634 *Clientes Inativos* (+{inativo_dias} dias)\n\n" + "\n".join(lines)
                messages.append({"type": "inativo", "message": msg, "count": len(inativos)})
                results["inativo"] = {"found": len(inativos)}

        sent_results = []
        if send and messages:
            if selected_contact_ids is not None:
                if selected_contact_ids:
                    ph = ",".join("?" for _ in selected_contact_ids)
                    active_contacts = [dict(r) for r in conn.execute(
                        f"SELECT * FROM whatsapp_contacts WHERE active=1 AND id IN ({ph})",
                        selected_contact_ids
                    ).fetchall()]
                else:
                    active_contacts = []
            else:
                active_contacts = [dict(r) for r in conn.execute("SELECT * FROM whatsapp_contacts WHERE active=1").fetchall()]
            conn.close()
            conn = None
            log_entries = []
            for m in messages:
                for c in active_contacts:
                    full_msg = m["message"] + "\n\n_Enviado por Menina dos Raios_"
                    res = wa_send(c["phone"], full_msg, cfg)
                    log_entries.append({
                        "phone": c["phone"], "contact": c["name"], "trigger": m["type"],
                        "message": full_msg, "status": "sent" if res["ok"] else "error",
                        "response": _wa_log_response(res,cfg)
                    })
                    sent_results.append({
                        "contact": c["name"], "phone": c["phone"],
                        "trigger": m["type"], "ok": res["ok"], "response": res.get("response",""),
                        "hint": "" if res["ok"] else _wa_failure_hint(res.get("response",""))
                    })
            conn = get_db()
            try:
                for item in log_entries:
                    conn.execute(
                        "INSERT INTO whatsapp_log (id,phone,contact,event_type,message,status,response) VALUES (?,?,?,?,?,?,?)",
                        [str(uuid.uuid4()), item["phone"], item["contact"], item["trigger"], item["message"],
                         item["status"], item["response"]]
                    )
                conn.commit()
            except Exception:
                conn.rollback()
                raise
            finally:
                conn.close()
                conn = None

        return {
            "triggers_fired": len(messages),
            "details": results,
            "messages": [{"type": m["type"], "preview": m["message"], "count": m["count"]} for m in messages],
            "sent": sent_results if send else None,
            "dry_run": not send
        }
    except HTTPException:
        raise
    except Exception as e:
        # Em vez de devolver 500 sem detalhes, retornamos o erro real para
        # facilitar diagnÃ³stico no frontend (toast mostra a mensagem).
        import traceback
        tb = traceback.format_exc()
        print(f"[check_wa_triggers] ERRO: {type(e).__name__}: {e}\n{tb}")
        raise HTTPException(500, f"{type(e).__name__}: {e}")
    finally:
        if conn is not None:
            conn.close()

@app.post("/api/whatsapp/send")
def wa_send_endpoint(body: dict, x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    try:
        cfg  = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM whatsapp_config").fetchall()}
        ids = body.get("contact_ids") or []
        if ids:
            ph  = ",".join(["?"] * len(ids))
            contacts = [dict(r) for r in conn.execute(f"SELECT * FROM whatsapp_contacts WHERE id IN ({ph}) AND active=1", ids).fetchall()]
        else:
            contacts = [dict(r) for r in conn.execute("SELECT * FROM whatsapp_contacts WHERE active=1").fetchall()]
    finally:
        conn.close()
    message    = (body.get("message") or "").strip()
    event_type = body.get("event_type", "manual")
    results    = []
    for c in contacts:
        res = wa_send(c["phone"], message, cfg)
        results.append({"contact": c["name"], "phone": c["phone"], "ok": res["ok"], "response": res["response"],
                        "hint": "" if res["ok"] else _wa_failure_hint(res.get("response",""))})
    conn = get_db()
    try:
        for c,res in zip(contacts,results):
            raw_res={"ok":res["ok"],"response":res["response"]}
            if res.get("hint"): raw_res["hint"]=res["hint"]
            conn.execute("INSERT INTO whatsapp_log (id,phone,contact,event_type,message,status,response) VALUES (?,?,?,?,?,?,?)",
                         [str(uuid.uuid4()), c["phone"], c["name"], event_type, message,
                          "sent" if res["ok"] else "error", _wa_log_response(raw_res,cfg)])
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return {"results": results, "total": len(results)}

@app.post("/api/whatsapp/test/{cid}")
def wa_test(cid: str, x_token: str = Header(...)):
    require_admin(x_token)
    conn    = get_db()
    contact = conn.execute("SELECT * FROM whatsapp_contacts WHERE id=?", [cid]).fetchone()
    cfg = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM whatsapp_config").fetchall()}
    conn.close()
    if not contact: raise HTTPException(404, "Contato n\u00e3o encontrado.")
    msg = ("\u2705 *Menina dos Raios \u2013 Teste de Notifica\u00e7\u00e3o*\n\n"
           "Este \u00e9 um teste do sistema de alertas.\n"
           f"Contato: {contact['name']}\n"
           f"Hora: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    res  = wa_send(contact["phone"], msg, cfg)
    if not res.get("ok"):
        res["hint"]=_wa_failure_hint(res.get("response",""))
    conn = get_db()
    try:
        conn.execute("INSERT INTO whatsapp_log (id,phone,contact,event_type,message,status,response) VALUES (?,?,?,?,?,?,?)",
                     [str(uuid.uuid4()), contact["phone"], contact["name"], "test", msg,
                      "sent" if res["ok"] else "error", _wa_log_response(res,cfg)])
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return res

@app.get("/api/whatsapp/baileys-status")
def wa_baileys_status(x_token: str = Header(...)):
    """Proxy para o status/QR do serviÃ§o Baileys local (http://127.0.0.1:3001)."""
    require_admin(x_token)
    conn = get_db()
    try:
        cfg  = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM whatsapp_config").fetchall()}
    finally:
        conn.close()
    api_url = (cfg.get("api_url") or "").strip().rstrip("/")
    token   = cfg.get("api_token", "")
    if not api_url:
        api_url = "http://127.0.0.1:3001"
    # Falha de conexÃ£o retorna status "offline" de forma graciosa (200), evitando
    # que o frontend mostre "Internal Server Error" e fique piscando no polling.
    try:
        try:
            import httpx as _hx
            r = _hx.get(f"{api_url}/qr", headers={"x-api-key": token}, timeout=5)
        except ImportError:
            import requests as _rq
            r = _rq.get(f"{api_url}/qr", headers={"x-api-key": token}, timeout=5)
        if r.status_code >= 500:
            return {"connected": False, "qr": None,
                    "message": f"ServiÃ§o Baileys respondeu erro {r.status_code}. Reinicie o serviÃ§o no servidor."}
        return r.json()
    except Exception as e:
        return {"connected": False, "qr": None,
                "message": f"ServiÃ§o Baileys offline ou inacessÃ­vel â€” verifique se ele estÃ¡ rodando no servidor ({type(e).__name__})."}

@app.post("/api/whatsapp/baileys-connection/{action}")
def wa_baileys_connection(action: str, x_token: str = Header(...)):
    require_admin(x_token)
    if action not in ("reconnect","disconnect"):
        raise HTTPException(400,"Acao de conexao invalida.")
    conn=get_db()
    try:
        cfg={r["key"]:r["value"] for r in conn.execute("SELECT key,value FROM whatsapp_config").fetchall()}
    finally:
        conn.close()
    if cfg.get("provider","ultramsg")!="baileys":
        raise HTTPException(400,"Os controles de QR Code estao disponiveis somente para o provedor Baileys.")
    api_url=(cfg.get("api_url") or "http://127.0.0.1:3001").strip().rstrip("/")
    token=cfg.get("api_token","")
    try:
        import httpx as _hx
        response=_hx.post(f"{api_url}/{action}",headers={"x-api-key":token},timeout=12)
        payload=response.json() if response.text else {}
    except ImportError:
        try:
            import requests as _rq
            response=_rq.post(f"{api_url}/{action}",headers={"x-api-key":token},timeout=12)
            payload=response.json() if response.text else {}
        except ImportError:
            raise HTTPException(503,"Nenhuma biblioteca HTTP disponivel no servidor.")
    except Exception as e:
        raise HTTPException(503,f"Servico Baileys offline ou inacessivel: {e}")
    if response.status_code not in (200,201):
        raise HTTPException(response.status_code,payload.get("error") or payload.get("message") or response.text[:300])
    return payload

@app.get("/api/whatsapp/log")
def get_wa_log(page: int = 1, limit: int = 10, x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    total = conn.execute("SELECT COUNT(*) AS c FROM whatsapp_log").fetchone()["c"]
    offset = (page - 1) * limit
    rows = conn.execute("SELECT * FROM whatsapp_log ORDER BY sent_at DESC LIMIT ? OFFSET ?", [limit, offset]).fetchall()
    conn.close()
    return {"logs": [dict(r) for r in rows], "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

@app.delete("/api/whatsapp/log/{log_id}")
def delete_wa_log(log_id: str, x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    conn.execute("DELETE FROM whatsapp_log WHERE id=?", [log_id])
    conn.commit()
    conn.close()
    return {"ok": True}

# â”€â”€ WhatsApp Auto Rules â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_whatsapp_rules_fields = ["name","rule_type","enabled","event_key","keyword","message","recipients","priority","schedule_from","schedule_to","cooldown_min","repeatable","max_per_day"]

@app.get("/api/whatsapp/auto-rules")
def list_wa_auto_rules(x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    rows = conn.execute("SELECT * FROM whatsapp_auto_rules ORDER BY priority ASC, name").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/whatsapp/auto-rules")
def create_wa_auto_rule(body: dict, x_token: str = Header(...)):
    require_admin(x_token)
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "Nome da regra obrigatÃ³rio.")
    rid = str(uuid.uuid4())
    vals = {"id": rid, **{k: body.get(k) for k in _whatsapp_rules_fields}}
    conn = get_db()
    conn.execute("""INSERT INTO whatsapp_auto_rules (id, name, rule_type, enabled, event_key, keyword, message, recipients, priority, schedule_from, schedule_to, cooldown_min, repeatable, max_per_day) VALUES (:id, :name, :rule_type, :enabled, :event_key, :keyword, :message, :recipients, :priority, :schedule_from, :schedule_to, :cooldown_min, :repeatable, :max_per_day)""", vals)
    conn.commit()
    conn.close()
    return {"ok": True, "id": rid}

@app.put("/api/whatsapp/auto-rules/{rid}")
def update_wa_auto_rule(rid: str, body: dict, x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    try:
        for k in _whatsapp_rules_fields:
            if k in body:
                conn.execute(f"UPDATE whatsapp_auto_rules SET {k}=?, updated_at=datetime('now') WHERE id=?", [body[k], rid])
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        finally:
            raise
    finally:
        conn.close()
    return {"ok": True}

@app.delete("/api/whatsapp/auto-rules/{rid}")
def delete_wa_auto_rule(rid: str, x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    conn.execute("DELETE FROM whatsapp_auto_rules WHERE id=?", [rid])
    conn.commit()
    conn.close()
    return {"ok": True}

# â”€â”€ WhatsApp Templates â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.get("/api/whatsapp/templates")
def list_wa_templates(x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    rows = conn.execute("SELECT * FROM whatsapp_templates ORDER BY category, name").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.get("/api/whatsapp/motivation-status")
def motivation_status(x_token: str = Header(...)):
    require_admin(x_token);conn=get_db()
    cfg={r["key"]:r["value"] for r in conn.execute("SELECT key,value FROM whatsapp_config").fetchall()}
    count=len(_motivation_templates(conn));contacts=conn.execute("SELECT COUNT(*) c FROM whatsapp_contacts WHERE active=1").fetchone()["c"]
    conn.close()
    return {"enabled":cfg.get("motivation_enabled","1")=="1","time":cfg.get("motivation_time","07:00"),
            "last_success":cfg.get("motivation_last_success",""),"last_attempt":cfg.get("motivation_last_attempt",""),
            "templates":count,"active_contacts":contacts}

@app.post("/api/whatsapp/motivation-send-now")
def motivation_send_now(x_token: str = Header(...)):
    require_admin(x_token);return send_daily_motivation(force=True)

@app.post("/api/whatsapp/templates")
def create_wa_template(body: dict, x_token: str = Header(...)):
    require_admin(x_token)
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "Nome do template obrigatÃ³rio.")
    tid = str(uuid.uuid4())
    conn = get_db()
    conn.execute("INSERT INTO whatsapp_templates (id, name, category, content) VALUES (?,?,?,?)",
                 [tid, name, body.get("category",""), body.get("content","")])
    conn.commit()
    conn.close()
    return {"ok": True, "id": tid}

@app.put("/api/whatsapp/templates/{tid}")
def update_wa_template(tid: str, body: dict, x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    try:
        for f in ["name","category","content"]:
            if f in body:
                conn.execute(f"UPDATE whatsapp_templates SET {f}=?, updated_at=datetime('now') WHERE id=?", [body[f], tid])
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        finally:
            raise
    finally:
        conn.close()
    return {"ok": True}

@app.delete("/api/whatsapp/templates/{tid}")
def delete_wa_template(tid: str, x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    conn.execute("DELETE FROM whatsapp_templates WHERE id=?", [tid])
    conn.commit()
    conn.close()
    return {"ok": True}

# â”€â”€ WhatsApp Bot Settings â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_BOT_SETTINGS_KEYS = ["bot_active","auto_reply_enabled","auto_reply_from","auto_reply_to",
                      "min_interval_secs","block_groups","block_duplicate","max_per_day_per_client","test_mode"]

@app.get("/api/whatsapp/bot-settings")
def get_wa_bot_settings(x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    rows = conn.execute("SELECT key, value FROM whatsapp_config WHERE key IN ({})".format(
        ",".join("?" for _ in _BOT_SETTINGS_KEYS)), _BOT_SETTINGS_KEYS).fetchall()
    conn.close()
    return {r["key"]: r["value"] for r in rows}

@app.put("/api/whatsapp/bot-settings")
def save_wa_bot_settings(body: dict, x_token: str = Header(...)):
    require_admin(x_token)
    conn = get_db()
    try:
        for k in _BOT_SETTINGS_KEYS:
            if k in body:
                conn.execute("INSERT OR REPLACE INTO whatsapp_config (key, value) VALUES (?,?)", [k, str(body[k])])
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        finally:
            raise
    finally:
        conn.close()
    return {"ok": True}

# â”€â”€ WhatsApp Test Send â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.post("/api/whatsapp/test-message")
def wa_test_message(body: dict, x_token: str = Header(...)):
    require_admin(x_token)
    msg  = (body.get("message") or "").strip()
    phone = re.sub(r'\D', '', body.get("phone") or "")
    if not msg:
        raise HTTPException(400, "Mensagem vazia.")
    if not phone:
        raise HTTPException(400, "Telefone obrigatÃ³rio para teste.")
    if not phone.startswith("55"):
        phone = "55" + phone
    conn = get_db()
    try:
        cfg = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM whatsapp_config").fetchall()}
    finally:
        conn.close()
    res = wa_send(phone, msg, cfg)
    return {"ok": res.get("ok"), "response": res.get("response")}

# â”€â”€ Notas enviadas pelo aplicativo Android (banco independente) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def get_app_notes_db():
    conn=sqlite3.connect(APP_NOTES_DB_PATH,timeout=20)
    conn.row_factory=sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        conn.executescript("""
          CREATE TABLE IF NOT EXISTS app_notes(
            id TEXT PRIMARY KEY, external_id TEXT NOT NULL UNIQUE, client TEXT NOT NULL DEFAULT '',
            note_date TEXT NOT NULL DEFAULT '', total REAL NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL, updated_at TEXT NOT NULL, source TEXT NOT NULL DEFAULT 'android',
            status TEXT NOT NULL DEFAULT 'pending', completed_at TEXT
          );
          CREATE TABLE IF NOT EXISTS app_note_items(
            id TEXT PRIMARY KEY, note_id TEXT NOT NULL, product TEXT NOT NULL DEFAULT '',
            quantity REAL NOT NULL DEFAULT 0, quantity_provided INTEGER NOT NULL DEFAULT 1,
            weight REAL NOT NULL DEFAULT 0, unit TEXT NOT NULL DEFAULT '',
            unit_price REAL NOT NULL DEFAULT 0, price_provided INTEGER NOT NULL DEFAULT 1,
            position INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY(note_id) REFERENCES app_notes(id) ON DELETE CASCADE
          );
          CREATE TABLE IF NOT EXISTS app_note_submissions(
            external_id TEXT PRIMARY KEY, note_id TEXT NOT NULL, received_at TEXT NOT NULL,
            FOREIGN KEY(note_id) REFERENCES app_notes(id) ON DELETE CASCADE
          );
          CREATE TABLE IF NOT EXISTS app_notes_meta(key TEXT PRIMARY KEY,value TEXT NOT NULL);
          CREATE TABLE IF NOT EXISTS app_calendar_events(
            id TEXT PRIMARY KEY, title TEXT NOT NULL DEFAULT '', details TEXT NOT NULL DEFAULT '',
            due_date TEXT NOT NULL DEFAULT '', notify_days_before INTEGER NOT NULL DEFAULT 2,
            reminders_per_day INTEGER NOT NULL DEFAULT 4, status TEXT NOT NULL DEFAULT 'pending',
            created_at TEXT NOT NULL, updated_at TEXT NOT NULL, completed_at TEXT
          );
          CREATE TABLE IF NOT EXISTS app_vales(
            id TEXT PRIMARY KEY,
            client_id TEXT NOT NULL UNIQUE,
            solicitante_nome TEXT NOT NULL DEFAULT '',
            solicitante_key TEXT NOT NULL DEFAULT '',
            amount REAL NOT NULL DEFAULT 0,
            request_date TEXT NOT NULL DEFAULT '',
            signature_png_base64 TEXT NOT NULL DEFAULT '',
            signature_format TEXT NOT NULL DEFAULT 'png',
            registered_by_user_id TEXT NOT NULL DEFAULT '',
            registered_by_username TEXT NOT NULL DEFAULT '',
            registered_by_name TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL DEFAULT 'sincronizado',
            source TEXT NOT NULL DEFAULT 'android_app',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
          );
          CREATE INDEX IF NOT EXISTS idx_app_notes_client ON app_notes(client);
          CREATE INDEX IF NOT EXISTS idx_app_notes_date ON app_notes(note_date);
          CREATE INDEX IF NOT EXISTS idx_app_note_items_note ON app_note_items(note_id);
          CREATE INDEX IF NOT EXISTS idx_app_note_submissions_note ON app_note_submissions(note_id);
          CREATE INDEX IF NOT EXISTS idx_app_calendar_due_status ON app_calendar_events(due_date,status);
          CREATE INDEX IF NOT EXISTS idx_app_vales_solicitante ON app_vales(solicitante_key);
          CREATE INDEX IF NOT EXISTS idx_app_vales_request_date ON app_vales(request_date);
          CREATE INDEX IF NOT EXISTS idx_app_vales_status ON app_vales(status);
        """)
        item_columns={r[1] for r in conn.execute("PRAGMA table_info(app_note_items)").fetchall()}
        note_columns={r[1] for r in conn.execute("PRAGMA table_info(app_notes)").fetchall()}
        if "status" not in note_columns:
            conn.execute("ALTER TABLE app_notes ADD COLUMN status TEXT NOT NULL DEFAULT 'pending'")
        if "completed_at" not in note_columns:
            conn.execute("ALTER TABLE app_notes ADD COLUMN completed_at TEXT")
        if "weight" not in item_columns:
            conn.execute("ALTER TABLE app_note_items ADD COLUMN weight REAL NOT NULL DEFAULT 0")
            conn.execute("UPDATE app_note_items SET weight=quantity WHERE weight=0")
        if "quantity_provided" not in item_columns:
            conn.execute("ALTER TABLE app_note_items ADD COLUMN quantity_provided INTEGER NOT NULL DEFAULT 1")
        if "price_provided" not in item_columns:
            conn.execute("ALTER TABLE app_note_items ADD COLUMN price_provided INTEGER NOT NULL DEFAULT 1")
        conn.execute("INSERT OR IGNORE INTO app_note_submissions(external_id,note_id,received_at) SELECT external_id,id,created_at FROM app_notes")
        merged=conn.execute("SELECT value FROM app_notes_meta WHERE key='merge_client_day_v1'").fetchone()
        if not merged:
            groups=conn.execute("""SELECT lower(trim(client)) client_key,note_date,group_concat(id) ids
                FROM app_notes WHERE trim(client)<>'' GROUP BY lower(trim(client)),note_date HAVING count(*)>1""").fetchall()
            for group in groups:
                ids=group["ids"].split(","); keep=ids[0]
                for duplicate in ids[1:]:
                    conn.execute("UPDATE app_note_items SET note_id=? WHERE note_id=?",(keep,duplicate))
                    conn.execute("UPDATE app_note_submissions SET note_id=? WHERE note_id=?",(keep,duplicate))
                    conn.execute("DELETE FROM app_notes WHERE id=?",(duplicate,))
                item_ids=conn.execute("SELECT id FROM app_note_items WHERE note_id=? ORDER BY position,id",(keep,)).fetchall()
                for position,item_row in enumerate(item_ids):
                    conn.execute("UPDATE app_note_items SET position=? WHERE id=?",(position,item_row["id"]))
                total=conn.execute("SELECT COALESCE(sum(weight*unit_price),0) total FROM app_note_items WHERE note_id=? AND price_provided=1",(keep,)).fetchone()["total"]
                conn.execute("UPDATE app_notes SET total=?,updated_at=? WHERE id=?",(round(float(total or 0),2),datetime.now().isoformat(timespec="seconds"),keep))
            conn.execute("INSERT OR REPLACE INTO app_notes_meta(key,value) VALUES('merge_client_day_v1','done')")
        conn.commit()
        return conn
    except Exception:
        conn.rollback()
        conn.close()
        raise

def _app_note_catalog_prices():
    """PreÃ§os atuais da aba Produtos do Monteiro, com aliases seguros de unidade."""
    main=get_db()
    try:
        rows=main.execute("SELECT name,suggested_price FROM paladar_products WHERE active=1").fetchall()
    finally:
        main.close()
    return app_note_catalog_from_rows(rows,_normalize_name)

def _app_note_dict(conn,row,catalog_prices=None):
    return app_note_dict_from_row(conn,row,catalog_prices,_normalize_name)

@app.post("/api/app-notes/mobile")
def create_app_note_mobile(body:dict,x_app_token:str=Header("",alias="x-app-token")):
    if not APP_NOTES_TOKEN or not hmac.compare_digest(x_app_token,APP_NOTES_TOKEN):
        raise HTTPException(401,"Aplicativo nÃ£o autorizado.")
    external_id=str(body.get("external_id") or "").strip()[:80]
    if not external_id: raise HTTPException(400,"Identificador da nota obrigatÃ³rio.")
    client,note_date,items,total=_clean_app_note(body)
    conn=get_app_notes_db()
    try:
        submission=conn.execute("SELECT note_id FROM app_note_submissions WHERE external_id=?",(external_id,)).fetchone()
        if submission:
            existing=conn.execute("SELECT * FROM app_notes WHERE id=?",(submission["note_id"],)).fetchone()
            result=_app_note_dict(conn,existing); return {"ok":True,"duplicate":True,"merged":False,"note":result}
        note_id=str(uuid.uuid4()); now=datetime.now().isoformat(timespec="seconds")
        same_day=None
        if client:
            same_day=conn.execute("SELECT * FROM app_notes WHERE lower(trim(client))=lower(trim(?)) AND note_date=? ORDER BY created_at LIMIT 1",(client,note_date)).fetchone()
        merged_note=bool(same_day)
        if same_day:
            note_id=same_day["id"]
            start_position=conn.execute("SELECT COALESCE(max(position),-1)+1 p FROM app_note_items WHERE note_id=?",(note_id,)).fetchone()["p"]
        else:
            start_position=0
            conn.execute("INSERT INTO app_notes(id,external_id,client,note_date,total,created_at,updated_at) VALUES(?,?,?,?,?,?,?)",
                         (note_id,external_id,client,note_date,total,now,now))
        for item in items:
            conn.execute("INSERT INTO app_note_items(id,note_id,product,quantity,quantity_provided,weight,unit,unit_price,price_provided,position) VALUES(?,?,?,?,?,?,?,?,?,?)",
                (str(uuid.uuid4()),note_id,item["product"],item["quantity"],int(item["quantity_provided"]),item["weight"],item["unit"],item["unit_price"],int(item["price_provided"]),start_position+item["position"]))
        conn.execute("INSERT INTO app_note_submissions(external_id,note_id,received_at) VALUES(?,?,?)",(external_id,note_id,now))
        accumulated=conn.execute("SELECT COALESCE(sum(weight*unit_price),0) total FROM app_note_items WHERE note_id=? AND price_provided=1",(note_id,)).fetchone()["total"]
        conn.execute("UPDATE app_notes SET total=?,updated_at=?,status='pending',completed_at=NULL WHERE id=?",(round(float(accumulated or 0),2),now,note_id))
        conn.commit(); row=conn.execute("SELECT * FROM app_notes WHERE id=?",(note_id,)).fetchone()
        result=_app_note_dict(conn,row)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return {"ok":True,"duplicate":False,"merged":merged_note,"note":result}

@app.get("/api/app-notes")
def list_app_notes(client:Optional[str]=None,month:Optional[str]=None,year:Optional[str]=None,status:Optional[str]=None,x_token:str=Header("")):
    require_auth(x_token); conn=get_app_notes_db(); where=[]; params=[]
    try:
        if client: where.append("client LIKE ?"); params.append("%"+client.strip()[:120]+"%")
        if month: where.append("substr(note_date,4,2)=?"); params.append(month.zfill(2)[:2])
        if year: where.append("substr(note_date,7,4)=?"); params.append(year[:4])
        if status in ("pending","completed"): where.append("status=?"); params.append(status)
        sql="SELECT * FROM app_notes"+(" WHERE "+" AND ".join(where) if where else "")+" ORDER BY created_at DESC LIMIT 1000"
        rows=conn.execute(sql,params).fetchall(); catalog_prices=_app_note_catalog_prices(); notes=[_app_note_dict(conn,r,catalog_prices) for r in rows]
        return {"notes":notes,"count":len(notes),"total":round(sum(float(n["total"] or 0) for n in notes),2)}
    finally:
        conn.close()

@app.put("/api/app-notes/{note_id}")
def update_app_note(note_id:str,body:dict,x_token:str=Header("")):
    require_admin_or_editor(x_token); client,note_date,items,total=_clean_app_note(body); conn=get_app_notes_db()
    try:
        if not conn.execute("SELECT 1 FROM app_notes WHERE id=?",(note_id,)).fetchone(): raise HTTPException(404,"Nota não encontrada.")
        new_status=str(body.get("status") or "").strip()
        status_sql=""
        status_params=[]
        if new_status in ("pending","completed"):
            status_sql=",status=?,completed_at=?"; status_params=[new_status,datetime.now().isoformat(timespec="seconds") if new_status=="completed" else None]
        conn.execute("UPDATE app_notes SET client=?,note_date=?,total=?,updated_at=?"+status_sql+" WHERE id=?",
                     [client,note_date,total,datetime.now().isoformat(timespec="seconds")]+status_params+[note_id])
        conn.execute("DELETE FROM app_note_items WHERE note_id=?",(note_id,))
        for item in items:
            conn.execute("INSERT INTO app_note_items(id,note_id,product,quantity,quantity_provided,weight,unit,unit_price,price_provided,position) VALUES(?,?,?,?,?,?,?,?,?,?)",
                (str(uuid.uuid4()),note_id,item["product"],item["quantity"],int(item["quantity_provided"]),item["weight"],item["unit"],item["unit_price"],int(item["price_provided"]),item["position"]))
        conn.commit(); row=conn.execute("SELECT * FROM app_notes WHERE id=?",(note_id,)).fetchone(); result=_app_note_dict(conn,row)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return {"ok":True,"note":result}

@app.put("/api/app-notes/{note_id}/status")
def update_app_note_status(note_id:str,body:dict,x_token:str=Header("")):
    require_admin_or_editor(x_token)
    status=str(body.get("status") or "").strip()
    if status not in ("pending","completed"): raise HTTPException(400,"Status invÃ¡lido.")
    now=datetime.now().isoformat(timespec="seconds"); conn=get_app_notes_db()
    cur=conn.execute("UPDATE app_notes SET status=?,completed_at=?,updated_at=? WHERE id=?",
        (status,now if status=="completed" else None,now,note_id))
    conn.commit(); conn.close()
    if not cur.rowcount: raise HTTPException(404,"Nota nÃ£o encontrada.")
    return {"ok":True,"status":status}

@app.delete("/api/app-notes/{note_id}")
def delete_app_note(note_id:str,x_token:str=Header("")):
    require_admin_or_editor(x_token); conn=get_app_notes_db(); cur=conn.execute("DELETE FROM app_notes WHERE id=?",(note_id,)); conn.commit(); conn.close()
    if not cur.rowcount: raise HTTPException(404,"Nota nÃ£o encontrada.")
    return {"ok":True}

# â”€â”€ SPA catch-all â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# IMPORTANTE: este endpoint TEM que ficar por Ãºltimo, pois @app.get("/{full_path:path}")
# captura QUALQUER GET. Se for declarado antes de outros @app.get("/api/..."),
# o catch-all intercepta as chamadas e devolve index.html (HTML), causando
# "Unexpected token '<'" no frontend ao tentar JSON.parse.
# â”€â”€ Calendario de notificacoes do aplicativo Android â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _clean_calendar_event(body:dict):
    title=str(body.get("title") or "").strip()[:160]
    details=str(body.get("details") or body.get("description") or "").strip()[:2000]
    due_date=str(body.get("due_date") or "").strip()[:10]
    if not title:
        raise HTTPException(400,"Informe o titulo da notificacao.")
    try:
        datetime.strptime(due_date,"%Y-%m-%d")
    except Exception:
        raise HTTPException(400,"Informe uma data valida.")
    try:
        notify_days_before=int(body.get("notify_days_before",2))
        reminders_per_day=int(body.get("reminders_per_day",4))
    except Exception:
        raise HTTPException(400,"Antecedencia ou repeticao invalida.")
    notify_days_before=max(0,min(30,notify_days_before))
    reminders_per_day=max(1,min(12,reminders_per_day))
    status=str(body.get("status") or "pending").strip()
    if status not in ("pending","completed"):
        status="pending"
    return title,details,due_date,notify_days_before,reminders_per_day,status

def _auto_complete_expired_calendar_events(conn):
    today=date.today().isoformat()
    now=datetime.now().isoformat(timespec="seconds")
    conn.execute("""UPDATE app_calendar_events
        SET status='completed', completed_at=?, updated_at=?
        WHERE status='pending' AND due_date<>'' AND date(due_date)<date(?)""",
        (now,now,today))
    conn.commit()

@app.get("/api/app-calendar")
def list_app_calendar(status:Optional[str]=None,month:Optional[str]=None,year:Optional[str]=None,x_token:str=Header("")):
    require_monteiro_calendar(x_token)
    conn=get_app_notes_db()
    try:
        _auto_complete_expired_calendar_events(conn); where=[]; params=[]
        if status in ("pending","completed"): where.append("status=?"); params.append(status)
        if month: where.append("substr(due_date,6,2)=?"); params.append(month.zfill(2)[:2])
        if year: where.append("substr(due_date,1,4)=?"); params.append(year[:4])
        sql="SELECT * FROM app_calendar_events"+(" WHERE "+" AND ".join(where) if where else "")+" ORDER BY due_date ASC, created_at DESC LIMIT 1000"
        rows=conn.execute(sql,params).fetchall(); events=[_calendar_event_dict(r) for r in rows]
        return {"events":events,"count":len(events),"pending":sum(1 for e in events if e.get("status")=="pending")}
    finally:
        conn.close()

@app.post("/api/app-calendar")
def create_app_calendar_event(body:dict,x_token:str=Header("")):
    require_monteiro_calendar(x_token)
    title,details,due_date,notify_days_before,reminders_per_day,status=_clean_calendar_event(body)
    recurring=bool(body.get("recurring"))
    try:
        repeat_months=int(body.get("repeat_months",1))
    except Exception:
        repeat_months=1
    repeat_months=max(1,min(60,repeat_months if recurring else 1))
    now=datetime.now().isoformat(timespec="seconds")
    conn=get_app_notes_db()
    try:
        event_ids=[]
        series_id=str(uuid.uuid4()) if repeat_months>1 else ""
        for offset in range(repeat_months):
            event_id=str(uuid.uuid4()); event_ids.append(event_id)
            event_date=_add_months(due_date,offset)
            item_details=details
            if series_id:
                item_details=(details+"\n\n" if details else "")+"Recorrencia: "+str(offset+1)+"/"+str(repeat_months)
            conn.execute("""INSERT INTO app_calendar_events
                (id,title,details,due_date,notify_days_before,reminders_per_day,status,created_at,updated_at,completed_at)
                VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (event_id,title,item_details,event_date,notify_days_before,reminders_per_day,status,now,now,now if status=="completed" else None))
        conn.commit()
        rows=conn.execute("SELECT * FROM app_calendar_events WHERE id IN ({}) ORDER BY due_date ASC".format(",".join("?" for _ in event_ids)),event_ids).fetchall()
        events=[_calendar_event_dict(r) for r in rows]
        return {"ok":True,"event":events[0] if events else None,"events":events,"created":len(events)}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

@app.put("/api/app-calendar/{event_id}")
def update_app_calendar_event(event_id:str,body:dict,x_token:str=Header("")):
    require_monteiro_calendar(x_token)
    title,details,due_date,notify_days_before,reminders_per_day,status=_clean_calendar_event(body)
    now=datetime.now().isoformat(timespec="seconds"); completed_at=now if status=="completed" else None
    conn=get_app_notes_db()
    try:
        cur=conn.execute("""UPDATE app_calendar_events SET title=?,details=?,due_date=?,notify_days_before=?,
            reminders_per_day=?,status=?,completed_at=?,updated_at=? WHERE id=?""",
            (title,details,due_date,notify_days_before,reminders_per_day,status,completed_at,now,event_id))
        conn.commit()
        if not cur.rowcount:
            raise HTTPException(404,"Notificacao nao encontrada.")
        row=conn.execute("SELECT * FROM app_calendar_events WHERE id=?",(event_id,)).fetchone(); result=_calendar_event_dict(row)
        return {"ok":True,"event":result}
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        conn.close()

@app.put("/api/app-calendar/{event_id}/status")
def update_app_calendar_status(event_id:str,body:dict,x_token:str=Header("")):
    require_monteiro_calendar(x_token)
    status=str(body.get("status") or "").strip()
    if status not in ("pending","completed"): raise HTTPException(400,"Status invalido.")
    now=datetime.now().isoformat(timespec="seconds"); conn=get_app_notes_db()
    try:
        cur=conn.execute("UPDATE app_calendar_events SET status=?,completed_at=?,updated_at=? WHERE id=?",
            (status,now if status=="completed" else None,now,event_id))
        conn.commit()
        if not cur.rowcount: raise HTTPException(404,"Notificacao nao encontrada.")
        return {"ok":True,"status":status}
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        conn.close()

@app.delete("/api/app-calendar/{event_id}")
def delete_app_calendar_event(event_id:str,x_token:str=Header("")):
    require_monteiro_calendar(x_token)
    conn=get_app_notes_db()
    try:
        cur=conn.execute("DELETE FROM app_calendar_events WHERE id=?",(event_id,)); conn.commit()
        if not cur.rowcount: raise HTTPException(404,"Notificacao nao encontrada.")
        return {"ok":True}
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        conn.close()

@app.get("/api/app-calendar/mobile")
def list_app_calendar_mobile(x_app_token:str=Header("",alias="x-app-token")):
    if not APP_CALENDAR_TOKEN or not hmac.compare_digest(x_app_token,APP_CALENDAR_TOKEN):
        raise HTTPException(401,"Aplicativo nao autorizado.")
    conn=get_app_notes_db()
    try:
        _auto_complete_expired_calendar_events(conn)
        today=date.today().isoformat()
        rows=conn.execute("""SELECT * FROM app_calendar_events
            WHERE status='pending' AND date(due_date)>=date(?)
            ORDER BY due_date ASC, created_at DESC LIMIT 200""",(today,)).fetchall()
        events=[]
        for r in rows:
            event=_calendar_event_dict(r)
            if event.get("in_reminder_window"):
                events.append(event)
        return {"events":events,"count":len(events),"checked_at":datetime.now().isoformat(timespec="seconds")}
    finally:
        conn.close()

@app.get("/{full_path:path}")
def serve_spa(full_path:str):
    # â”€â”€ SEGURANÃ‡A: bloqueio de path traversal â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Sem essas guardas, GET /../app.py ou GET /../bm_monteiro.db vazam o
    # cÃ³digo fonte e o banco inteiro. FastAPI decodifica %2e%2e antes de
    # entregar full_path, entÃ£o a checagem cobre ambas as formas.
    if full_path:
        # Rejeita absolutos e qualquer segmento ".." na URL
        if full_path.startswith("/") or full_path.startswith("\\"):
            return JSONResponse({"error":"Caminho invÃ¡lido."},404)
        parts=full_path.replace("\\","/").split("/")
        if ".." in parts or "" in parts:
            return JSONResponse({"error":"Caminho invÃ¡lido."},404)
    for base in [STATIC_DIR,STATIC_DIR/"client"]:
        if not full_path:
            continue
        try:
            base_resolved=base.resolve()
            f_resolved=(base/full_path).resolve()
            # Defesa em profundidade: confirma que o caminho resolvido fica
            # estritamente dentro de base (cobre symlinks e edge-cases).
            if not f_resolved.is_relative_to(base_resolved):
                continue
        except Exception:
            continue
        if f_resolved.exists() and f_resolved.is_file():
            headers={"Cache-Control":"no-store, no-cache, must-revalidate"} if f_resolved.suffix.lower() in (".html",".htm") else None
            media_type="text/html; charset=utf-8" if f_resolved.suffix.lower() in (".html",".htm") else None
            return FileResponse(str(f_resolved),headers=headers,media_type=media_type)
    idx=find_index()
    if idx: return FileResponse(str(idx),headers={"Cache-Control":"no-store, no-cache, must-revalidate"},media_type="text/html; charset=utf-8")
    return JSONResponse({"error":"Frontend nÃ£o compilado."},404)

if __name__=="__main__":
    # ProduÃ§Ã£o: app roda em servidor (Hostinger VPS), iniciado pelo systemd
    # ("systemctl restart menina"). Deploy via ATUALIZAR.bat (scp + restart).
    for company in COMPANY_DBS:
        init_db(company)
    port=int(os.environ.get("PORT",8765))
    uvicorn.run(app,host="0.0.0.0",port=port,log_level="warning")
